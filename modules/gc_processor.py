"""
modules/gc_processor.py
Core GC data processing engine.
Loaded by app.py — do not run directly.
"""

import os, re, zipfile, xml.etree.ElementTree as ET
from datetime import datetime
import numpy as np
import pandas as pd
import yaml


def _safe_file_token(value, default='GC'):
    text = str(value or '').strip()
    if not text:
        text = default
    text = re.sub(r'[^A-Za-z0-9_.-]+', '_', text).strip('_.-')
    return text or default


def _output_file_prefix(metadata):
    explicit = str((metadata or {}).get('output_prefix') or '').strip()
    if explicit:
        return _safe_file_token(explicit)
    date_token = str((metadata or {}).get('output_date') or '').strip()
    if not date_token:
        date_token = datetime.now().strftime('%Y%m%d')
    catalyst_token = _safe_file_token((metadata or {}).get('catalyst_id'), 'Unknown')
    return f'{_safe_file_token(date_token)}_{catalyst_token}'


_DEFAULT_UNKNOWN_AREA_RESPONSE_FACTORS = {
    5: 0.0018326957690202206,
    6: 0.0007242622194262482,
}


_COELUTED_SPLIT_RULES = {
    'cis-2-butene/butane': (
        {'label': 'c2C4H8', 'cn': 4, 'det': 'FID', 'fraction': 0.5},
        {'label': 'nC4H10', 'cn': 4, 'det': 'FID', 'fraction': 0.5},
    ),
}


_DEFAULT_SELECTIVITY_GROUP_ORDER = [
    'CO', 'CH4', 'C2-C4 Paraffins', 'C2-C4 Olefins', 'C5+',
    'Methanol', 'CO2', 'Other C products'
]


# ─────────────────────────────────────────────────────────────────────────────
# LOAD REACTION CONFIG
# ─────────────────────────────────────────────────────────────────────────────

def load_reaction_config(config_path):
    with open(config_path, 'r') as f:
        return yaml.safe_load(f)

def list_reaction_configs(config_dir):
    configs = []
    for fname in sorted(os.listdir(config_dir)):
        if fname.endswith('.yaml') and fname != 'custom_template.yaml':
            path = os.path.join(config_dir, fname)
            with open(path, 'r') as f:
                cfg = yaml.safe_load(f)
            configs.append({
                'file':        fname,
                'name':        cfg.get('name', fname),
                'description': cfg.get('description', ''),
                'reactant':    cfg.get('reactant', ''),
                'inlet_species': cfg.get('inlet_species', []),
            })
    return configs


# ─────────────────────────────────────────────────────────────────────────────
# XLSX PARSER (direct XML — avoids openpyxl styling bug)
# ─────────────────────────────────────────────────────────────────────────────

def col_to_idx(letters):
    r = 0
    for ch in letters.upper():
        r = r * 26 + (ord(ch) - ord('A') + 1)
    return r - 1

_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_RELS_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_PKG_RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'


def _xlsx_shared_strings(z):
    if 'xl/sharedStrings.xml' not in z.namelist():
        return []
    with z.open('xl/sharedStrings.xml') as f:
        tree = ET.parse(f)
    ns = {'x': _XLSX_NS}
    out = []
    for si in tree.findall('x:si', ns):
        texts = [t.text or '' for t in si.findall('.//x:t', ns)]
        out.append(''.join(texts))
    return out


def _xlsx_worksheet_paths(z):
    names = set(z.namelist())
    if 'xl/workbook.xml' in names and 'xl/_rels/workbook.xml.rels' in names:
        with z.open('xl/_rels/workbook.xml.rels') as f:
            rels_tree = ET.parse(f)
        rels = {}
        for rel in rels_tree.getroot().findall(f'{{{_PKG_RELS_NS}}}Relationship'):
            rels[rel.get('Id')] = rel.get('Target')

        with z.open('xl/workbook.xml') as f:
            workbook_tree = ET.parse(f)
        paths = []
        for sheet in workbook_tree.findall(f'.//{{{_XLSX_NS}}}sheet'):
            rid = sheet.get(f'{{{_RELS_NS}}}id')
            target = rels.get(rid, '')
            if not target:
                continue
            target = target.lstrip('/')
            if not target.startswith('xl/'):
                target = 'xl/' + target
            if target in names:
                paths.append(target)
        if paths:
            return paths

    if 'xl/worksheets/sheet.xml' in names:
        return ['xl/worksheets/sheet.xml']
    return sorted(n for n in names
                  if n.startswith('xl/worksheets/') and n.endswith('.xml'))


def _parse_gc_sheet(stree, strings):
    rows = stree.findall(f'.//{{{_XLSX_NS}}}row')

    def cell_val(c):
        t = c.get('t', '')
        if t == 'inlineStr':
            texts = [x.text or '' for x in c.findall(f'.//{{{_XLSX_NS}}}t')]
            return ''.join(texts)
        v = c.find(f'{{{_XLSX_NS}}}v')
        if v is None:
            return None
        if t == 's':
            try:
                return strings[int(v.text)]
            except (TypeError, ValueError, IndexError):
                return None
        return v.text

    def row_dict(row):
        d = {}
        refs = {}
        for c in row.findall(f'{{{_XLSX_NS}}}c'):
            ref = c.get('r', '')
            letters = ''.join(ch for ch in ref if ch.isalpha())
            if not letters:
                continue
            val = cell_val(c)
            if val is not None:
                idx = col_to_idx(letters)
                d[idx] = val
                refs[idx] = ref
        return d, refs

    row_data = [row_dict(row) for row in rows]
    row_dicts = [d for d, _ in row_data]
    row_refs = [refs for _, refs in row_data]

    sequence_name = 'Unknown'
    for d in row_dicts[:8]:
        for idx, val in d.items():
            if str(val).strip().lower() == 'sequence name':
                sequence_name = d.get(idx + 1) or d.get(idx + 2) or sequence_name
                break

    species_row_idx = measure_row_idx = None
    for idx in range(min(20, max(0, len(row_dicts) - 1))):
        n_amount = sum(
            str(v).strip().lower() == 'amount'
            for v in row_dicts[idx + 1].values())
        if n_amount:
            species_row_idx = idx
            measure_row_idx = idx + 1
            break
    if species_row_idx is None:
        return {'sequence_name': sequence_name, 'injections': []}

    species_row = row_dicts[species_row_idx]
    measure_row = row_dicts[measure_row_idx]

    amount_cols = {}
    area_cols = {}
    sorted_species = sorted(
        (idx, str(name).strip()) for idx, name in species_row.items()
        if str(name).strip())
    for col_idx, mtype in measure_row.items():
        mtype_norm = str(mtype).strip().lower()
        if mtype_norm not in {'amount', 'area', 'peak area'}:
            continue
        best, best_dist = None, 999
        for sc_idx, sc_name in sorted_species:
            dist = abs(col_idx - sc_idx)
            if dist <= 3 and dist < best_dist:
                best, best_dist = sc_name, dist
        if best:
            if mtype_norm == 'amount':
                amount_cols[col_idx] = best
            else:
                area_cols[col_idx] = best

    injections = []
    for row_idx, d in enumerate(row_dicts[measure_row_idx + 1:], start=measure_row_idx + 1):
        if not d:
            continue
        label = str(d.get(0, '')).strip()
        if not label:
            continue
        amounts = {}
        areas = {}
        amount_refs = {}
        area_refs = {}
        for cidx, sp in amount_cols.items():
            val = d.get(cidx)
            if val in (None, ''):
                continue
            try:
                amounts[sp] = float(val)
                if cidx in row_refs[row_idx]:
                    amount_refs[sp] = row_refs[row_idx][cidx]
            except (TypeError, ValueError):
                pass
        for cidx, sp in area_cols.items():
            val = d.get(cidx)
            if val in (None, ''):
                continue
            try:
                areas[sp] = float(val)
                if cidx in row_refs[row_idx]:
                    area_refs[sp] = row_refs[row_idx][cidx]
            except (TypeError, ValueError):
                pass
        m = re.search(r'(\d+)\s*$', label)
        injections.append({
            'label':     label,
            'inj_num':   int(m.group(1)) if m else None,
            'is_bypass': 'bypass' in label.lower(),
            'amounts':   amounts,
            'areas':     areas,
            'source_refs': {
                'label': row_refs[row_idx].get(0),
                'amounts': amount_refs,
                'areas': area_refs,
            },
        })

    return {'sequence_name': sequence_name, 'injections': injections}


def parse_xlsx(filepath):
    with zipfile.ZipFile(filepath) as z:
        strings = _xlsx_shared_strings(z)
        for sheet_idx, sheet_path in enumerate(_xlsx_worksheet_paths(z)):
            with z.open(sheet_path) as f:
                data = _parse_gc_sheet(ET.parse(f), strings)
            if data['injections']:
                data['worksheet'] = sheet_path
                data['worksheet_index'] = sheet_idx
                return data
    raise ValueError('No usable GC worksheet found in the XLSX file.')


# ─────────────────────────────────────────────────────────────────────────────
# MOLAR FLOW CALCULATIONS
# ─────────────────────────────────────────────────────────────────────────────

def find_ch4_tcd_key(species_config):
    for header, cfg in species_config.items():
        if cfg['label'] == 'CH4_TCD': return header
    return None

def find_ch4_fid_key(species_config):
    for header, cfg in species_config.items():
        if cfg['label'] == 'CH4': return header
    return None

def find_ar_key(species_config, is_label=None):
    for header, cfg in species_config.items():
        if cfg['label'] == 'Ar': return header
    return None


def _infer_carbon_number_from_name(name):
    text = str(name or '').strip().lower()
    m = re.search(r'\bc\s*([0-9]+)\b', text)
    if not m:
        m = re.search(r'c([0-9]+)', text)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return 0
    if 'pent' in text:
        return 5
    if 'hex' in text:
        return 6
    return 0


def _response_reference_header(cn, species_config):
    if int(cn or 0) == 5:
        preferred = ('Pentane', 'nC5H12')
    elif int(cn or 0) == 6:
        preferred = ('Hexane', 'C6H14')
    else:
        return None
    for target in preferred:
        for header, cfg in species_config.items():
            if header == target or cfg.get('label') == target:
                return header
    return None


def _augment_species_config_for_area_fallback(data, species_config):
    added = []
    for inj in data.get('injections', []):
        for header in (inj.get('areas') or {}):
            if header in species_config:
                continue
            cn = _infer_carbon_number_from_name(header)
            if cn in (5, 6):
                species_config[header] = {
                    'label': header,
                    'cn': cn,
                    'det': 'FID',
                    'area_fallback': True,
                }
                added.append(header)
    return sorted(set(added))


def _global_response_factor(data, reference_header):
    vals = []
    for inj in data.get('injections', []):
        amount = (inj.get('amounts') or {}).get(reference_header)
        area = (inj.get('areas') or {}).get(reference_header)
        if amount is not None and area and area > 0:
            vals.append(float(amount) / float(area))
    return float(np.mean(vals)) if vals else None


def _unknown_response_factor_key(cn):
    return f'c{int(cn or 0)}_unknown_response_factor'


def _unknown_area_response_factor(header, cn, metadata=None):
    if 'unknown' not in str(header or '').strip().lower():
        return None
    cn = int(cn or 0)
    default = _DEFAULT_UNKNOWN_AREA_RESPONSE_FACTORS.get(cn)
    if default is None:
        return None
    return _metadata_float(metadata or {}, _unknown_response_factor_key(cn), default)


def apply_area_response_fallbacks(data, species_config, metadata=None):
    """Convert area-only C5/C6 peaks to amounts with pentane/hexane response factors."""
    _augment_species_config_for_area_fallback(data, species_config)
    refs = {
        5: _response_reference_header(5, species_config),
        6: _response_reference_header(6, species_config),
    }
    global_rf = {
        cn: _global_response_factor(data, ref) if ref else None
        for cn, ref in refs.items()
    }
    converted = []
    skipped = []
    for inj in data.get('injections', []):
        amounts = inj.setdefault('amounts', {})
        areas = inj.get('areas') or {}
        source_refs = inj.setdefault('source_refs', {})
        area_refs = source_refs.get('areas') or {}
        amount_refs = source_refs.setdefault('amounts', {})
        derived_refs = source_refs.setdefault('area_derived', {})
        for header, area in areas.items():
            if header in amounts and amounts.get(header) not in (None, ''):
                continue
            cfg = species_config.get(header)
            cn = int((cfg or {}).get('cn') or _infer_carbon_number_from_name(header) or 0)
            if cn not in (5, 6) or not area or area <= 0:
                continue
            fixed_rf = _unknown_area_response_factor(header, cn, metadata)
            ref_header = refs.get(cn)
            ref_amount = None
            ref_area = None
            if fixed_rf:
                rf = fixed_rf
                source = 'settings/default factor'
                ref_header = f'C{cn} unknown RF setting'
            else:
                if not ref_header:
                    skipped.append((header, cn, 'missing reference species'))
                    continue
                ref_amount = amounts.get(ref_header)
                ref_area = areas.get(ref_header)
                if ref_amount is not None and ref_area and ref_area > 0:
                    rf = float(ref_amount) / float(ref_area)
                    source = 'same injection'
                else:
                    rf = global_rf.get(cn)
                    source = 'run average'
            if not rf:
                skipped.append((header, cn, f'missing {ref_header} amount/area response factor'))
                continue
            amounts[header] = float(area) * float(rf)
            derived_refs[header] = {
                'area': area_refs.get(header),
                'reference_header': ref_header,
                'reference_amount': amount_refs.get(ref_header) if ref_header in amount_refs else None,
                'reference_area': area_refs.get(ref_header) if ref_header in area_refs else None,
                'response_factor': rf,
                'response_factor_key': _unknown_response_factor_key(cn) if fixed_rf else None,
                'response_source': source,
                'carbon_number': cn,
            }
            converted.append({
                'label': inj.get('label'),
                'species': header,
                'carbon_number': cn,
                'area': float(area),
                'reference': ref_header,
                'response_source': source,
                'amount': float(amounts[header]),
            })
    data['area_derived_amounts'] = converted
    data['area_derivation_skipped'] = [
        {'species': sp, 'carbon_number': cn, 'reason': reason}
        for sp, cn, reason in sorted(set(skipped))
    ]
    return converted


def _augment_species_config_for_coeluted_splits(data, species_config):
    added = []
    seen_headers = set()
    for inj in data.get('injections', []):
        seen_headers.update((inj.get('amounts') or {}).keys())
        seen_headers.update((inj.get('areas') or {}).keys())

    for header in sorted(seen_headers):
        rule = _COELUTED_SPLIT_RULES.get(str(header or '').strip().lower())
        if not rule:
            continue
        for idx, cfg in enumerate(rule, start=1):
            split_header = f'{header}__split_{idx}'
            if split_header in species_config:
                continue
            species_config[split_header] = {
                'label': cfg['label'],
                'cn': cfg['cn'],
                'det': cfg['det'],
                'source_header': header,
                'flow_fraction': float(cfg['fraction']),
                'split_source': header,
            }
            added.append(split_header)
    return added


def compute_flows(amounts, F_Ar_sccm, species_config, use_ch4_bridge):
    ar_key = find_ar_key(species_config)
    C_Ar = amounts.get(ar_key) if ar_key else None
    if not C_Ar or C_Ar == 0:
        return {}

    ch4_ratio = None
    if use_ch4_bridge:
        tcd_key = find_ch4_tcd_key(species_config)
        fid_key = find_ch4_fid_key(species_config)
        c_tcd = amounts.get(tcd_key) if tcd_key else None
        c_fid = amounts.get(fid_key) if fid_key else None
        if c_tcd and c_fid and c_tcd > 0 and c_fid > 0:
            ch4_ratio = c_tcd / c_fid

    flows = {}
    for sp_header, cfg in species_config.items():
        C_A = amounts.get(cfg.get('source_header') or sp_header)
        if C_A is None or C_A == 0: continue
        label = cfg['label']
        if cfg['det'] == 'TCD':
            flow = F_Ar_sccm * (C_A / C_Ar)
        else:
            use_bridge_for_species = cfg.get('fid_bridge', True) is not False
            if use_ch4_bridge and ch4_ratio is not None and use_bridge_for_species:
                flow = F_Ar_sccm * ch4_ratio * (C_A / C_Ar)
            elif not use_ch4_bridge or not use_bridge_for_species:
                flow = F_Ar_sccm * (C_A / C_Ar)
            else:
                continue
        flow *= float(cfg.get('flow_fraction') or 1.0)
        flows[label] = flows.get(label, 0.0) + flow
    return flows

def build_flow_table(data, F_Ar_sccm, species_config):
    ch4_tcd_key = find_ch4_tcd_key(species_config)
    ch4_fid_key = find_ch4_fid_key(species_config)
    has_bridge  = any(
        inj['amounts'].get(ch4_tcd_key) and inj['amounts'].get(ch4_fid_key)
        for inj in data['injections']
    )
    records = []
    for inj in data['injections']:
        flows = compute_flows(inj['amounts'], F_Ar_sccm, species_config, has_bridge)
        row = {'label': inj['label'], 'inj_num': inj['inj_num'],
               'is_bypass': inj['is_bypass'],
               'source_kind': inj.get('source_kind', 'main')}
        row.update(flows)
        records.append(row)
    return pd.DataFrame(records), has_bridge


def _mean_amounts(injections):
    buckets = {}
    for inj in injections:
        for key, val in inj.get('amounts', {}).items():
            if val is None:
                continue
            buckets.setdefault(key, []).append(float(val))
    return {
        key: float(np.mean(vals))
        for key, vals in buckets.items()
        if vals
    }


def _metadata_choice(metadata, key, default='auto'):
    raw = str((metadata or {}).get(key, default) or default).strip().lower()
    return raw or default


def _mark_source_kind(data, source_kind):
    for inj in (data or {}).get('injections', []):
        inj.setdefault('source_kind', source_kind)


def _bypass_like_score(inj, reaction_config, species_config):
    label = str(inj.get('label') or '').strip().lower()
    if 'bypass' in label:
        return 100.0

    inlet_labels = set(_inlet_labels(reaction_config))
    inlet_labels.update({'Ar', reaction_config.get('reactant')})
    product_count = 0
    inlet_count = 0
    for header, val in (inj.get('amounts') or {}).items():
        if val in (None, ''):
            continue
        cfg = species_config.get(header) or {}
        label_name = cfg.get('label') or header
        cn = int(cfg.get('cn') or 0)
        if label_name in inlet_labels:
            inlet_count += 1
        elif cn > 0:
            product_count += 1

    # Bypass rows usually contain inlet gases only. Reaction rows usually have
    # several product amounts, so penalize product-rich edge rows strongly.
    return inlet_count * 2.0 - product_count * 3.0


def _apply_same_file_bypass_selection(data, metadata, reaction_config, species_config):
    injections = list((data or {}).get('injections', []))
    mode = _metadata_choice(metadata, 'same_file_bypass_mode', 'auto')
    requested = _metadata_int(metadata or {}, 'same_file_bypass_rows')
    if requested is None:
        requested = _metadata_int(metadata or {}, 'bypass_points_used', 3)
    requested = max(0, int(requested or 0))

    for inj in injections:
        inj['is_bypass'] = False

    if mode in {'none', 'off', 'manual'} or not injections:
        return None

    label_indices = [
        idx for idx, inj in enumerate(injections)
        if 'bypass' in str(inj.get('label') or '').lower()
    ]
    selected = []
    detect_method = ''

    if mode in {'label', 'labels', 'labeled'}:
        selected = label_indices
        detect_method = 'label'
    elif mode in {'first', 'beginning', 'start'}:
        selected = list(range(min(requested, len(injections))))
        detect_method = 'first'
    elif mode in {'last', 'end'}:
        n = min(requested, len(injections))
        selected = list(range(len(injections) - n, len(injections))) if n else []
        detect_method = 'last'
    else:
        if label_indices:
            selected = label_indices
            detect_method = 'label'
        else:
            n = min(requested or 3, len(injections))
            first = list(range(n))
            last = list(range(len(injections) - n, len(injections))) if n else []
            first_score = sum(_bypass_like_score(injections[i], reaction_config, species_config) for i in first)
            last_score = sum(_bypass_like_score(injections[i], reaction_config, species_config) for i in last)
            if first_score > 0 or last_score > 0:
                selected = first if first_score >= last_score else last
                detect_method = 'auto_first' if selected == first else 'auto_last'

    selected = sorted(set(idx for idx in selected if 0 <= idx < len(injections)))
    for idx in selected:
        injections[idx]['is_bypass'] = True
        injections[idx]['bypass_source'] = 'same_file'

    metadata['same_file_bypass_mode'] = mode
    metadata['same_file_bypass_rows'] = requested
    metadata['same_file_bypass_detected'] = len(selected)
    metadata['same_file_bypass_detect_method'] = detect_method or 'none'

    if not selected:
        return None

    out = dict(data or {})
    out['injections'] = [injections[idx] for idx in selected]
    out['source'] = 'same_file'
    out['source_sheet_name'] = 'Raw Original'
    out['worksheet_index'] = (data or {}).get('worksheet_index', 0)
    return out


def _select_bypass_data(bypass_data, metadata):
    injections = list((bypass_data or {}).get('injections', []))
    omit = max(0, _metadata_int(metadata or {}, 'bypass_omit_initial', 0) or 0)
    requested = _metadata_int(metadata or {}, 'bypass_points_used')
    selected = injections[min(omit, len(injections)):]
    if requested and requested > 0:
        selected = selected[:min(requested, len(selected))]

    out = dict(bypass_data or {})
    out['injections'] = selected
    out['bypass_total_points'] = len(injections)
    out['bypass_omit_initial'] = omit
    out['bypass_points_requested'] = requested
    out['bypass_selected_points'] = len(selected)
    out['source'] = (bypass_data or {}).get('source')
    out['source_sheet_name'] = (bypass_data or {}).get('source_sheet_name')
    return out


def _inlet_labels(reaction_config):
    labels = []
    for item in reaction_config.get('inlet_species', []):
        if isinstance(item, dict) and item.get('label'):
            labels.append(item['label'])
        elif isinstance(item, str):
            labels.append(item)
    return labels


def infer_inlet_flows_from_bypass(bypass_data, F_Ar_sccm, reaction_config):
    """Infer inlet molar flows from a separate bypass GC workbook.

    The bypass workbook supplies inlet concentrations.  The entered Ar MFC
    flow remains the absolute normalisation anchor:

        F_i,in = F_Ar,in * C_i,bypass / C_Ar,bypass
    """
    species_config = reaction_config['species']
    inlet_labels = set(_inlet_labels(reaction_config))
    mean_amounts = _mean_amounts(bypass_data.get('injections', []))
    ar_key = find_ar_key(species_config)
    C_Ar = mean_amounts.get(ar_key) if ar_key else None
    if not C_Ar or C_Ar <= 0:
        return {}, 'Bypass file did not contain a usable Ar amount.'

    inferred = {'Ar': float(F_Ar_sccm)}
    for header, cfg in species_config.items():
        label = cfg.get('label')
        if not label or label == 'Ar' or label not in inlet_labels:
            continue
        C_i = mean_amounts.get(header)
        if C_i is not None and C_i > 0:
            inferred[label] = float(F_Ar_sccm) * float(C_i) / float(C_Ar)
    return inferred, None


def get_cn(label, species_config):
    for cfg in species_config.values():
        if cfg['label'] == label: return cfg['cn']
    return 0


def _reaction_type(metadata=None, reaction_config=None):
    raw = ''
    if reaction_config:
        raw = reaction_config.get('type') or reaction_config.get('name') or ''
    if not raw and metadata:
        raw = metadata.get('reaction_type') or metadata.get('reaction') or ''
    return re.sub(r'[^a-z0-9]+', '_', str(raw).strip().lower()).strip('_')


def _conversion_basis_candidates(reactant_label, reaction_config=None):
    candidates = []
    if reaction_config:
        for key in ('conversion_basis', 'reactant_basis'):
            val = reaction_config.get(key)
            if val:
                candidates.append(str(val))
        fallbacks = reaction_config.get('conversion_fallbacks') or []
        if isinstance(fallbacks, str):
            fallbacks = [x.strip() for x in fallbacks.split(',') if x.strip()]
        candidates.extend(str(x) for x in fallbacks if x)
    candidates.append(reactant_label)
    seen = set()
    out = []
    for item in candidates:
        if item and item not in seen:
            seen.add(item)
            out.append(item)
    return out


def _configured_product_labels(reaction_config=None):
    if not reaction_config:
        return None
    raw = (
        reaction_config.get('product_labels') or
        reaction_config.get('selectivity_products') or
        reaction_config.get('carbon_products')
    )
    if not raw:
        return None
    if isinstance(raw, str):
        labels = [x.strip() for x in raw.split(',') if x.strip()]
    else:
        labels = [str(x) for x in raw if x]
    return set(labels) if labels else None


def _select_existing_column(df, candidates):
    for col in candidates:
        if col in df.columns and _has_numeric_data(df[col]):
            return col
    for col in candidates:
        if col in df.columns:
            return col
    return candidates[0] if candidates else None


def _selectivity_group_order(metadata=None):
    raw = (metadata or {}).get('selectivity_group_order')
    if isinstance(raw, str):
        order = [x.strip() for x in raw.split(',') if x.strip()]
    elif raw:
        order = [str(x) for x in raw if x]
    else:
        order = []
    for group in _DEFAULT_SELECTIVITY_GROUP_ORDER:
        if group not in order:
            order.append(group)
    return order

def _has_numeric_data(series):
    return pd.to_numeric(series, errors='coerce').notna().any()


def _is_duplicate_tcd_product(label, df):
    if not isinstance(label, str) or not label.endswith('_TCD'):
        return False
    fid_label = label[:-4]
    return fid_label in df.columns and _has_numeric_data(df[fid_label])


def calculate_results(df, reactant_label, F_reactant_inlet, species_config,
                      reaction_config=None):
    df = df.copy()
    basis_candidates = _conversion_basis_candidates(reactant_label, reaction_config)
    reactant_basis_label = _select_existing_column(df, basis_candidates)
    if reactant_basis_label in df.columns and F_reactant_inlet:
        df['conversion'] = (F_reactant_inlet - df[reactant_basis_label]) / F_reactant_inlet
    else:
        df['conversion'] = np.nan

    meta_cols = {
        'label', 'inj_num', 'is_bypass', 'source_kind', 'analysis_include',
        'row_status', 'conversion', 'time_on_stream_h'
    }
    reactant_exclusions = set(basis_candidates)
    reactant_exclusions.add(reactant_label)
    allowed_products = _configured_product_labels(reaction_config)
    carbon_cols = [
        c for c in df.columns
        if c not in meta_cols
        and get_cn(c, species_config) > 0
        and c not in reactant_exclusions
        and (allowed_products is None or c in allowed_products)
        and not _is_duplicate_tcd_product(c, df)
    ]

    if carbon_cols:
        product_C  = sum(get_cn(c, species_config) * df[c].fillna(0) for c in carbon_cols)
        reactant_C = get_cn(reactant_basis_label, species_config) * df[reactant_basis_label].fillna(0) \
                     if reactant_basis_label in df.columns else 0
        total_C_out = product_C + reactant_C
        with np.errstate(divide='ignore', invalid='ignore'):
            df_sel = pd.DataFrame({
                f'S_{c}': np.where(product_C > 0,
                    get_cn(c, species_config) * df[c].fillna(0) / product_C, np.nan)
                for c in carbon_cols
            }, index=df.index)
    else:
        total_C_out = pd.Series(np.nan, index=df.index)
        df_sel = pd.DataFrame(index=df.index)

    return df, df_sel, total_C_out, carbon_cols


# ─────────────────────────────────────────────────────────────────────────────
# PLOTTING
# ─────────────────────────────────────────────────────────────────────────────

def _plot_x_values(rxn):
    temps = []
    for label in rxn['label'].astype(str):
        matches = re.findall(r'(-?\d+(?:\.\d+)?)\s*(?:°\s*)?C\b', label, flags=re.I)
        vals = [float(v) for v in matches if -50 <= float(v) <= 700]
        temps.append(vals[-1] if vals else np.nan)
    if np.isfinite(temps).sum() >= 2 and len(set(v for v in temps if np.isfinite(v))) >= 2:
        return np.array(temps, dtype=float), 'Temperature (°C)'
    return pd.to_numeric(rxn['inj_num'], errors='coerce').to_numpy(dtype=float), 'Injection number'


def _analysis_mask(df):
    if 'analysis_include' in df.columns:
        return df['analysis_include'].fillna(False).astype(bool)
    return ~df['is_bypass']


def _analysis_reaction_rows(df):
    return df[(~df['is_bypass']) & _analysis_mask(df)].copy()


def _metadata_float(metadata, key, default=None):
    raw = metadata.get(key, default)
    if raw in (None, ''):
        return default
    try:
        return float(raw)
    except (TypeError, ValueError):
        text = str(raw)
        m = re.search(r'(-?\d+(?:\.\d+)?)', text)
        if m:
            return float(m.group(1))
    return default


def _metadata_int(metadata, key, default=None):
    val = _metadata_float(metadata, key, default=None)
    if val is None:
        return default
    return int(round(val))


def _infer_run_duration_h(metadata):
    duration = _metadata_float(metadata, 'run_duration_h')
    if duration and duration > 0:
        return duration
    notes = str(metadata.get('notes', '') or '')
    m = re.search(r'(\d+(?:\.\d+)?)\s*(?:h|hr|hrs|hour|hours)\b', notes, flags=re.I)
    if m:
        return float(m.group(1))
    return None


def _add_time_on_stream_column(df, metadata):
    df = df.copy()
    duration = _infer_run_duration_h(metadata)
    interval_min = _metadata_float(metadata, 'injection_interval_min')
    df['time_on_stream_h'] = np.nan
    df['analysis_include'] = False
    df['row_status'] = np.where(
        df['is_bypass'],
        'Bypass / inlet normalization',
        'Reaction included'
    )

    rxn_idx_all = list(df.index[~df['is_bypass']])
    rxn_idx = list(rxn_idx_all)
    rejected = max(0, _metadata_int(metadata, 'rejected_initial_injections', 0) or 0)
    if rejected:
        rxn_idx = rxn_idx[min(rejected, len(rxn_idx)):]
    requested = _metadata_int(metadata, 'registered_reaction_injections')
    if requested and requested > 0:
        rxn_idx = rxn_idx[:min(requested, len(rxn_idx))]

    included = set(rxn_idx)
    df.loc[list(included), 'analysis_include'] = True
    for pos, idx in enumerate(rxn_idx_all):
        if idx in included:
            continue
        if pos < rejected:
            df.loc[idx, 'row_status'] = 'Excluded: initial reaction point'
        else:
            df.loc[idx, 'row_status'] = 'Excluded: beyond requested reaction count'

    if not rxn_idx:
        return df, 0

    if interval_min and interval_min > 0:
        times = np.arange(len(rxn_idx), dtype=float) * float(interval_min) / 60.0
    elif len(rxn_idx) == 1:
        times = np.array([0.0])
    elif duration and duration > 0:
        times = np.linspace(0.0, float(duration), len(rxn_idx))
    else:
        return df, len(rxn_idx)
    df.loc[rxn_idx, 'time_on_stream_h'] = times
    return df, len(rxn_idx)


def _selectivity_groups(df_sel, species_config):
    groups = {
        'CO': [],
        'CH4': [],
        'C2-C4 Olefins': [],
        'C2-C4 Paraffins': [],
        'C5+': [],
        'Methanol': [],
        'CO2': [],
        'Other C products': [],
    }
    olefins = {
        'C2H4', 'C3H6', 'C4H6', 't2C4H8', '1C4H8', 'c2C4H8',
        'iC4H8', 'VCH'
    }
    paraffins = {'C2H6', 'C3H8', 'nC4H10', 'iC4H10'}
    cn_lookup = {cfg['label']: cfg.get('cn', 0) for cfg in species_config.values()}
    for col in df_sel.columns:
        label = col.replace('S_', '')
        if label == 'CO':
            groups['CO'].append(col)
        elif label == 'CH4':
            groups['CH4'].append(col)
        elif label in {'Methanol', 'MeOH', 'CH3OH'}:
            groups['Methanol'].append(col)
        elif label == 'CO2':
            groups['CO2'].append(col)
        elif label in olefins:
            groups['C2-C4 Olefins'].append(col)
        elif label in paraffins:
            groups['C2-C4 Paraffins'].append(col)
        elif cn_lookup.get(label, 0) >= 5:
            groups['C5+'].append(col)
        else:
            groups['Other C products'].append(col)
    return {k: v for k, v in groups.items() if v}


def _nice_upper(value, floor=5.0):
    if not np.isfinite(value) or value <= floor:
        return floor
    exp = 10 ** np.floor(np.log10(value))
    for mult in (1, 2, 5, 10):
        upper = mult * exp
        if value <= upper:
            return float(upper)
    return float(10 * exp)


def _legend_label_width(draw, label, font, sub_font):
    width = 0
    for ch in str(label):
        use_font = sub_font if ch.isdigit() else font
        bbox = draw.textbbox((0, 0), ch, font=use_font)
        width += bbox[2] - bbox[0]
    return width


def _draw_legend_label(draw, x, y, label, font, sub_font, fill=(0, 0, 0)):
    cursor = int(x)
    sub_offset = max(2, int(getattr(font, 'size', 14) * 0.35))
    for ch in str(label):
        is_sub = ch.isdigit()
        use_font = sub_font if is_sub else font
        y_pos = int(y) + (sub_offset if is_sub else 0)
        draw.text((cursor, y_pos), ch, fill=fill, font=use_font)
        bbox = draw.textbbox((0, 0), ch, font=use_font)
        cursor += bbox[2] - bbox[0]


def _draw_time_on_stream_plot(df, df_sel, total_C_out, C_in_flow,
                              reactant_label, metadata, species_config,
                              output_dir):
    from PIL import Image, ImageDraw, ImageFont

    rxn = df[
        (~df['is_bypass']) &
        _analysis_mask(df) &
        pd.to_numeric(df.get('time_on_stream_h'), errors='coerce').notna()
    ].copy()
    if rxn.empty:
        return _draw_stacked_selectivity_plot(
            df, df_sel, total_C_out, C_in_flow,
            reactant_label, metadata, species_config, output_dir)

    width, height = 1250, 900
    margin = {'l': 110, 'r': 108, 't': 88, 'b': 175}
    x0, y0 = margin['l'], margin['t']
    x1, y1 = width - margin['r'], height - margin['b']
    plot_w, plot_h = x1 - x0, y1 - y0

    img = Image.new('RGB', (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    def load_font(size, bold=False):
        names = ['arialbd.ttf', 'arial.ttf'] if bold else ['arial.ttf', 'segoeui.ttf']
        windir = os.environ.get('WINDIR', r'C:\Windows')
        for name in names:
            try:
                return ImageFont.truetype(os.path.join(windir, 'Fonts', name), size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    font = load_font(16)
    small_font = load_font(14)
    legend_sub_font = load_font(11)
    axis_font = load_font(28)
    title_font = load_font(34)

    def txt(x, y, text, fill=(0, 0, 0), anchor=None, font_obj=None):
        kwargs = {'fill': fill, 'font': font_obj or font}
        if anchor:
            kwargs['anchor'] = anchor
        draw.text((int(x), int(y)), str(text), **kwargs)

    def text_w(text, font_obj=None):
        bbox = draw.textbbox((0, 0), str(text), font=font_obj or font)
        return bbox[2] - bbox[0]

    def rotated_txt(x, y, text, angle, font_obj=None):
        font_use = font_obj or font
        bbox = draw.textbbox((0, 0), str(text), font=font_use)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        layer = Image.new('RGBA', (tw + 12, th + 12), (255, 255, 255, 0))
        layer_draw = ImageDraw.Draw(layer)
        layer_draw.text((6, 6), str(text), fill=(0, 0, 0, 255), font=font_use)
        rotated = layer.rotate(angle, expand=True)
        img.paste(rotated, (int(x - rotated.width / 2), int(y - rotated.height / 2)), rotated)

    x_vals = pd.to_numeric(rxn['time_on_stream_h'], errors='coerce').to_numpy(dtype=float)
    duration = _infer_run_duration_h(metadata)
    x_min = 0.0
    x_max = max(float(duration or 0), float(np.nanmax(x_vals)), 1.0)

    conv_vals = pd.to_numeric(rxn.get('conversion', pd.Series(index=rxn.index, dtype=float)),
                              errors='coerce').to_numpy() * 100.0
    conv_upper = _nice_upper(np.nanmax(conv_vals) if np.isfinite(conv_vals).any() else 5.0, floor=5.0)

    def xp(v):
        return x0 + (float(v) - x_min) / (x_max - x_min) * plot_w

    def y_left(v):
        v = max(0.0, min(float(conv_upper), float(v)))
        return y1 - (v / conv_upper) * plot_h

    def y_right(v):
        v = max(0.0, min(100.0, float(v)))
        return y1 - (v / 100.0) * plot_h

    # Axes and grid.
    draw.line((x0, y1, x1, y1), fill=(0, 0, 0), width=3)
    draw.line((x0, y0, x0, y1), fill=(0, 0, 0), width=3)
    draw.line((x1, y0, x1, y1), fill=(0, 0, 0), width=3)

    for v in np.linspace(0, conv_upper, 6):
        y = y_left(v)
        draw.line((x0 - 9, y, x0, y), fill=(0, 0, 0), width=2)
        txt(x0 - 16, y - 9, f'{v:g}', anchor='ra', font_obj=font)
    for v in range(0, 101, 20):
        y = y_right(v)
        draw.line((x1, y, x1 + 9, y), fill=(0, 0, 0), width=2)
        txt(x1 + 16, y - 9, str(v), font_obj=font)

    if x_max <= 14:
        step = 2.0
    elif x_max <= 28:
        step = 4.0
    else:
        step = max(1.0, round(x_max / 6.0))
    x_ticks = np.arange(0, x_max + step * 0.5, step)
    for v in x_ticks:
        if v > x_max + 1e-9:
            continue
        x = xp(v)
        draw.line((x, y1, x, y1 + 9), fill=(0, 0, 0), width=2)
        label = f'{v:.0f}' if abs(v - round(v)) < 0.05 else f'{v:.1f}'
        txt(x, y1 + 18, label, anchor='ma', font_obj=font)

    txt(x0 + plot_w / 2, y1 + 76, 'Time on stream (h)', anchor='mm', font_obj=axis_font)
    rotated_txt(x0 - 74, y0 + plot_h / 2, f'{reactant_label} Conversion (%)', 90, font_obj=axis_font)
    rotated_txt(x1 + 82, y0 + plot_h / 2, 'Carbon-based selectivity (%)', -90, font_obj=axis_font)
    title = metadata.get('catalyst_id') or metadata.get('source_file') or 'GC run'
    txt(x0 + plot_w / 2, 38, title, anchor='mm', font_obj=title_font)

    groups = _selectivity_groups(df_sel, species_config)
    group_order = _selectivity_group_order(metadata)
    palette = {
        'CO': (50, 130, 210),
        'CH4': (75, 175, 70),
        'C2-C4 Paraffins': (220, 38, 38),
        'C2-C4 Olefins': (105, 88, 205),
        'C5+': (247, 205, 64),
        'Methanol': (34, 184, 194),
        'CO2': (216, 216, 216),
        'Other C products': (165, 165, 165),
    }
    group_values = {}
    for group in group_order:
        cols = groups.get(group, [])
        if not cols:
            continue
        vals = df_sel.loc[rxn.index, cols].fillna(0).sum(axis=1).to_numpy() * 100.0
        if np.isfinite(vals).any() and np.nanmax(vals) > 0.02:
            group_values[group] = vals

    finite_x = x_vals[np.isfinite(x_vals)]
    if len(finite_x) > 1:
        diffs = np.diff(np.sort(finite_x))
        spacing = float(np.median(diffs[diffs > 0])) if np.any(diffs > 0) else 1.0
        bar_px = max(7, min(32, int(abs(xp(x_min + spacing) - xp(x_min)) * 0.82)))
    else:
        bar_px = 20

    for i, xv in enumerate(x_vals):
        if not np.isfinite(xv):
            continue
        x = xp(xv)
        base = 0.0
        for group in group_order:
            vals = group_values.get(group)
            if vals is None:
                continue
            val = max(0.0, float(vals[i]))
            if val <= 0:
                continue
            y_top = y_right(base + val)
            y_bot = y_right(base)
            draw.rectangle((x - bar_px / 2, y_top, x + bar_px / 2, y_bot),
                           fill=palette[group], outline=(255, 255, 255))
            base += val

    def marker(x, y, color, shape, filled=True, size=7):
        x, y = float(x), float(y)
        if shape == 'circle_open':
            draw.ellipse((x - size, y - size, x + size, y + size), fill=color, outline=color, width=2)
        elif shape == 'square':
            draw.rectangle((x - size, y - size, x + size, y + size), fill=color if filled else (255, 255, 255), outline=color, width=2)
        elif shape == 'diamond':
            draw.polygon([(x, y - size), (x + size, y), (x, y + size), (x - size, y)], fill=color, outline=color)
        elif shape == 'triangle_up':
            draw.polygon([(x, y - size), (x + size, y + size), (x - size, y + size)], fill=color, outline=color)
        elif shape == 'triangle_down':
            draw.polygon([(x - size, y - size), (x + size, y - size), (x, y + size)], fill=color, outline=color)
        elif shape == 'pentagon':
            pts = []
            for j in range(5):
                a = -np.pi / 2 + j * 2 * np.pi / 5
                pts.append((x + np.cos(a) * size, y + np.sin(a) * size))
            draw.polygon(pts, fill=color, outline=color)
        elif shape == 'x':
            draw.line((x - size, y - size, x + size, y + size), fill=color, width=3)
            draw.line((x - size, y + size, x + size, y - size), fill=color, width=3)

    def draw_polyline(points, color, dashed=False, width_line=3):
        if len(points) < 2:
            return
        if not dashed:
            draw.line(points, fill=color, width=width_line)
            return
        for p0, p1 in zip(points[:-1], points[1:]):
            dx, dy = p1[0] - p0[0], p1[1] - p0[1]
            dist = max((dx * dx + dy * dy) ** 0.5, 1.0)
            pos = 0.0
            while pos < dist:
                end = min(pos + 10.0, dist)
                sx, sy = p0[0] + dx * pos / dist, p0[1] + dy * pos / dist
                ex, ey = p0[0] + dx * end / dist, p0[1] + dy * end / dist
                draw.line((sx, sy, ex, ey), fill=color, width=width_line)
                pos += 16.0

    conv_label = f'{reactant_label} Conversion'
    conv_points = [(xp(x), y_left(y)) for x, y in zip(x_vals, conv_vals) if np.isfinite(x) and pd.notna(y)]
    color, shape, filled = (0, 0, 0), 'circle_open', True
    draw_polyline(conv_points, color, dashed=False)
    for x, y in conv_points:
        marker(x, y, color, shape, filled, size=8)

    legend_items = [(conv_label, 'line', (0, 0, 0))]
    for group in group_order:
        if group in group_values:
            legend_items.append((group, 'box', palette[group]))
    legend_y = height - 72
    legend_x = x0 + 80
    row_gap = 48
    col_gap = 255
    for i, (label, kind, color) in enumerate(legend_items):
        col = i % 4
        row = i // 4
        lx = legend_x + col * col_gap
        ly = legend_y + row * row_gap
        if kind == 'line':
            draw.line((lx, ly + 12, lx + 48, ly + 12), fill=color, width=4)
            marker(lx + 24, ly + 12, color, 'circle_open', True, size=8)
        else:
            draw.rectangle((lx, ly, lx + 34, ly + 24), fill=color)
        _draw_legend_label(draw, lx + 58, ly - 1, label, font, legend_sub_font)

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_plot.png')
    img.save(path)
    return path


def _draw_stacked_selectivity_plot(df, df_sel, total_C_out, C_in_flow,
                                   reactant_label, metadata, species_config,
                                   output_dir):
    from PIL import Image, ImageDraw, ImageFont

    rxn = _analysis_reaction_rows(df)
    width, height = 1250, 900
    margin = {'l': 122, 'r': 64, 't': 106, 'b': 178}
    x0, y0 = margin['l'], margin['t']
    x1, y1 = width - margin['r'], height - margin['b']
    plot_w, plot_h = x1 - x0, y1 - y0

    img = Image.new('RGB', (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(img)
    def load_font(size, bold=False):
        names = ['arialbd.ttf', 'arial.ttf'] if bold else ['arial.ttf', 'segoeui.ttf']
        paths = []
        windir = os.environ.get('WINDIR', r'C:\Windows')
        for name in names:
            paths.append(os.path.join(windir, 'Fonts', name))
        for path in paths:
            try:
                return ImageFont.truetype(path, size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    font = load_font(16)
    small_font = load_font(14)
    legend_sub_font = load_font(10)
    title_font = load_font(20, bold=True)

    def txt(x, y, text, fill=(0, 0, 0), anchor=None, font_obj=None):
        kwargs = {'fill': fill, 'font': font_obj or font}
        if anchor:
            kwargs['anchor'] = anchor
        draw.text((int(x), int(y)), str(text), **kwargs)

    def text_w(text, font_obj=None):
        bbox = draw.textbbox((0, 0), str(text), font=font_obj or font)
        return bbox[2] - bbox[0]

    title = metadata.get('catalyst_id') or metadata.get('source_file') or 'GC run'
    txt(width / 2, 36, title, anchor='mm', font_obj=title_font)

    x_vals, x_label = _plot_x_values(rxn)
    valid_x = np.isfinite(x_vals)
    if not valid_x.any():
        x_vals = np.arange(len(rxn), dtype=float)
        valid_x = np.ones(len(rxn), dtype=bool)
        x_label = 'Reaction point'

    x_min, x_max = float(np.nanmin(x_vals)), float(np.nanmax(x_vals))
    if x_max == x_min:
        x_min -= 0.5
        x_max += 0.5
    x_pad = (x_max - x_min) * 0.03
    x_min -= x_pad
    x_max += x_pad

    def xp(v):
        return x0 + (float(v) - x_min) / (x_max - x_min) * plot_w

    def yp(v):
        v = max(0.0, min(105.0, float(v)))
        return y1 - (v / 105.0) * plot_h

    # Grid, axes, and ticks.
    for pct in range(0, 101, 25):
        y = yp(pct)
        draw.line((x0, y, x1, y), fill=(215, 215, 215), width=1)
        txt(x0 - 42, y - 8, str(pct), fill=(45, 45, 45), font_obj=small_font)
    draw.line((x0, y1, x1, y1), fill=(0, 0, 0), width=2)
    draw.line((x0, y0, x0, y1), fill=(0, 0, 0), width=2)
    txt(x0, y0 - 34, 'Selectivity / Conversion / Carbon Balance (%)')
    txt(x0 + plot_w / 2, height - 110, x_label, anchor='mm')

    tick_count = min(7, len(rxn))
    tick_idx = np.linspace(0, len(rxn) - 1, tick_count).round().astype(int)
    for idx in sorted(set(tick_idx)):
        if not np.isfinite(x_vals[idx]):
            continue
        x = xp(x_vals[idx])
        draw.line((x, y1, x, y1 + 5), fill=(0, 0, 0), width=1)
        val = x_vals[idx]
        label = f'{val:.0f}' if abs(val - round(val)) < 0.05 else f'{val:.1f}'
        txt(x, y1 + 12, label, fill=(45, 45, 45), anchor='ma', font_obj=small_font)

    palette = {
        'CO': (50, 130, 210),
        'CH4': (54, 211, 112),
        'C2-C4 Olefins': (116, 122, 230),
        'C2-C4 Paraffins': (216, 78, 71),
        'C5+': (247, 205, 64),
        'Methanol': (34, 184, 194),
        'CO2': (214, 214, 214),
        'Other C products': (180, 180, 180),
    }
    groups = _selectivity_groups(df_sel, species_config)
    group_order = _selectivity_group_order(metadata)
    group_values = {}
    for group in group_order:
        cols = groups.get(group, [])
        if cols:
            group_values[group] = df_sel.loc[rxn.index, cols].fillna(0).sum(axis=1).to_numpy() * 100.0

    if len(rxn) > 1:
        sorted_x = np.sort(x_vals[valid_x])
        diffs = np.diff(sorted_x)
        spacing = float(np.nanmedian(diffs[diffs > 0])) if np.any(diffs > 0) else 1.0
        bar_px = max(6, min(28, int(abs(xp(x_min + spacing) - xp(x_min)) * 0.72)))
    else:
        bar_px = 20

    for i, xv in enumerate(x_vals):
        if not np.isfinite(xv):
            continue
        x = xp(xv)
        base = 0.0
        for group in group_order:
            vals = group_values.get(group)
            if vals is None:
                continue
            val = max(0.0, float(vals[i]))
            if val <= 0:
                continue
            y_top = yp(base + val)
            y_bot = yp(base)
            draw.rectangle((x - bar_px / 2, y_top, x + bar_px / 2, y_bot),
                           fill=palette[group], outline=(255, 255, 255))
            base += val

    def line_points(vals):
        return [
            (xp(xv), yp(yv))
            for xv, yv in zip(x_vals, vals)
            if np.isfinite(xv) and pd.notna(yv)
        ]

    conv = pd.to_numeric(rxn.get('conversion', pd.Series(index=rxn.index, dtype=float)),
                         errors='coerce').to_numpy() * 100.0
    conv_pts = line_points(conv)
    if len(conv_pts) >= 2:
        draw.line(conv_pts, fill=(0, 0, 0), width=3)
    for x, y in conv_pts:
        draw.ellipse((x - 4, y - 4, x + 4, y + 4), fill=(0, 0, 0))

    if C_in_flow > 0:
        cb = (total_C_out.loc[rxn.index].to_numpy(dtype=float) / C_in_flow) * 100.0
        cb_pts = line_points(cb)
        if len(cb_pts) >= 2:
            for p0, p1 in zip(cb_pts[:-1], cb_pts[1:]):
                dx, dy = p1[0] - p0[0], p1[1] - p0[1]
                dist = max((dx * dx + dy * dy) ** 0.5, 1.0)
                pos = 0.0
                while pos < dist:
                    end = min(pos + 8.0, dist)
                    sx = p0[0] + dx * pos / dist
                    sy = p0[1] + dy * pos / dist
                    ex = p0[0] + dx * end / dist
                    ey = p0[1] + dy * end / dist
                    draw.line((sx, sy, ex, ey), fill=(160, 30, 45), width=2)
                    pos += 13.0
        for x, y in cb_pts:
            draw.polygon([(x, y - 5), (x + 4, y + 4), (x - 4, y + 4)],
                         fill=(160, 30, 45))

    # Legend.
    legend_y = height - 72
    legend_x = x0
    row_h = 30
    legend_gap = 28

    def reserve_legend_slot(label):
        nonlocal legend_x, legend_y
        item_w = 42 + _legend_label_width(draw, label, small_font, legend_sub_font) + legend_gap
        if legend_x + item_w > x1:
            legend_x = x0
            legend_y += row_h
        x = legend_x
        legend_x += item_w
        return x

    label = f'{reactant_label} Conversion'
    lx = reserve_legend_slot(label)
    draw.line((lx, legend_y + 7, lx + 28, legend_y + 7), fill=(0, 0, 0), width=3)
    draw.ellipse((lx + 11, legend_y + 2, lx + 21, legend_y + 12), fill=(0, 0, 0))
    _draw_legend_label(draw, lx + 36, legend_y, label, small_font, legend_sub_font)

    if C_in_flow > 0:
        label = 'Carbon Balance'
        lx = reserve_legend_slot(label)
        for start in range(0, 28, 13):
            draw.line((lx + start, legend_y + 7, lx + min(start + 8, 28), legend_y + 7),
                      fill=(160, 30, 45), width=2)
        draw.polygon([(lx + 15, legend_y + 2), (lx + 20, legend_y + 12),
                      (lx + 10, legend_y + 12)], fill=(160, 30, 45))
        _draw_legend_label(draw, lx + 36, legend_y, label, small_font, legend_sub_font)

    for group in group_order:
        if group not in group_values:
            continue
        lx = reserve_legend_slot(group)
        draw.rectangle((lx, legend_y + 1, lx + 28, legend_y + 13),
                       fill=palette[group], outline=(120, 120, 120))
        _draw_legend_label(draw, lx + 36, legend_y, group, small_font, legend_sub_font)

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_plot.png')
    img.save(path)
    return path


def _draw_co_oxidation_plot(df, df_sel, total_C_out, C_in_flow,
                            reactant_label, metadata, species_config,
                            output_dir):
    from PIL import Image, ImageDraw, ImageFont

    rxn = _analysis_reaction_rows(df)
    if rxn.empty:
        return _draw_stacked_selectivity_plot(
            df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
            species_config, output_dir)

    width, height = 1200, 760
    margin = {'l': 112, 'r': 58, 't': 86, 'b': 145}
    x0, y0 = margin['l'], margin['t']
    x1, y1 = width - margin['r'], height - margin['b']
    plot_w, plot_h = x1 - x0, y1 - y0

    img = Image.new('RGB', (width, height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    def load_font(size, bold=False):
        names = ['arialbd.ttf', 'arial.ttf'] if bold else ['arial.ttf', 'segoeui.ttf']
        windir = os.environ.get('WINDIR', r'C:\Windows')
        for name in names:
            try:
                return ImageFont.truetype(os.path.join(windir, 'Fonts', name), size=size)
            except OSError:
                continue
        return ImageFont.load_default()

    font = load_font(16)
    small_font = load_font(13)
    sub_font = load_font(10)
    axis_font = load_font(24)
    title_font = load_font(30)

    def txt(x, y, text, fill=(0, 0, 0), anchor=None, font_obj=None):
        kwargs = {'fill': fill, 'font': font_obj or font}
        if anchor:
            kwargs['anchor'] = anchor
        draw.text((int(x), int(y)), str(text), **kwargs)

    def rotated_txt(x, y, text, angle, font_obj=None):
        font_use = font_obj or font
        bbox = draw.textbbox((0, 0), str(text), font=font_use)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
        layer = Image.new('RGBA', (tw + 12, th + 12), (255, 255, 255, 0))
        layer_draw = ImageDraw.Draw(layer)
        layer_draw.text((6, 6), str(text), fill=(0, 0, 0, 255), font=font_use)
        rotated = layer.rotate(angle, expand=True)
        img.paste(rotated, (int(x - rotated.width / 2), int(y - rotated.height / 2)), rotated)

    if pd.to_numeric(rxn.get('time_on_stream_h'), errors='coerce').notna().any():
        x_vals = pd.to_numeric(rxn['time_on_stream_h'], errors='coerce').to_numpy(dtype=float)
        duration = _infer_run_duration_h(metadata)
        x_label = 'Time on stream (h)'
        x_min = 0.0
        x_max = max(float(duration or 0), float(np.nanmax(x_vals)), 1.0)
    else:
        x_vals, x_label = _plot_x_values(rxn)
        finite = x_vals[np.isfinite(x_vals)]
        if finite.size:
            x_min, x_max = float(np.nanmin(finite)), float(np.nanmax(finite))
        else:
            x_vals = np.arange(len(rxn), dtype=float)
            x_min, x_max = 0.0, max(float(len(rxn) - 1), 1.0)
            x_label = 'Reaction point'
        if x_max == x_min:
            x_min -= 0.5
            x_max += 0.5
        pad = (x_max - x_min) * 0.03
        x_min -= pad
        x_max += pad

    conv_vals = pd.to_numeric(rxn.get('conversion', pd.Series(index=rxn.index, dtype=float)),
                              errors='coerce').to_numpy(dtype=float) * 100.0
    co2_vals = None
    if 'S_CO2' in df_sel.columns:
        co2_vals = pd.to_numeric(df_sel.loc[rxn.index, 'S_CO2'], errors='coerce').to_numpy(dtype=float) * 100.0
    cb_vals = None
    if C_in_flow > 0:
        cb_vals = pd.to_numeric(total_C_out.loc[rxn.index], errors='coerce').to_numpy(dtype=float) / C_in_flow * 100.0

    all_vals = [conv_vals]
    if co2_vals is not None:
        all_vals.append(co2_vals)
    if cb_vals is not None:
        all_vals.append(cb_vals)
    finite_y = np.concatenate([vals[np.isfinite(vals)] for vals in all_vals if vals is not None])
    y_max = max(105.0, float(np.nanmax(finite_y)) * 1.05) if finite_y.size else 105.0
    y_min = 0.0

    def xp(v):
        return x0 + (float(v) - x_min) / (x_max - x_min) * plot_w

    def yp(v):
        v = max(y_min, min(y_max, float(v)))
        return y1 - (v - y_min) / (y_max - y_min) * plot_h

    for pct in np.linspace(0, y_max, 6):
        y = yp(pct)
        draw.line((x0, y, x1, y), fill=(218, 218, 218), width=1)
        txt(x0 - 14, y - 8, f'{pct:.0f}', fill=(45, 45, 45), anchor='ra', font_obj=small_font)
    draw.line((x0, y1, x1, y1), fill=(0, 0, 0), width=2)
    draw.line((x0, y0, x0, y1), fill=(0, 0, 0), width=2)

    tick_count = min(7, len(rxn))
    tick_idx = np.linspace(0, len(rxn) - 1, tick_count).round().astype(int)
    for idx in sorted(set(tick_idx)):
        if not np.isfinite(x_vals[idx]):
            continue
        x = xp(x_vals[idx])
        draw.line((x, y1, x, y1 + 7), fill=(0, 0, 0), width=1)
        val = x_vals[idx]
        label = f'{val:.0f}' if abs(val - round(val)) < 0.05 else f'{val:.1f}'
        txt(x, y1 + 16, label, fill=(45, 45, 45), anchor='ma', font_obj=small_font)

    title = metadata.get('catalyst_id') or metadata.get('source_file') or 'CO oxidation'
    txt(width / 2, 36, title, anchor='mm', font_obj=title_font)
    txt(x0 + plot_w / 2, height - 76, x_label, anchor='mm', font_obj=axis_font)
    rotated_txt(x0 - 76, y0 + plot_h / 2, 'CO conversion / selectivity (%)', 90, font_obj=axis_font)

    def points(vals):
        return [
            (xp(xv), yp(yv))
            for xv, yv in zip(x_vals, vals)
            if np.isfinite(xv) and np.isfinite(yv)
        ]

    def draw_series(vals, color, marker='circle', dashed=False):
        pts = points(vals)
        if len(pts) >= 2:
            if dashed:
                for p0, p1 in zip(pts[:-1], pts[1:]):
                    dx, dy = p1[0] - p0[0], p1[1] - p0[1]
                    dist = max((dx * dx + dy * dy) ** 0.5, 1.0)
                    pos = 0.0
                    while pos < dist:
                        end = min(pos + 9.0, dist)
                        draw.line((
                            p0[0] + dx * pos / dist,
                            p0[1] + dy * pos / dist,
                            p0[0] + dx * end / dist,
                            p0[1] + dy * end / dist,
                        ), fill=color, width=3)
                        pos += 15.0
            else:
                draw.line(pts, fill=color, width=3)
        for x, y in pts:
            if marker == 'triangle':
                draw.polygon([(x, y - 6), (x + 5, y + 5), (x - 5, y + 5)], fill=color)
            elif marker == 'square':
                draw.rectangle((x - 5, y - 5, x + 5, y + 5), fill=color)
            else:
                draw.ellipse((x - 5, y - 5, x + 5, y + 5), fill=color)

    conv_color = (0, 0, 0)
    co2_color = (75, 175, 70)
    cb_color = (160, 30, 45)
    draw_series(conv_vals, conv_color, marker='circle')
    if co2_vals is not None:
        draw_series(co2_vals, co2_color, marker='square')
    if cb_vals is not None:
        draw_series(cb_vals, cb_color, marker='triangle', dashed=True)

    legend = [('CO Conversion', conv_color, 'circle', False)]
    if co2_vals is not None:
        legend.append(('CO2 Selectivity', co2_color, 'square', False))
    if cb_vals is not None:
        legend.append(('Carbon Balance', cb_color, 'triangle', True))
    lx = x0
    ly = height - 118
    for label, color, marker, dashed in legend:
        if dashed:
            for start in range(0, 34, 15):
                draw.line((lx + start, ly + 8, lx + min(start + 9, 34), ly + 8), fill=color, width=3)
        else:
            draw.line((lx, ly + 8, lx + 34, ly + 8), fill=color, width=3)
        if marker == 'square':
            draw.rectangle((lx + 13, ly + 3, lx + 23, ly + 13), fill=color)
        elif marker == 'triangle':
            draw.polygon([(lx + 18, ly + 2), (lx + 24, ly + 14), (lx + 12, ly + 14)], fill=color)
        else:
            draw.ellipse((lx + 13, ly + 3, lx + 23, ly + 13), fill=color)
        _draw_legend_label(draw, lx + 44, ly, label, font, sub_font)
        lx += 250

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_plot.png')
    img.save(path)
    return path


def make_plots(df, df_sel, total_C_out, C_in_flow, reactant_label,
               ss_mask, metadata, carbon_cols, species_config, output_dir):
    style = str(metadata.get('plot_style') or 'auto').strip().lower()
    has_time_axis = bool(_infer_run_duration_h(metadata) or _metadata_float(metadata, 'injection_interval_min'))
    if _reaction_type(metadata=metadata) == 'co_oxidation':
        return _draw_co_oxidation_plot(
            df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
            species_config, output_dir)
    if style in {'single_axis_stacked', 'stacked_preview'}:
        return _draw_stacked_selectivity_plot(
            df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
            species_config, output_dir)
    if has_time_axis:
        return _draw_time_on_stream_plot(
            df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
            species_config, output_dir)
    return _draw_stacked_selectivity_plot(
        df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
        species_config, output_dir)

    from PIL import Image, ImageDraw, ImageFont

    width, height = 1500, 1100
    bg = (15, 17, 23)
    panel_bg = (22, 27, 34)
    fg = (232, 232, 232)
    muted = (139, 148, 158)
    grid = (45, 51, 59)
    green = (57, 211, 83)
    yellow = (227, 179, 65)
    colors = [
        (88, 166, 255), (247, 129, 102), (86, 211, 100),
        (227, 179, 65), (188, 140, 255), (255, 123, 114),
        (121, 192, 255), (255, 166, 87), (126, 231, 135),
        (210, 168, 255),
    ]

    img = Image.new('RGB', (width, height), bg)
    draw = ImageDraw.Draw(img)
    font = ImageFont.load_default()

    def txt(x, y, s, fill=fg):
        draw.text((int(x), int(y)), str(s), fill=fill, font=font)

    title = (
        f"{metadata.get('catalyst_id', 'Unknown')}   |   "
        f"T={metadata.get('temperature','?')}   "
        f"P={metadata.get('pressure','?')}   "
        f"GHSV={metadata.get('ghsv','?')}"
    )
    txt(36, 24, title, fg)

    rxn = _analysis_reaction_rows(df)
    inj = pd.to_numeric(rxn['inj_num'], errors='coerce')
    ss_rxn = ss_mask[rxn.index]

    def panel(box, title_text):
        x0, y0, x1, y1 = box
        draw.rounded_rectangle(box, radius=8, fill=panel_bg, outline=grid, width=1)
        txt(x0 + 14, y0 + 10, title_text, fg)
        return (x0 + 56, y0 + 44, x1 - 26, y1 - 48)

    def draw_line_chart(box, series, y_label='', percent=False, y_floor=0.0):
        px0, py0, px1, py1 = box
        plot_w, plot_h = px1 - px0, py1 - py0
        for frac in np.linspace(0, 1, 5):
            y = py1 - frac * plot_h
            draw.line((px0, y, px1, y), fill=grid, width=1)
        draw.line((px0, py1, px1, py1), fill=muted, width=1)
        draw.line((px0, py0, px0, py1), fill=muted, width=1)
        valid_x = inj.dropna()
        if valid_x.empty or not series:
            txt(px0 + 12, py0 + 20, 'No plottable data', muted)
            return
        x_min, x_max = float(valid_x.min()), float(valid_x.max())
        if x_max == x_min:
            x_max = x_min + 1.0
        all_y = []
        for _, vals, _ in series:
            all_y.extend([float(v) for v in vals if pd.notna(v)])
        y_max = max(all_y) if all_y else 1.0
        y_min = min(y_floor, min(all_y) if all_y else y_floor)
        if y_max <= y_min:
            y_max = y_min + 1.0
        y_max += (y_max - y_min) * 0.08

        def xp(v):
            return px0 + (float(v) - x_min) / (x_max - x_min) * plot_w

        def yp(v):
            return py1 - (float(v) - y_min) / (y_max - y_min) * plot_h

        for idx, (label, vals, color) in enumerate(series):
            points = [
                (xp(x), yp(y))
                for x, y in zip(inj, vals)
                if pd.notna(x) and pd.notna(y)
            ]
            if len(points) >= 2:
                draw.line(points, fill=color, width=3)
            for x, y in points:
                draw.ellipse((x - 3, y - 3, x + 3, y + 3), fill=color)
            lx = px0 + 10 + (idx % 5) * 150
            ly = py0 + 8 + (idx // 5) * 18
            draw.line((lx, ly + 6, lx + 22, ly + 6), fill=color, width=3)
            txt(lx + 28, ly, label, fg)
        txt(px0 - 48, py0 - 2, f"{y_max:.1f}{'%' if percent else ''}", muted)
        txt(px0 - 40, py1 - 10, f"{y_min:.1f}", muted)
        txt(px0 + plot_w / 2 - 40, py1 + 22, 'Injection number', muted)
        txt(px0 - 50, py0 + plot_h / 2, y_label, muted)

    flow_box = panel((36, 70, 1464, 520), 'Molar flows vs. injection number')
    plotted = set()
    flow_series = []
    priority = [reactant_label, 'CO2', 'CO', 'H2', 'CH4', 'C2H4', 'C2H6', 'C3H6', 'C3H8']
    for sp in priority + carbon_cols:
        if sp in plotted or sp not in rxn.columns:
            continue
        vals = pd.to_numeric(rxn[sp], errors='coerce')
        if vals.notna().sum() > 0:
            flow_series.append((sp, vals, colors[len(flow_series) % len(colors)]))
            plotted.add(sp)
        if len(flow_series) >= 10:
            break
    draw_line_chart(flow_box, flow_series, 'sccm')

    conv_box = panel((36, 560, 720, 1040), f'{reactant_label} conversion')
    conv_series = []
    if 'conversion' in rxn.columns:
        vals = pd.to_numeric(rxn['conversion'], errors='coerce') * 100.0
        if vals.notna().sum() > 0:
            conv_series.append(('conversion', vals, colors[0]))
    draw_line_chart(conv_box, conv_series, '%', percent=True)
    if ss_rxn.any() and 'conversion' in rxn.columns:
        ss_mean = pd.to_numeric(rxn.loc[ss_rxn, 'conversion'], errors='coerce').mean() * 100
        txt(86, 590, f'SS avg: {ss_mean:.2f}%', green)

    bar_box = panel((780, 560, 1464, 1040), 'Carbon selectivity (steady state)')
    bx0, by0, bx1, by1 = bar_box
    if not df_sel.empty and ss_rxn.any():
        ss_sel = df_sel.loc[rxn.index][ss_rxn].mean() * 100
        ss_sel = ss_sel[ss_sel > 0.05].sort_values(ascending=False).head(10)
        if not ss_sel.empty:
            max_val = max(100.0, float(ss_sel.max()) * 1.15)
            bar_w = max(16, int((bx1 - bx0) / max(len(ss_sel), 1) * 0.62))
            gap = max(8, int((bx1 - bx0 - len(ss_sel) * bar_w) / max(len(ss_sel), 1)))
            for idx, (col, val) in enumerate(ss_sel.items()):
                x0 = bx0 + gap // 2 + idx * (bar_w + gap)
                h = (float(val) / max_val) * (by1 - by0 - 36)
                y0 = by1 - h
                color = colors[idx % len(colors)]
                draw.rectangle((x0, y0, x0 + bar_w, by1), fill=color)
                txt(x0, by1 + 8, col.replace('S_', '')[:10], muted)
                txt(x0, y0 - 16, f'{val:.1f}', fg)
        else:
            txt(bx0 + 12, by0 + 20, 'No selectivity above threshold', muted)
        if C_in_flow > 0:
            cb = (total_C_out[rxn.index][ss_rxn].mean() / C_in_flow) * 100
            cb_color = green if 90 <= cb <= 110 else yellow
            txt(bx1 - 160, by0 + 8, f'C balance: {cb:.1f}%', cb_color)
    else:
        txt(bx0 + 12, by0 + 20, 'No steady-state selectivity data', muted)

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_plot.png')
    img.save(path)
    return path

    fig = plt.figure(figsize=(15, 11))
    fig.patch.set_facecolor('#0f1117')
    fig.suptitle(
        f"{metadata['catalyst_id']}   ·   T={metadata.get('temperature','?')}   "
        f"P={metadata.get('pressure','?')}   GHSV={metadata.get('ghsv','?')}",
        fontsize=11, fontweight='bold', y=0.98, color='#e8e8e8')

    gs  = gridspec.GridSpec(2, 2, figure=fig, hspace=0.48, wspace=0.32)
    rxn = _analysis_reaction_rows(df)
    inj = rxn['inj_num']

    ax_bg = '#161b22'
    ax_fg = '#e8e8e8'
    grid_c = '#2d333b'

    def style_ax(ax):
        ax.set_facecolor(ax_bg)
        ax.tick_params(colors=ax_fg, labelsize=8)
        ax.xaxis.label.set_color(ax_fg)
        ax.yaxis.label.set_color(ax_fg)
        ax.title.set_color(ax_fg)
        for spine in ax.spines.values():
            spine.set_edgecolor(grid_c)
        ax.grid(True, color=grid_c, alpha=0.6, linewidth=0.6)

    def shade(ax):
        if ss_mask[rxn.index].any():
            ss_inj = inj[ss_mask[rxn.index]]
            ax.axvspan(ss_inj.iloc[0] - 0.5, ss_inj.iloc[-1] + 0.5,
                       alpha=0.12, color='#39d353', zorder=0)

    colors = ['#58a6ff','#f78166','#56d364','#e3b341','#bc8cff',
              '#ff7b72','#79c0ff','#ffa657','#7ee787','#d2a8ff']

    # Panel 1: Molar flows
    ax1 = fig.add_subplot(gs[0, :])
    style_ax(ax1)
    plotted, ci = set(), 0
    priority = [reactant_label, 'CO2', 'CO', 'H2', 'CH4', 'C2H4', 'C2H6', 'C3H6', 'C3H8']
    for sp in priority + carbon_cols:
        if sp in rxn.columns and sp not in plotted and rxn[sp].notna().sum() > 0:
            ax1.plot(inj, rxn[sp], marker='o', markersize=3, linewidth=1.8,
                     label=sp, color=colors[ci % len(colors)])
            plotted.add(sp); ci += 1
    shade(ax1)
    ax1.set_xlabel('Injection number', fontsize=9)
    ax1.set_ylabel('Molar flow (sccm equiv.)', fontsize=9)
    ax1.set_title('Molar Flows vs. Injection Number')
    leg = ax1.legend(fontsize=7, ncol=6, loc='upper right',
                     facecolor='#1c2128', edgecolor=grid_c, labelcolor=ax_fg)

    # Panel 2: Conversion
    ax2 = fig.add_subplot(gs[1, 0])
    style_ax(ax2)
    if 'conversion' in rxn.columns and rxn['conversion'].notna().any():
        ax2.plot(inj, rxn['conversion'] * 100,
                 color='#58a6ff', marker='o', markersize=4, linewidth=2)
        shade(ax2)
        ss_rxn = ss_mask[rxn.index]
        if ss_rxn.any():
            ss_mean = rxn.loc[ss_rxn, 'conversion'].mean() * 100
            ax2.axhline(ss_mean, color='#39d353', linestyle='--', linewidth=1.5,
                        label=f'SS avg: {ss_mean:.2f}%')
            ax2.legend(fontsize=8, facecolor='#1c2128', edgecolor=grid_c, labelcolor=ax_fg)
    ax2.set_xlabel('Injection number', fontsize=9)
    ax2.set_ylabel('Conversion (%)', fontsize=9)
    ax2.set_title(f'{reactant_label} Conversion')

    # Panel 3: Selectivity bar
    ax3 = fig.add_subplot(gs[1, 1])
    style_ax(ax3)
    ss_rxn = ss_mask[rxn.index]
    if not df_sel.empty and ss_rxn.any():
        ss_sel = df_sel.loc[rxn.index][ss_rxn].mean() * 100
        ss_sel = ss_sel[ss_sel > 0.05].sort_values(ascending=False)
        if not ss_sel.empty:
            labels_bar = [s.replace('S_', '') for s in ss_sel.index]
            bar_colors = [colors[i % len(colors)] for i in range(len(ss_sel))]
            ax3.bar(range(len(ss_sel)), ss_sel.values, color=bar_colors, edgecolor='#0f1117', linewidth=0.5)
            ax3.set_xticks(range(len(ss_sel)))
            ax3.set_xticklabels(labels_bar, rotation=35, ha='right', fontsize=8)
            ax3.set_ylim(0, 115)
            if C_in_flow > 0:
                cb = (total_C_out[rxn.index][ss_rxn].mean() / C_in_flow) * 100
                cb_color = '#39d353' if 90 <= cb <= 110 else '#e3b341'
                ax3.text(0.98, 0.97, f'C balance: {cb:.1f}%',
                         transform=ax3.transAxes, ha='right', va='top', fontsize=9,
                         color=cb_color, fontweight='bold',
                         bbox=dict(boxstyle='round,pad=0.4', fc='#1c2128', ec=cb_color, alpha=0.9))
    ax3.set_ylabel('Selectivity (%)', fontsize=9)
    ax3.set_title('Carbon Selectivity (Steady State)')

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_plot.png')
    plt.savefig(path, dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
    plt.close()
    return path


# ─────────────────────────────────────────────────────────────────────────────
# SAVE OUTPUTS
# ─────────────────────────────────────────────────────────────────────────────

def _excel_sheet_ref(sheet, cell_ref):
    return f"'{sheet}'!{cell_ref}"


def _raw_amount_ref(inj, species_header, sheet_name='Raw Original'):
    ref = (inj.get('source_refs') or {}).get('amounts', {}).get(species_header)
    if not ref:
        return None
    return _excel_sheet_ref(sheet_name, ref)


def _raw_area_ref(inj, species_header, sheet_name='Raw Original'):
    ref = (inj.get('source_refs') or {}).get('areas', {}).get(species_header)
    if not ref:
        return None
    return _excel_sheet_ref(sheet_name, ref)


def _amount_expr_for_header(inj, species_header, sheet_name='Raw Original',
                            fixed_rf_cells=None):
    raw = _raw_amount_ref(inj, species_header, sheet_name=sheet_name)
    if raw:
        return raw
    derived = (inj.get('source_refs') or {}).get('area_derived', {}).get(species_header)
    if not derived:
        return None
    area_ref = derived.get('area')
    ref_amount = derived.get('reference_amount')
    ref_area = derived.get('reference_area')
    rf_key = derived.get('response_factor_key')
    if area_ref and rf_key and fixed_rf_cells and fixed_rf_cells.get(rf_key):
        return (
            f'({_excel_sheet_ref(sheet_name, area_ref)}*'
            f'{fixed_rf_cells[rf_key]})')
    if area_ref and ref_amount and ref_area:
        return (
            f'({_excel_sheet_ref(sheet_name, area_ref)}*'
            f'{_excel_sheet_ref(sheet_name, ref_amount)}/'
            f'{_excel_sheet_ref(sheet_name, ref_area)})')
    rf = derived.get('response_factor')
    if area_ref and rf:
        return f'({_excel_sheet_ref(sheet_name, area_ref)}*{float(rf):.12g})'
    return None


def _copy_source_sheet_to_workbook(out_wb, source_filepath, worksheet_index=0,
                                   sheet_name='Raw Original', insert_at=1):
    from copy import copy
    from openpyxl import load_workbook

    src_wb = load_workbook(source_filepath, data_only=False)
    src_ws = src_wb.worksheets[min(max(int(worksheet_index or 0), 0), len(src_wb.worksheets) - 1)]
    raw_ws = out_wb.create_sheet(sheet_name, insert_at)
    for row in src_ws.iter_rows():
        for cell in row:
            dst = raw_ws[cell.coordinate]
            dst.value = cell.value
            if cell.has_style:
                dst.font = copy(cell.font)
                dst.fill = copy(cell.fill)
                dst.border = copy(cell.border)
                dst.alignment = copy(cell.alignment)
                dst.number_format = cell.number_format
    for key, dim in src_ws.column_dimensions.items():
        raw_ws.column_dimensions[key].width = dim.width
    for key, dim in src_ws.row_dimensions.items():
        raw_ws.row_dimensions[key].height = dim.height
    for merged in src_ws.merged_cells.ranges:
        raw_ws.merge_cells(str(merged))
    raw_ws.freeze_panes = src_ws.freeze_panes
    return raw_ws


def _flow_formula_for_header(inj, species_header, species_config, has_bridge,
                             ar_key, ch4_tcd_key, ch4_fid_key, ar_inlet_cell,
                             fixed_rf_cells=None):
    cfg = species_config.get(species_header, {})
    raw = _amount_expr_for_header(
        inj, cfg.get('source_header') or species_header,
        fixed_rf_cells=fixed_rf_cells)
    ar_raw = _raw_amount_ref(inj, ar_key)
    if not raw or not ar_raw:
        return None
    if cfg.get('det') == 'TCD':
        formula = f'{ar_inlet_cell}*{raw}/{ar_raw}'
        fraction = float(cfg.get('flow_fraction') or 1.0)
        return f'={formula}*{fraction:g}' if fraction != 1.0 else f'={formula}'
    use_bridge_for_species = cfg.get('fid_bridge', True) is not False
    if has_bridge and use_bridge_for_species:
        ch4_tcd = _raw_amount_ref(inj, ch4_tcd_key)
        ch4_fid = _raw_amount_ref(inj, ch4_fid_key)
        if ch4_tcd and ch4_fid:
            formula = f'{ar_inlet_cell}*({ch4_tcd}/{ch4_fid})*({raw}/{ar_raw})'
            fraction = float(cfg.get('flow_fraction') or 1.0)
            return f'={formula}*{fraction:g}' if fraction != 1.0 else f'={formula}'
    formula = f'{ar_inlet_cell}*{raw}/{ar_raw}'
    fraction = float(cfg.get('flow_fraction') or 1.0)
    return f'={formula}*{fraction:g}' if fraction != 1.0 else f'={formula}'


def _autosize_sheet(ws, max_width=36):
    for col in ws.columns:
        letter = col[0].column_letter
        width = 10
        for cell in col:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def _write_bypass_processed_sheet(wb, bypass_data, species_config, inlet_labels,
                                  ar_inlet_cell, insert_at=3,
                                  source_sheet_name='Bypass Original'):
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet('Bypass Processed', insert_at)
    ws['A1'] = 'Bypass-derived inlet flows'
    ws['A1'].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(['Species', 'Average inlet sccm', 'Trace'])
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='5B9BD5')

    label_to_header = {cfg['label']: header for header, cfg in species_config.items()}
    ar_key = find_ar_key(species_config)
    active_labels = [label for label in inlet_labels if label != 'Ar' and label_to_header.get(label)]

    data_header_row = max(7, 5 + len(active_labels))
    headers = ['Bypass point', 'Raw label', 'Injection #', 'Raw Ar amount']
    for label in active_labels:
        headers.extend([f'Raw {label} amount', f'{label}/Ar ratio', f'{label} inlet sccm'])
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(data_header_row, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='70AD47')
        cell.alignment = Alignment(horizontal='center')

    injections = bypass_data.get('injections', []) if bypass_data else []
    inlet_cols = {}
    for i, inj in enumerate(injections, start=1):
        row = data_header_row + i
        label_ref = (inj.get('source_refs') or {}).get('label')
        ws.cell(row, 1).value = i
        ws.cell(row, 2).value = f'={_excel_sheet_ref(source_sheet_name, label_ref)}' if label_ref else inj.get('label', '')
        ws.cell(row, 3).value = inj.get('inj_num')
        ar_ref = _raw_amount_ref(inj, ar_key, sheet_name=source_sheet_name)
        ws.cell(row, 4).value = f'={ar_ref}' if ar_ref else None
        col = 5
        for label in active_labels:
            header = label_to_header.get(label)
            raw_ref = _raw_amount_ref(inj, header, sheet_name=source_sheet_name)
            raw_cell = ws.cell(row, col)
            ratio_cell = ws.cell(row, col + 1)
            inlet_cell = ws.cell(row, col + 2)
            raw_cell.value = f'={raw_ref}' if raw_ref else None
            ratio_cell.value = f'=IF(D{row}>0,{raw_cell.coordinate}/D{row},"")'
            inlet_cell.value = f'=IF({ratio_cell.coordinate}<>"",{ratio_cell.coordinate}*{ar_inlet_cell},"")'
            inlet_cols.setdefault(label, []).append(inlet_cell.coordinate)
            col += 3

    summary_refs = {}
    for idx, label in enumerate(active_labels, start=4):
        ws.cell(idx, 1).value = label
        if inlet_cols.get(label):
            refs = ','.join(inlet_cols[label])
            ws.cell(idx, 2).value = f'=AVERAGE({refs})'
            ws.cell(idx, 3).value = f'Average of {label}/Ar normalized by Settings Ar inlet'
            summary_refs[label] = _excel_sheet_ref('Bypass Processed', f'$B${idx}')

    ws.freeze_panes = f'A{data_header_row + 1}'
    _autosize_sheet(ws)
    return summary_refs


def _write_area_fallback_sheet(wb, raw_data, insert_at=4, fixed_rf_cells=None):
    from openpyxl.styles import Alignment, Font, PatternFill

    converted = raw_data.get('area_derived_amounts', []) if raw_data else []
    skipped = raw_data.get('area_derivation_skipped', []) if raw_data else []
    if not converted and not skipped:
        return None

    ws = wb.create_sheet('Area Fallbacks', insert_at)
    ws['A1'] = 'Area-only C5/C6 amount estimates'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = 'C5/C6 unknowns use editable Settings response factors; named C5/C6 compounds use pentane/hexane response factors when available.'
    headers = [
        'Raw label', 'Species', 'Carbon #', 'Raw peak area', 'Reference',
        'Reference amount', 'Reference area', 'Response source',
        'Derived amount'
    ]
    header_row = 4
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, col)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='C65911')
        cell.alignment = Alignment(horizontal='center')

    for inj in raw_data.get('injections', []):
        label_ref = (inj.get('source_refs') or {}).get('label')
        raw_label = f'={_excel_sheet_ref("Raw Original", label_ref)}' if label_ref else inj.get('label', '')
        derived = (inj.get('source_refs') or {}).get('area_derived', {})
        for species, info in derived.items():
            row = ws.max_row + 1
            area_ref = info.get('area')
            ref_amount = info.get('reference_amount')
            ref_area = info.get('reference_area')
            ws.cell(row, 1).value = raw_label
            ws.cell(row, 2).value = species
            ws.cell(row, 3).value = info.get('carbon_number')
            ws.cell(row, 4).value = f'={_excel_sheet_ref("Raw Original", area_ref)}' if area_ref else None
            ws.cell(row, 5).value = info.get('reference_header')
            ws.cell(row, 6).value = f'={_excel_sheet_ref("Raw Original", ref_amount)}' if ref_amount else None
            ws.cell(row, 7).value = f'={_excel_sheet_ref("Raw Original", ref_area)}' if ref_area else None
            ws.cell(row, 8).value = info.get('response_source')
            if area_ref and ref_amount and ref_area:
                ws.cell(row, 9).value = (
                    f'={_excel_sheet_ref("Raw Original", area_ref)}*'
                    f'{_excel_sheet_ref("Raw Original", ref_amount)}/'
                    f'{_excel_sheet_ref("Raw Original", ref_area)}')
            elif area_ref and info.get('response_factor_key') and fixed_rf_cells and fixed_rf_cells.get(info.get('response_factor_key')):
                ws.cell(row, 9).value = (
                    f'={_excel_sheet_ref("Raw Original", area_ref)}*'
                    f'{fixed_rf_cells[info["response_factor_key"]]}')
            elif area_ref and info.get('response_factor'):
                ws.cell(row, 9).value = (
                    f'={_excel_sheet_ref("Raw Original", area_ref)}*'
                    f'{float(info["response_factor"]):.12g}')

    if skipped:
        start = ws.max_row + 3
        ws.cell(start, 1).value = 'Skipped area-only peaks'
        ws.cell(start, 1).font = Font(bold=True)
        for col, header in enumerate(['Species', 'Carbon #', 'Reason'], start=1):
            cell = ws.cell(start + 1, col)
            cell.value = header
            cell.font = Font(bold=True)
        for item in skipped:
            ws.append([item.get('species'), item.get('carbon_number'), item.get('reason')])

    ws.freeze_panes = 'A5'
    _autosize_sheet(ws)
    return ws


def _write_gc_analysis_workbook(df, df_sel, total_C_out, C_in_flow, reactant_label,
                                metadata, species_config, output_dir, source_filepath,
                                raw_data, inlet_flows, has_bridge,
                                bypass_filepath=None, bypass_data=None,
                                ss_mask=None):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    path = os.path.join(output_dir, f'{_output_file_prefix(metadata)}_gc_analysis.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Processed'
    try:
        _copy_source_sheet_to_workbook(
            wb, source_filepath, raw_data.get('worksheet_index', 0) if raw_data else 0,
            sheet_name='Raw Original', insert_at=1)
    except Exception:
        raw_ws = wb.create_sheet('Raw Original', 1)
        raw_ws.append(['Raw sheet copy failed; processed formulas use parsed values where needed.'])

    has_separate_bypass_file = bool(bypass_filepath and bypass_data)
    has_bypass = bool(bypass_data and bypass_data.get('injections'))
    if has_separate_bypass_file:
        try:
            _copy_source_sheet_to_workbook(
                wb, bypass_filepath, bypass_data.get('worksheet_index', 0),
                sheet_name='Bypass Original', insert_at=2)
        except Exception:
            byp_ws = wb.create_sheet('Bypass Original', 2)
            byp_ws.append(['Bypass raw sheet copy failed; parsed bypass values are preserved on Bypass Processed.'])

    area_fallback_count = len(raw_data.get('area_derived_amounts', []) if raw_data else [])
    area_skipped_count = len(raw_data.get('area_derivation_skipped', []) if raw_data else [])

    label_to_header = {cfg['label']: header for header, cfg in species_config.items()}
    ar_key = find_ar_key(species_config)
    ch4_tcd_key = find_ch4_tcd_key(species_config)
    ch4_fid_key = find_ch4_fid_key(species_config)
    inlet_order = ['Ar', reactant_label]
    inlet_order.extend(label for label in inlet_flows if label not in inlet_order)

    settings = wb.create_sheet('Settings')
    settings.append(('Parameter', 'Value', 'Notes'))
    setting_rows = {}

    def add_setting(name, value, note=''):
        row_idx = settings.max_row + 1
        settings.append((name, value, note))
        setting_rows[name] = row_idx
        return _excel_sheet_ref('Settings', f'$B${row_idx}')

    duration_cell = add_setting('run_duration_h', metadata.get('run_duration_h'), 'Optional axis extent for time-on-stream plots.')
    interval_cell = add_setting('injection_interval_min', metadata.get('injection_interval_min'), 'Used to convert accepted injection count to time.')
    add_setting('rejected_initial_injections', metadata.get('rejected_initial_injections'), 'Initial reaction rows excluded from the plotted time axis.')
    add_setting('registered_reaction_injections', metadata.get('registered_reaction_injections'), 'Accepted/plotted reaction point count when specified.')
    npoints_cell = add_setting('plot_reaction_points', metadata.get('plot_reaction_points'), 'Number of rows with assigned time_on_stream_h.')
    add_setting('plot_style', metadata.get('plot_style', 'auto'), 'GUI-selected plot rendering mode.')
    add_setting('inlet_flow_source', metadata.get('inlet_flow_source'), 'manual or bypass-derived.')
    add_setting('bypass_source', metadata.get('bypass_source'), 'manual, separate_file, or same_file.')
    add_setting('bypass_file', metadata.get('bypass_file'), 'Separate bypass workbook or same input file used for inlet normalization.')
    add_setting('same_file_bypass_mode', metadata.get('same_file_bypass_mode'), 'How same-file bypass rows were selected.')
    add_setting('same_file_bypass_rows', metadata.get('same_file_bypass_rows'), 'Default same-file edge rows checked when labels are absent.')
    add_setting('same_file_bypass_detected', metadata.get('same_file_bypass_detected'), 'Same-file rows marked as bypass before omission/averaging.')
    add_setting('same_file_bypass_detect_method', metadata.get('same_file_bypass_detect_method'), 'Detection method used for same-file bypass rows.')
    add_setting('bypass_used', 'yes' if has_bypass else 'no', 'Whether bypass data are integrated into this workbook.')
    add_setting('bypass_omit_initial', metadata.get('bypass_omit_initial'), 'Initial bypass rows omitted before inlet averaging.')
    add_setting('bypass_points_used', metadata.get('bypass_points_used'), 'Number of bypass rows averaged after omission; blank means all remaining.')
    add_setting('bypass_selected_points', metadata.get('bypass_selected_points'), 'Actual bypass rows used for inlet averaging.')
    fixed_rf_cells = {}
    for cn in (5, 6):
        key = _unknown_response_factor_key(cn)
        fixed_rf_cells[key] = add_setting(
            key,
            _metadata_float(metadata, key, _DEFAULT_UNKNOWN_AREA_RESPONSE_FACTORS[cn]),
            f'Editable response factor used for C{cn} area-only unknowns.')
    add_setting('area_derived_amounts', area_fallback_count, 'Area-only C5/C6 unknowns use the editable response factors above; named C5/C6 peaks can use pentane/hexane response factors.')
    add_setting('area_derivation_skipped', area_skipped_count, 'Area-only C5/C6 peaks not converted because the reference response was unavailable.')
    inlet_setting_cells = {}
    for label in inlet_order:
        inlet_setting_cells[label] = add_setting(
            f'{label}_inlet_sccm', inlet_flows.get(label),
            'Ar is user-entered; non-Ar values can be bypass-derived.')
    settings['A1'].font = settings['B1'].font = settings['C1'].font = Font(bold=True)
    settings.freeze_panes = 'A2'
    ar_inlet_cell = inlet_setting_cells.get('Ar')

    if area_fallback_count or area_skipped_count:
        _write_area_fallback_sheet(
            wb, raw_data, insert_at=4 if has_separate_bypass_file else (3 if has_bypass else 2),
            fixed_rf_cells=fixed_rf_cells)

    if has_bypass and ar_inlet_cell:
        bypass_insert_at = 3 if has_separate_bypass_file else 2
        bypass_source_sheet = bypass_data.get('source_sheet_name') or (
            'Bypass Original' if has_separate_bypass_file else 'Raw Original')
        summary_refs = _write_bypass_processed_sheet(
            wb, bypass_data, species_config, inlet_order, ar_inlet_cell,
            insert_at=bypass_insert_at, source_sheet_name=bypass_source_sheet)
        for label, ref in summary_refs.items():
            row_idx = setting_rows.get(f'{label}_inlet_sccm')
            if row_idx:
                settings.cell(row_idx, 2).value = f'={ref}'

    reactant_inlet_cell = inlet_setting_cells.get(reactant_label)

    groups = _selectivity_groups(df_sel, species_config)
    group_order = _selectivity_group_order(metadata)
    present_groups = [g for g in group_order if g in groups]
    product_labels = [col.replace('S_', '') for col in df_sel.columns]
    flow_labels = [reactant_label] + product_labels

    headers = [
        'Use in SS summary', 'Row status', 'Accepted point', 'Time on stream h',
        'Injection #', 'Catalyst ID',
        f'{reactant_label} out sccm', f'{reactant_label} conversion %',
        'Product C out', 'Carbon balance %',
    ]
    headers.extend(f'Sel {g} %' for g in present_groups)
    headers.extend(f'Flow {label} sccm' for label in flow_labels)
    headers.append('Raw label')
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center')

    raw_injections = raw_data.get('injections', []) if raw_data else []
    processed_rows = []
    accepted_count = 0
    ss_lookup = ss_mask if ss_mask is not None else pd.Series(False, index=df.index)
    for df_idx, row in df.iterrows():
        inj = raw_injections[df_idx] if df_idx < len(raw_injections) else {}
        if inj.get('source_kind') == 'bypass_file':
            continue
        analysis_include = bool(row.get('analysis_include', not row.get('is_bypass', False)))
        in_ss = bool(ss_lookup.loc[df_idx]) if df_idx in ss_lookup.index else False
        accepted_point = ''
        if analysis_include:
            accepted_count += 1
            accepted_point = accepted_count
        status = row.get('row_status') or (
            'Bypass / inlet normalization' if row.get('is_bypass') else 'Reaction included')
        if analysis_include and not in_ss:
            status = 'Plotted reaction point; outside SS range'
        if in_ss:
            status = 'Included in SS summary'
        processed_rows.append((accepted_point, in_ss, status, df_idx, row, inj))

    flow_start_col = 11 + len(present_groups)
    raw_label_col = flow_start_col + len(flow_labels)
    raw_label_example_ref = None
    excluded_fill = PatternFill('solid', fgColor='FFF2CC')
    bypass_fill = PatternFill('solid', fgColor='DDEBF7')
    outside_fill = PatternFill('solid', fgColor='E7E6E6')
    for accepted_point, in_ss, status, df_idx, row, inj in processed_rows:
        excel_row = ws.max_row + 1
        label_ref = (inj.get('source_refs') or {}).get('label')
        raw_label_formula = f'={_excel_sheet_ref("Raw Original", label_ref)}' if label_ref else row.get('label', '')
        ws.cell(excel_row, 1).value = 'yes' if in_ss else 'no'
        ws.cell(excel_row, 2).value = status
        ws.cell(excel_row, 3).value = accepted_point
        ws.cell(excel_row, 4).value = (
            f'=IF(C{excel_row}<>"",'
            f'IF({interval_cell}>0,(C{excel_row}-1)*{interval_cell}/60,'
            f'IF({npoints_cell}>1,(C{excel_row}-1)*{duration_cell}/({npoints_cell}-1),0)),"")')
        ws.cell(excel_row, 5).value = row.get('inj_num')
        ws.cell(excel_row, 6).value = metadata.get('catalyst_id') or ''
        ws.cell(excel_row, raw_label_col).value = raw_label_formula
        if raw_label_example_ref is None:
            raw_label_example_ref = f"='Processed'!{ws.cell(excel_row, raw_label_col).coordinate}"

        flow_cell_by_label = {}
        for offset, label in enumerate(flow_labels):
            col = flow_start_col + offset
            header = label_to_header.get(label)
            formula = _flow_formula_for_header(
                inj, header, species_config, has_bridge,
                ar_key, ch4_tcd_key, ch4_fid_key, ar_inlet_cell,
                fixed_rf_cells=fixed_rf_cells) if header else None
            ws.cell(excel_row, col).value = formula if formula else row.get(label)
            flow_cell_by_label[label] = ws.cell(excel_row, col).coordinate

        reactant_flow_cell = flow_cell_by_label.get(reactant_label)
        ws.cell(excel_row, 7).value = f'={reactant_flow_cell}' if reactant_flow_cell else row.get(reactant_label)
        ws.cell(excel_row, 8).value = (
            f'=IF($B{excel_row}="Bypass / inlet normalization","",'
            f'({reactant_inlet_cell}-G{excel_row})/{reactant_inlet_cell}*100)')

        product_terms = []
        for label in product_labels:
            cn = get_cn(label, species_config)
            cell = flow_cell_by_label.get(label)
            if cell and cn:
                product_terms.append(f'{cn}*{cell}')
        ws.cell(excel_row, 9).value = '=' + '+'.join(product_terms) if product_terms else '=0'
        ws.cell(excel_row, 10).value = (
            f'=IF($B{excel_row}="Bypass / inlet normalization","",'
            f'(G{excel_row}+I{excel_row})/{reactant_inlet_cell}*100)')

        for g_idx, group in enumerate(present_groups, start=11):
            terms = []
            for sel_col in groups.get(group, []):
                label = sel_col.replace('S_', '')
                cn = get_cn(label, species_config)
                cell = flow_cell_by_label.get(label)
                if cell and cn:
                    terms.append(f'{cn}*{cell}')
            ws.cell(excel_row, g_idx).value = (
                f'=IF(I{excel_row}>0,({"+".join(terms)})/I{excel_row}*100,"")'
                if terms else '')

        row_fill = None
        if row.get('is_bypass'):
            row_fill = bypass_fill
        elif not bool(row.get('analysis_include', False)):
            row_fill = excluded_fill
        elif not in_ss:
            row_fill = outside_fill
        if row_fill:
            for cell in ws[excel_row]:
                cell.fill = row_fill

    guide_row = settings.max_row + 3
    settings.cell(guide_row, 1).value = 'Calculation illustration'
    settings.cell(guide_row, 1).font = Font(bold=True, size=13)
    guide_header = guide_row + 1
    for col, value in enumerate(['Concept', 'Formula / trace', 'Example value'], start=1):
        cell = settings.cell(guide_header, col)
        cell.value = value
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='7030A0')
    example_rows = [
        ('Catalyst ID row label', 'Processed row identifiers use the Catalyst ID entered in the GUI.', "='Processed'!F2"),
        ('Raw label carried over', "The final Processed column still points directly to Raw Original column A.", raw_label_example_ref or ''),
        ('Row inclusion', 'Processed column A says whether the row is included in the steady-state summary; highlighted rows remain visible but are not averaged.', "='Processed'!A2"),
        ('Time on stream', 'Processed!D2 = (accepted point - 1) * injection_interval_min / 60 when spacing is set.', "='Processed'!D2"),
        ('Bypass inlet flow', f'{reactant_label}_inlet_sccm comes from Bypass Processed average when bypass is used.', f'={reactant_inlet_cell}'),
        (f'{reactant_label} outlet flow', 'Processed!G2 references the calculated flow column for the reactant.', "='Processed'!G2"),
        (f'{reactant_label} conversion', f'({reactant_label} inlet - {reactant_label} outlet) / {reactant_label} inlet * 100.', "='Processed'!H2"),
        ('Product carbon out', 'Sum of carbon number times product flow for each product species.', "='Processed'!I2"),
        ('Carbon balance', f'({reactant_label} out + product carbon out) / {reactant_label} inlet * 100.', "='Processed'!J2"),
    ]
    if area_fallback_count:
        example_rows.insert(3, (
            'Area-only C5/C6 amount',
            'Derived amount = unknown peak area * editable Settings response factor; named C5/C6 peaks can use pentane/hexane response factors.',
            "='Area Fallbacks'!I5"))
    elif area_skipped_count:
        example_rows.insert(3, (
            'Area-only C5/C6 amount',
            'Area-only peaks were found, but no usable pentane/hexane reference response was available; see Area Fallbacks.',
            'not calculated'))
    if present_groups:
        example_rows.append((
            f'{present_groups[0]} selectivity',
            f'{present_groups[0]} carbon out / total product carbon out * 100.',
            "='Processed'!K2"))
    for concept, trace, example in example_rows:
        settings.append((concept, trace, example))

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.number_format = '0.000'
    for row in settings.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.number_format = '0.000'
    ws.freeze_panes = 'A2'
    _autosize_sheet(ws)
    _autosize_sheet(settings)
    try:
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(path)
    return path


def save_outputs(df, df_sel, total_C_out, C_in_flow,
                 reactant_label, ss_mask, metadata, species_config, output_dir,
                 source_filepath=None, raw_data=None, inlet_flows=None,
                 has_bridge=False, bypass_filepath=None, bypass_data=None):
    rxn    = df[~df['is_bypass']]
    ss_rxn = ss_mask[rxn.index]
    row    = dict(metadata)

    row['n_bypass']   = int(df['is_bypass'].sum())
    row['n_reaction'] = int((~df['is_bypass']).sum())
    if ss_rxn.any():
        row['ss_inj_start'] = int(rxn.loc[ss_rxn, 'inj_num'].min())
        row['ss_inj_end']   = int(rxn.loc[ss_rxn, 'inj_num'].max())

    if 'conversion' in rxn.columns and ss_rxn.any():
        row['conversion_%']     = round(rxn.loc[ss_rxn, 'conversion'].mean() * 100, 2)
        row['conversion_std_%'] = round(rxn.loc[ss_rxn, 'conversion'].std()  * 100, 3)

    if not df_sel.empty and ss_rxn.any():
        for col, val in (df_sel.loc[rxn.index][ss_rxn].mean() * 100).items():
            sp = col.replace('S_', '')
            row[f'sel_{sp}_%'] = round(val, 2)

    if C_in_flow > 0 and ss_rxn.any():
        row['carbon_balance_%'] = round(
            (total_C_out[rxn.index][ss_rxn].mean() / C_in_flow) * 100, 2)

    summary = pd.DataFrame([row])
    prefix = _output_file_prefix(metadata)
    summary_csv_path = os.path.join(output_dir, f'{prefix}_gc_summary.csv')
    flows_path = os.path.join(output_dir, f'{prefix}_gc_flows.csv')
    summary.to_csv(summary_csv_path, index=False)
    df_export = df.copy()
    if 'catalyst_id' not in df_export.columns:
        df_export.insert(0, 'catalyst_id', metadata.get('catalyst_id') or '')
    df_export.to_csv(flows_path, index=False)
    if source_filepath and raw_data is not None and inlet_flows is not None:
        summary_path = _write_gc_analysis_workbook(
            df, df_sel, total_C_out, C_in_flow, reactant_label, metadata,
            species_config, output_dir, source_filepath, raw_data, inlet_flows,
            has_bridge, bypass_filepath=bypass_filepath, bypass_data=bypass_data,
            ss_mask=ss_mask)
    else:
        summary_path = summary_csv_path
    return summary, summary_path, flows_path


# ─────────────────────────────────────────────────────────────────────────────
# MAIN ENTRY POINT (called by app.py)
# ─────────────────────────────────────────────────────────────────────────────

def run(filepath, output_dir, reaction_config, metadata, inlet_flows,
        ss_start, ss_end, bypass_filepath=None):
    """
    Main processing function called by the web app.
    Returns a dict with paths to output files and result summary.
    """
    os.makedirs(output_dir, exist_ok=True)

    species_config = {
        header: dict(cfg)
        for header, cfg in reaction_config['species'].items()
    }
    reactant_label = reaction_config['reactant']
    metadata = dict(metadata)
    metadata['reaction_type'] = reaction_config.get('type') or metadata.get('reaction_type') or reaction_config.get('name')
    if reaction_config.get('conversion_basis'):
        metadata['conversion_basis'] = reaction_config.get('conversion_basis')
    if reaction_config.get('selectivity_group_order'):
        metadata['selectivity_group_order'] = reaction_config.get('selectivity_group_order')
    for cn in (5, 6):
        key = _unknown_response_factor_key(cn)
        metadata[key] = _metadata_float(
            metadata, key, _DEFAULT_UNKNOWN_AREA_RESPONSE_FACTORS[cn])
    if 'bypass_points_used' not in metadata:
        metadata['bypass_points_used'] = 3
    if metadata.get('same_file_bypass_rows') in (None, ''):
        metadata['same_file_bypass_rows'] = metadata.get('bypass_points_used', 3)
    if not metadata.get('same_file_bypass_mode'):
        metadata['same_file_bypass_mode'] = 'auto'

    # Find Ar inlet flow
    F_Ar = 0
    for sp, flow in inlet_flows.items():
        for header, cfg in species_config.items():
            if cfg['label'] == sp and cfg['label'] == 'Ar':
                F_Ar = flow
    if F_Ar == 0:
        F_Ar = inlet_flows.get('Ar', 15.0)

    data = parse_xlsx(filepath)
    _mark_source_kind(data, 'main')
    if reaction_config.get('area_response_fallback', True) is not False:
        apply_area_response_fallbacks(data, species_config, metadata)
    if reaction_config.get('coeluted_splits', True) is not False:
        _augment_species_config_for_coeluted_splits(data, species_config)
    same_file_bypass_data = _apply_same_file_bypass_selection(
        data, metadata, reaction_config, species_config)
    inlet_source = 'manual'
    inferred_inlet_flows = {}
    bypass_data = None
    bypass_warning = None
    if bypass_filepath:
        bypass_data = parse_xlsx(bypass_filepath)
        _mark_source_kind(bypass_data, 'bypass_file')
        bypass_data['source'] = 'separate_file'
        bypass_data['source_sheet_name'] = 'Bypass Original'
        for inj in bypass_data['injections']:
            inj['is_bypass'] = True
        bypass_data = _select_bypass_data(bypass_data, metadata)
        metadata['bypass_total_points'] = bypass_data.get('bypass_total_points')
        metadata['bypass_omit_initial'] = bypass_data.get('bypass_omit_initial')
        metadata['bypass_points_used'] = bypass_data.get('bypass_points_requested')
        metadata['bypass_selected_points'] = bypass_data.get('bypass_selected_points')
        metadata['bypass_source'] = 'separate_file'
        inferred_inlet_flows, bypass_warning = infer_inlet_flows_from_bypass(
            bypass_data, F_Ar, reaction_config)
        if inferred_inlet_flows:
            inlet_flows = {**inlet_flows, **inferred_inlet_flows}
            inlet_source = 'bypass'
    elif same_file_bypass_data:
        bypass_data = _select_bypass_data(same_file_bypass_data, metadata)
        metadata['bypass_total_points'] = bypass_data.get('bypass_total_points')
        metadata['bypass_omit_initial'] = bypass_data.get('bypass_omit_initial')
        metadata['bypass_points_used'] = bypass_data.get('bypass_points_requested')
        metadata['bypass_selected_points'] = bypass_data.get('bypass_selected_points')
        metadata['bypass_source'] = 'same_file'
        inferred_inlet_flows, bypass_warning = infer_inlet_flows_from_bypass(
            bypass_data, F_Ar, reaction_config)
        if inferred_inlet_flows:
            inlet_flows = {**inlet_flows, **inferred_inlet_flows}
            inlet_source = 'bypass'

    if 'bypass_source' not in metadata:
        metadata['bypass_source'] = 'manual'
    metadata.setdefault('bypass_selected_points', 0)

    # Carbon in
    C_in_flow = 0.0
    for sp, flow in inlet_flows.items():
        for cfg in species_config.values():
            if cfg['label'] == sp:
                C_in_flow += cfg['cn'] * flow
                break

    metadata['inlet_flow_source'] = inlet_source
    if bypass_filepath:
        metadata['bypass_file'] = os.path.basename(bypass_filepath)
    elif bypass_data:
        metadata['bypass_file'] = 'same input file'
    for sp, flow in inlet_flows.items():
        metadata[f'inlet_{sp}_sccm'] = round(flow, 6)

    if bypass_filepath and bypass_data:
        existing_labels = {
            str(inj.get('label', '')).strip().lower()
            for inj in data['injections']
        }
        extra_bypass = [
            inj for inj in bypass_data['injections']
            if str(inj.get('label', '')).strip().lower() not in existing_labels
        ]
        data['injections'] = extra_bypass + data['injections']
    df, has_bridge = build_flow_table(data, F_Ar, species_config)
    df, plot_reaction_points = _add_time_on_stream_column(df, metadata)
    if plot_reaction_points:
        metadata['plot_reaction_points'] = plot_reaction_points
        metadata['run_duration_h'] = _infer_run_duration_h(metadata)
        metadata['injection_interval_min'] = _metadata_float(metadata, 'injection_interval_min')
        metadata['rejected_initial_injections'] = _metadata_int(metadata, 'rejected_initial_injections', 0)

    # Steady-state mask
    ss_mask = (
        (~df['is_bypass']) &
        _analysis_mask(df) &
        (df['inj_num'] >= ss_start) &
        (df['inj_num'] <= ss_end)
    )

    df, df_sel, total_C_out, carbon_cols = calculate_results(
        df, reactant_label, inlet_flows.get(reactant_label, 0),
        species_config, reaction_config=reaction_config)

    plot_path = make_plots(
        df, df_sel, total_C_out, C_in_flow,
        reactant_label, ss_mask, metadata, carbon_cols, species_config, output_dir)

    summary, summary_path, flows_path = save_outputs(
        df, df_sel, total_C_out, C_in_flow,
        reactant_label, ss_mask, metadata, species_config, output_dir,
        source_filepath=filepath, raw_data=data, inlet_flows=inlet_flows,
        has_bridge=has_bridge, bypass_filepath=bypass_filepath,
        bypass_data=bypass_data)

    # Build result dict for UI
    rxn    = df[~df['is_bypass']]
    ss_rxn = ss_mask[rxn.index]
    result = {
        'sequence_name':  data['sequence_name'],
        'n_bypass':       int(df['is_bypass'].sum()),
        'n_reaction':     int((~df['is_bypass']).sum()),
        'n_ss':           int(ss_rxn.sum()),
        'plot_reaction_points': int(plot_reaction_points),
        'run_duration_h':  _infer_run_duration_h(metadata),
        'injection_interval_min': _metadata_float(metadata, 'injection_interval_min'),
        'rejected_initial_injections': _metadata_int(metadata, 'rejected_initial_injections', 0),
        'plot_style':      metadata.get('plot_style', 'auto'),
        'area_derived_amounts': len(data.get('area_derived_amounts', [])),
        'area_derivation_skipped': len(data.get('area_derivation_skipped', [])),
        'c5_unknown_response_factor': metadata.get('c5_unknown_response_factor'),
        'c6_unknown_response_factor': metadata.get('c6_unknown_response_factor'),
        'bypass_omit_initial': _metadata_int(metadata, 'bypass_omit_initial', 0),
        'bypass_points_used': _metadata_int(metadata, 'bypass_points_used'),
        'bypass_selected_points': _metadata_int(metadata, 'bypass_selected_points'),
        'bypass_source': metadata.get('bypass_source'),
        'same_file_bypass_mode': metadata.get('same_file_bypass_mode'),
        'same_file_bypass_rows': _metadata_int(metadata, 'same_file_bypass_rows'),
        'same_file_bypass_detected': _metadata_int(metadata, 'same_file_bypass_detected', 0),
        'same_file_bypass_detect_method': metadata.get('same_file_bypass_detect_method'),
        'fid_bridge':     has_bridge,
        'plot_path':      plot_path,
        'summary_path':   summary_path,
        'flows_path':     flows_path,
        'output_dir':     output_dir,
        'inlet_source':   inlet_source,
        'inlet_flows':    {k: round(v, 6) for k, v in inlet_flows.items()},
        'inferred_inlet_flows': {
            k: round(v, 6) for k, v in inferred_inlet_flows.items()
        },
    }
    if bypass_filepath:
        result['bypass_file'] = os.path.basename(bypass_filepath)
    if bypass_warning:
        result['bypass_warning'] = bypass_warning
    if 'conversion' in rxn.columns and ss_rxn.any():
        result['conversion']     = round(rxn.loc[ss_rxn, 'conversion'].mean() * 100, 2)
        result['conversion_std'] = round(rxn.loc[ss_rxn, 'conversion'].std()  * 100, 3)
    if 'carbon_balance_%' in summary.columns:
        result['carbon_balance'] = round(summary['carbon_balance_%'].iloc[0], 2)

    sel_cols = [c for c in summary.columns if c.startswith('sel_') and summary[c].iloc[0] > 0.05]
    result['selectivities'] = {
        c.replace('sel_','').replace('_%',''): round(summary[c].iloc[0], 1)
        for c in sel_cols
    }

    return result
