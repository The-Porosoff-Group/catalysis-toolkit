"""BET / gas-isotherm processing for the Catalysis Data Toolkit.

The Micromeritics ASAP example supplied with the toolkit is a compact BIFF2
stream rather than an OLE-wrapped Excel workbook. The parser supports that
native export directly, ordinary XLSX workbooks, and delimited tables.
"""

from __future__ import annotations

import math
import os
import re
import struct
from datetime import datetime

import numpy as np


MODULE_INFO = {
    'name': 'BET / Isotherm',
    'description': 'Surface area, BET consistency, and pore metrics',
    'status': 'active',
    'icon': '📐',
}


_HEADER_FILL = '1F4E78'
_INPUT_FILL = 'FFF2CC'
_GOOD_FILL = 'E2F0D9'
_WARN_FILL = 'FCE4D6'
_N_A = 6.02214076e23


def _safe_token(value, default='BET'):
    text = re.sub(r'[^A-Za-z0-9_.-]+', '_', str(value or '')).strip('_.-')
    return text or default


def _to_float(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value or '').strip().replace(',', '')
    match = re.search(r'[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?', text)
    if not match:
        return None
    try:
        number = float(match.group(0))
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def _parse_biff2_cells(filepath):
    """Parse LABEL and NUMBER records from a direct BIFF2 instrument export."""
    with open(filepath, 'rb') as handle:
        payload = handle.read()
    if len(payload) < 4 or struct.unpack_from('<H', payload, 0)[0] != 0x0009:
        raise ValueError('The file is not a direct BIFF2 workbook.')

    cells = {}
    offset = 0
    while offset + 4 <= len(payload):
        opcode, length = struct.unpack_from('<HH', payload, offset)
        data_start = offset + 4
        data_end = data_start + length
        if data_end > len(payload):
            raise ValueError('The legacy XLS record stream is truncated.')
        if opcode in (0x0003, 0x0004) and length >= 8:
            row, column = struct.unpack_from('<HH', payload, data_start)
            if opcode == 0x0003 and length >= 15:
                cells[(row, column)] = struct.unpack_from('<d', payload, data_start + 7)[0]
            elif opcode == 0x0004:
                text_length = payload[data_start + 7]
                raw = payload[data_start + 8:data_start + 8 + text_length]
                cells[(row, column)] = raw.decode('cp1252', errors='replace')
        offset = data_end
    if not cells:
        raise ValueError('No cells were found in the legacy XLS export.')
    return [{'name': 'Micromeritics Export', 'cells': cells}]


def _parse_xlsx_cells(filepath):
    from openpyxl import load_workbook

    workbook = load_workbook(filepath, data_only=True, read_only=True)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            cells = {}
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cells[(cell.row - 1, cell.column - 1)] = cell.value
            sheets.append({'name': worksheet.title, 'cells': cells})
    finally:
        workbook.close()
    return sheets


def _parse_delimited(filepath):
    import pandas as pd

    frame = pd.read_csv(filepath, sep=None, engine='python', comment='#')
    if frame.shape[1] < 2:
        raise ValueError('A delimited BET file must contain at least two columns.')
    cells = {(0, col): str(label) for col, label in enumerate(frame.columns)}
    for row_index, values in enumerate(frame.itertuples(index=False), start=1):
        for col_index, value in enumerate(values):
            if value is not None and not (isinstance(value, float) and np.isnan(value)):
                cells[(row_index, col_index)] = value
    return [{'name': 'Delimited Data', 'cells': cells}]


def _load_sheets(filepath):
    extension = os.path.splitext(filepath)[1].lower()
    if extension in {'.csv', '.tsv', '.txt'}:
        return _parse_delimited(filepath)
    if extension == '.xlsx':
        return _parse_xlsx_cells(filepath)
    if extension == '.xls':
        try:
            return _parse_biff2_cells(filepath)
        except ValueError as direct_error:
            try:
                import xlrd
            except ImportError as exc:
                raise ValueError(
                    'This conventional .xls file requires the optional xlrd package. '
                    'The supplied Micromeritics BIFF export is supported directly.') from exc
            try:
                book = xlrd.open_workbook(filepath, formatting_info=False)
                sheets = []
                for worksheet in book.sheets():
                    cells = {}
                    for row in range(worksheet.nrows):
                        for column in range(worksheet.ncols):
                            value = worksheet.cell_value(row, column)
                            if value not in ('', None):
                                cells[(row, column)] = value
                    sheets.append({'name': worksheet.name, 'cells': cells})
                return sheets
            except Exception as exc:
                raise ValueError(f'Could not read the legacy XLS file: {direct_error}') from exc
    raise ValueError('BET accepts .xls, .xlsx, .csv, .tsv, or .txt files.')


def _cell_text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip())


def _find_label_value(sheets, label_pattern):
    pattern = re.compile(label_pattern, re.IGNORECASE)
    for sheet in sheets:
        cells = sheet['cells']
        for (row, column), value in cells.items():
            if pattern.search(_cell_text(value)):
                for offset in (1, 2):
                    candidate = cells.get((row, column + offset))
                    if candidate not in (None, ''):
                        return candidate
    return None


def _source_metrics(sheets):
    wanted = {
        'bet surface area': 'BET surface area',
        'langmuir surface area': 'Langmuir surface area',
        't-plot external surface area': 't-Plot external surface area',
        't-plot micropore volume': 't-Plot micropore volume',
        'micropore volume': 'Micropore volume',
        'micropore area': 'Micropore area',
        'external surface area': 'External surface area',
        'correlation coefficient': 'Source correlation coefficient',
        'molecular cross-sectional area': 'Molecular cross-sectional area',
        'single point surface area': 'Single-point surface area',
        'average particle size': 'Average particle size',
    }
    result = {}
    for sheet in sheets:
        cells = sheet['cells']
        for (row, column), value in cells.items():
            label = _cell_text(value).rstrip(':').lower()
            canonical = wanted.get(label)
            if canonical and canonical not in result:
                adjacent = cells.get((row, column + 1))
                if adjacent not in (None, '', ' '):
                    result[canonical] = adjacent
            pairs = (
                ('bjh adsorption cumulative surface area', 'BJH adsorption cumulative surface area'),
                ('bjh desorption cumulative surface area', 'BJH desorption cumulative surface area'),
                ('bjh adsorption cumulative volume', 'BJH adsorption cumulative pore volume'),
                ('bjh desorption cumulative volume', 'BJH desorption cumulative pore volume'),
            )
            for fragment, name in pairs:
                if fragment in label:
                    result.setdefault(name, cells.get((row + 1, column + 1)))
            if label.startswith('adsorption average pore width'):
                result.setdefault('Adsorption average pore width', cells.get((row, column + 1)))
            elif label.startswith('bjh adsorption average pore width'):
                result.setdefault('BJH adsorption average pore width', cells.get((row, column + 1)))
            elif label.startswith('bjh desorption average pore width'):
                result.setdefault('BJH desorption average pore width', cells.get((row, column + 1)))
            elif label.startswith('single point surface area'):
                result.setdefault('Single-point surface area', cells.get((row, column + 1)))
            elif label.startswith('single point adsorption total pore volume'):
                result.setdefault('Single-point total pore volume', cells.get((row + 1, column + 1)))
            if label == 'bet surface area report':
                for report_row in range(row + 1, row + 9):
                    report_label = _cell_text(cells.get((report_row, column))).rstrip(':').lower()
                    if report_label == 'slope':
                        result.setdefault('BET slope', cells.get((report_row, column + 1)))
                    elif report_label == 'y-intercept':
                        result.setdefault('BET y-intercept', cells.get((report_row, column + 1)))
    slope = _to_float(result.get('BET slope'))
    intercept = _to_float(result.get('BET y-intercept'))
    if slope is not None and intercept not in (None, 0):
        result['BET constant C'] = round(1.0 + slope / intercept, 6)

    cleaned = {}
    for key, value in result.items():
        if value in (None, ''):
            continue
        text = str(value)
        if '\ufffd' in text:
            if any(term in key.lower() for term in ('volume', 'slope', 'intercept')):
                replacement = '³'
            elif any(term in key.lower() for term in ('width', 'particle size')):
                replacement = 'Å'
            else:
                replacement = '²'
            text = text.replace('\ufffd', replacement)
        cleaned[key] = text if isinstance(value, str) else value
    return cleaned


def _extract_isotherm(sheets):
    for sheet in sheets:
        cells = sheet['cells']
        candidates = []
        for (row, column), value in cells.items():
            label = _cell_text(value).lower()
            if 'relative pressure' in label or label in {'p/p0', 'p/p°'}:
                candidates.append((row, column))
        for header_row, pressure_col in sorted(candidates):
            quantity_col = absolute_col = elapsed_col = None
            for column in range(pressure_col + 1, pressure_col + 6):
                label = _cell_text(cells.get((header_row, column))).lower()
                if 'quantity adsorbed' in label or 'volume adsorbed' in label:
                    quantity_col = column
                elif 'absolute pressure' in label:
                    absolute_col = column
                elif 'elapsed' in label or 'time' in label:
                    elapsed_col = column
            if quantity_col is None:
                continue
            points = []
            blank_run = 0
            max_row = max((row for row, _ in cells), default=header_row)
            for row in range(header_row + 1, max_row + 1):
                pressure = _to_float(cells.get((row, pressure_col)))
                quantity = _to_float(cells.get((row, quantity_col)))
                if pressure is None or quantity is None:
                    blank_run += 1
                    if blank_run >= 4 and points:
                        break
                    continue
                blank_run = 0
                if not (0 < pressure < 1.5):
                    continue
                points.append({
                    'relative_pressure': pressure,
                    'absolute_pressure_mmhg': _to_float(cells.get((row, absolute_col))) if absolute_col is not None else None,
                    'quantity_cm3_g_stp': quantity,
                    'elapsed_time': _cell_text(cells.get((row, elapsed_col))) if elapsed_col is not None else '',
                    'source_sheet': sheet['name'],
                    'source_row': row + 1,
                })
            if len(points) >= 5:
                turning = int(np.nanargmax([point['relative_pressure'] for point in points]))
                for index, point in enumerate(points):
                    point['branch'] = 'Adsorption' if index <= turning else 'Desorption'
                return points
    raise ValueError('Could not find an isotherm table with relative pressure and quantity adsorbed columns.')


def parse_bet_file(filepath):
    sheets = _load_sheets(filepath)
    return {
        'sample_id': _cell_text(_find_label_value(sheets, r'^sample\s*:?$')),
        'sample_mass_g': _to_float(_find_label_value(sheets, r'^sample\s+mass\s*:?$')),
        'adsorptive': _cell_text(_find_label_value(sheets, r'analysis\s+adsorptive')) or 'N2',
        'points': _extract_isotherm(sheets),
        'source_metrics': _source_metrics(sheets),
        'source_file': os.path.basename(filepath),
    }


def _linear_fit(x, y):
    x_mean = float(np.mean(x))
    y_mean = float(np.mean(y))
    denominator = float(np.sum((x - x_mean) ** 2))
    if denominator <= 0:
        raise ValueError('BET fitting points must span more than one pressure.')
    slope = float(np.sum((x - x_mean) * (y - y_mean)) / denominator)
    intercept = y_mean - slope * x_mean
    fitted = slope * x + intercept
    residual = y - fitted
    ss_res = float(np.sum(residual ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    return slope, intercept, 1.0 - ss_res / ss_tot if ss_tot > 0 else 1.0, fitted


def _fit_window(x, quantity, cross_section_nm2=0.162,
                molar_volume_cm3_mol=22414.0):
    transform = x / (quantity * (1.0 - x))
    slope, intercept, r_squared, fitted = _linear_fit(x, transform)
    denominator = slope + intercept
    monolayer_capacity = 1.0 / denominator if denominator > 0 else math.nan
    c_constant = 1.0 + slope / intercept if intercept != 0 else math.nan
    monolayer_pressure = 1.0 / (math.sqrt(c_constant) + 1.0) if c_constant > 0 else math.nan
    surface_area = (monolayer_capacity / molar_volume_cm3_mol * _N_A
                    * cross_section_nm2 * 1e-18) if monolayer_capacity > 0 else math.nan
    rouquerol_values = quantity * (1.0 - x)
    tolerance = max(float(np.ptp(rouquerol_values)) * 0.02, 1e-10)
    monotonic_fraction = float(np.mean(np.diff(rouquerol_values) >= -tolerance))
    interpolated_q = (float(np.interp(monolayer_pressure, x, quantity))
                      if math.isfinite(monolayer_pressure) and x[0] <= monolayer_pressure <= x[-1]
                      else math.nan)
    qm_deviation = (abs(interpolated_q - monolayer_capacity) / monolayer_capacity
                    if monolayer_capacity > 0 and math.isfinite(interpolated_q) else math.inf)
    return {
        'slope': slope, 'intercept': intercept, 'r_squared': r_squared,
        'fitted': fitted, 'transform': transform,
        'monolayer_capacity_cm3_g_stp': monolayer_capacity,
        'c_constant': c_constant, 'monolayer_pressure': monolayer_pressure,
        'surface_area_m2_g': surface_area,
        'rouquerol_monotonic_fraction': monotonic_fraction,
        'qm_deviation_fraction': qm_deviation,
    }


def recommend_bet_window(pressure, quantity, cross_section_nm2=0.162,
                         molar_volume_cm3_mol=22414.0):
    pressure = np.asarray(pressure, dtype=float)
    quantity = np.asarray(quantity, dtype=float)
    valid = np.isfinite(pressure) & np.isfinite(quantity) & (pressure > 0) & (pressure < 0.45) & (quantity > 0)
    x, q = pressure[valid], quantity[valid]
    if len(x) < 5:
        raise ValueError('At least five positive adsorption points below p/p0 = 0.45 are required.')
    order = np.argsort(x)
    x, q = x[order], q[order]
    candidates = []
    for start in range(0, len(x) - 4):
        for stop in range(start + 5, len(x) + 1):
            xs, qs = x[start:stop], q[start:stop]
            if xs[-1] - xs[0] < 0.045:
                continue
            try:
                fit = _fit_window(xs, qs, cross_section_nm2, molar_volume_cm3_mol)
            except (ValueError, np.linalg.LinAlgError, FloatingPointError):
                continue
            c_value = fit['c_constant']
            pmono = fit['monolayer_pressure']
            physical = (fit['slope'] > 0 and fit['intercept'] > 0
                        and math.isfinite(c_value) and c_value > 1
                        and math.isfinite(pmono) and xs[0] <= pmono <= xs[-1]
                        and fit['rouquerol_monotonic_fraction'] >= 0.9
                        and fit['qm_deviation_fraction'] <= 0.20)
            score = (fit['r_squared'] + min(len(xs), 15) * 0.0008
                     + min(xs[-1] - xs[0], 0.25) * 0.004
                     - min(fit['qm_deviation_fraction'], 1.0) * 0.02)
            candidates.append((physical, score, start, stop, fit))
    if not candidates:
        raise ValueError('No usable BET fitting window could be identified.')
    physical_candidates = [item for item in candidates if item[0]]
    chosen = max(physical_candidates or candidates, key=lambda item: item[1])
    _, _, start, stop, fit = chosen
    return float(x[start]), float(x[stop - 1]), fit, bool(physical_candidates)


def analyze_bet(parsed, p_min=None, p_max=None, cross_section_nm2=0.162,
                molar_volume_cm3_mol=22414.0, liquid_molar_volume_cm3_mol=34.65):
    if liquid_molar_volume_cm3_mol <= 0:
        raise ValueError('Adsorbate liquid molar volume must be greater than zero.')
    adsorption = [point for point in parsed['points'] if point['branch'] == 'Adsorption']
    pressure = np.asarray([point['relative_pressure'] for point in adsorption], dtype=float)
    quantity = np.asarray([point['quantity_cm3_g_stp'] for point in adsorption], dtype=float)
    recommended_min, recommended_max, _, recommendation_physical = recommend_bet_window(
        pressure, quantity, cross_section_nm2, molar_volume_cm3_mol)
    automatic = p_min is None or p_max is None
    used_min = recommended_min if p_min is None else float(p_min)
    used_max = recommended_max if p_max is None else float(p_max)
    if used_min < 0 or used_max <= used_min or used_max >= 1:
        raise ValueError('The BET p/p0 window must satisfy 0 <= minimum < maximum < 1.')
    included = (pressure >= used_min) & (pressure <= used_max) & (quantity > 0)
    if int(np.sum(included)) < 5:
        raise ValueError('The selected BET window contains fewer than five adsorption points.')
    fit = _fit_window(pressure[included], quantity[included], cross_section_nm2, molar_volume_cm3_mol)
    flags = []
    if fit['c_constant'] <= 0 or not math.isfinite(fit['c_constant']):
        flags.append('Unphysical BET constant (C <= 0); revise the p/p0 window.')
    elif fit['c_constant'] <= 1:
        flags.append('BET constant C is positive but <= 1; adsorption energetics may be nonphysical.')
    if fit['r_squared'] < 0.995:
        flags.append('BET linear-fit R² is below 0.995.')
    if fit['rouquerol_monotonic_fraction'] < 0.9:
        flags.append('Rouquerol criterion failed: V(1 - p/p0) is not consistently increasing.')
    if not (used_min <= fit['monolayer_pressure'] <= used_max):
        flags.append('Calculated monolayer pressure is outside the selected fit window.')
    if fit['qm_deviation_fraction'] > 0.20:
        flags.append('Calculated monolayer capacity differs from the isotherm value by more than 20%.')
    if not recommendation_physical:
        flags.append('No window passed every automated consistency criterion; inspect the fit manually.')
    source_micropore = parsed['source_metrics'].get('t-Plot micropore volume',
                                                    parsed['source_metrics'].get('Micropore volume'))
    if source_micropore is not None and (_to_float(source_micropore) or 0) < 0:
        flags.append('Source report contains a negative t-plot micropore volume.')
    highest_pressure_index = int(np.nanargmax(pressure))
    high_pressure_quantity = float(quantity[highest_pressure_index])
    total_pore_volume = (high_pressure_quantity * float(liquid_molar_volume_cm3_mol)
                         / float(molar_volume_cm3_mol))
    average_pore_diameter = (4000.0 * total_pore_volume / fit['surface_area_m2_g']
                             if fit['surface_area_m2_g'] > 0 else math.nan)
    return {
        **fit,
        'used_p_min': used_min, 'used_p_max': used_max,
        'recommended_p_min': recommended_min, 'recommended_p_max': recommended_max,
        'window_source': 'Automatic consistency recommendation' if automatic else 'User-selected',
        'n_points': int(np.sum(included)), 'flags': flags,
        'included_mask': included, 'adsorption_pressure': pressure,
        'adsorption_quantity': quantity, 'cross_section_nm2': float(cross_section_nm2),
        'molar_volume_cm3_mol': float(molar_volume_cm3_mol),
        'liquid_molar_volume_cm3_mol': float(liquid_molar_volume_cm3_mol),
        'total_pore_volume_cm3_g': total_pore_volume,
        'average_pore_diameter_nm': average_pore_diameter,
        'pore_volume_relative_pressure': float(pressure[highest_pressure_index]),
    }


def default_plot_settings(sample_id='BET sample'):
    return {
        'title': str(sample_id or 'BET isotherm'),
        'isotherm_x_label': 'Relative pressure (p/p₀)',
        'isotherm_y_label': 'Quantity adsorbed (cm³ STP/g)',
        'bet_x_label': 'Relative pressure (p/p₀)',
        'bet_y_label': '1 / [Q(p₀/p − 1)]',
        'tick_font_size': 11, 'axis_font_size': 13, 'title_font_size': 15,
        'legend_font_size': 10, 'line_width': 2.0, 'marker_size': 5.0,
        'png_dpi': 300, 'figure_width': 11.0, 'figure_height': 4.8,
        'x_axis_min': 0.0, 'x_axis_max': 1.0,
        'isotherm_y_min': None, 'isotherm_y_max': None,
        'bet_y_min': None, 'bet_y_max': None,
        'show_isotherm': True, 'show_bet_plot': True,
        'show_desorption': True, 'show_fit_window': True, 'show_grid': False,
        'adsorption_color': '#3282D2', 'desorption_color': '#E58B2A',
        'fit_color': '#D64545', 'window_color': '#3FAE62',
    }


def _hex_color(value, default):
    text = str(value or '').strip()
    if re.fullmatch(r'#[0-9A-Fa-f]{6}', text):
        return text.upper()
    if re.fullmatch(r'#[0-9A-Fa-f]{3}', text):
        return '#' + ''.join(character * 2 for character in text[1:]).upper()
    return default


def _optional_number(value):
    return None if value in (None, '') else _to_float(value)


def normalize_plot_settings(settings, sample_id='BET sample'):
    defaults = default_plot_settings(sample_id)
    settings = settings or {}
    normalized = dict(defaults)
    for key in ('title', 'isotherm_x_label', 'isotherm_y_label', 'bet_x_label', 'bet_y_label'):
        if key in settings:
            normalized[key] = str(settings[key])
    for key, low, high in (
        ('tick_font_size', 8, 30), ('axis_font_size', 9, 36),
        ('title_font_size', 10, 42), ('legend_font_size', 8, 28),
        ('line_width', 0.5, 6), ('marker_size', 1, 16),
        ('png_dpi', 72, 600), ('figure_width', 6, 20), ('figure_height', 3, 14)):
        value = _to_float(settings.get(key))
        if value is not None:
            normalized[key] = max(low, min(high, value))
    for key in ('x_axis_min', 'x_axis_max', 'isotherm_y_min', 'isotherm_y_max', 'bet_y_min', 'bet_y_max'):
        if key in settings:
            normalized[key] = _optional_number(settings.get(key))
    for key in ('show_isotherm', 'show_bet_plot', 'show_desorption', 'show_fit_window', 'show_grid'):
        value = settings.get(key, defaults[key])
        normalized[key] = value if isinstance(value, bool) else str(value).lower() in {'1', 'true', 'yes', 'on'}
    if not normalized['show_isotherm'] and not normalized['show_bet_plot']:
        normalized['show_isotherm'] = True
    for key in ('adsorption_color', 'desorption_color', 'fit_color', 'window_color'):
        normalized[key] = _hex_color(settings.get(key), defaults[key])
    validate_plot_axis_ranges(normalized)
    return normalized


def validate_plot_axis_ranges(settings):
    for label, low_key, high_key in (
        ('X-axis', 'x_axis_min', 'x_axis_max'),
        ('Isotherm Y-axis', 'isotherm_y_min', 'isotherm_y_max'),
        ('BET Y-axis', 'bet_y_min', 'bet_y_max')):
        low, high = _optional_number(settings.get(low_key)), _optional_number(settings.get(high_key))
        if low is not None and high is not None and high <= low:
            raise ValueError(f'{label} maximum must be greater than its minimum.')


def make_plot(parsed, analysis, output_dir, metadata, settings=None):
    from modules.characterization_plot import render_bet_plot

    settings = normalize_plot_settings(settings, metadata.get('sample_id'))
    adsorption = [point for point in parsed['points'] if point['branch'] == 'Adsorption']
    desorption = [point for point in parsed['points'] if point['branch'] == 'Desorption']
    ax = np.asarray([point['relative_pressure'] for point in adsorption], dtype=float)
    aq = np.asarray([point['quantity_cm3_g_stp'] for point in adsorption], dtype=float)
    os.makedirs(output_dir, exist_ok=True)
    prefix = _safe_token(metadata.get('output_prefix') or metadata.get('sample_id'), 'BET')
    path = os.path.join(output_dir, f'{prefix}_bet_plot.png')
    render_bet_plot(
        path, ax, aq,
        np.asarray([point['relative_pressure'] for point in desorption], dtype=float),
        np.asarray([point['quantity_cm3_g_stp'] for point in desorption], dtype=float),
        analysis['included_mask'], analysis['slope'], analysis['intercept'],
        analysis['r_squared'], settings)
    return path, settings


def _style_header(row):
    from openpyxl.styles import Alignment, Font, PatternFill
    for cell in row:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=_HEADER_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(worksheet, max_width=46):
    from openpyxl.utils import get_column_letter
    for column_index, column in enumerate(worksheet.columns, start=1):
        letter = get_column_letter(column_index)
        width = max((len(str(cell.value)) for cell in column if cell.value is not None), default=8)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_workbook(parsed, analysis, plot_path, output_dir, metadata, plot_settings=None):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font, PatternFill

    os.makedirs(output_dir, exist_ok=True)
    prefix = _safe_token(metadata.get('output_prefix') or metadata.get('sample_id'), 'BET')
    path = os.path.join(output_dir, f'{prefix}_bet_analysis.xlsx')
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary.sheet_view.showGridLines = False
    summary['A1'] = 'BET / Isotherm Analysis'
    summary['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    summary['A1'].fill = PatternFill('solid', fgColor=_HEADER_FILL)
    summary.merge_cells('A1:B1')
    for values in (
        ('Sample ID', metadata.get('sample_id') or parsed.get('sample_id')),
        ('Source file', parsed.get('source_file')), ('Adsorptive', parsed.get('adsorptive')),
        ('Sample mass (g)', metadata.get('sample_mass_g') or parsed.get('sample_mass_g')),
        ('Processed', datetime.now().isoformat(timespec='seconds')), ()):
        summary.append(values)
    summary.append(('Calculated metric', 'Value'))
    _style_header(summary[8])
    metric_rows = [
        ('BET surface area (m²/g)', analysis['surface_area_m2_g']),
        ('BET constant C', analysis['c_constant']),
        ('Monolayer capacity Qm (cm³ STP/g)', analysis['monolayer_capacity_cm3_g_stp']),
        ('Monolayer relative pressure', analysis['monolayer_pressure']),
        ('Slope', analysis['slope']), ('Y-intercept', analysis['intercept']),
        ('R²', analysis['r_squared']), ('Fit p/p₀ minimum', analysis['used_p_min']),
        ('Fit p/p₀ maximum', analysis['used_p_max']),
        ('Recommended p/p₀ minimum', analysis['recommended_p_min']),
        ('Recommended p/p₀ maximum', analysis['recommended_p_max']),
        ('Points in fit', analysis['n_points']), ('Window source', analysis['window_source']),
        ('Molecular cross-section (nm²)', analysis['cross_section_nm2']),
        ('STP molar volume (cm³/mol)', analysis['molar_volume_cm3_mol']),
        ('Total pore volume at highest p/p₀ (cm³/g)', analysis['total_pore_volume_cm3_g']),
        ('Average pore diameter, 4V/A (nm)', analysis['average_pore_diameter_nm']),
        ('Pore-volume relative pressure', analysis['pore_volume_relative_pressure']),
        ('Adsorbate liquid molar volume (cm³/mol)', analysis['liquid_molar_volume_cm3_mol']),
    ]
    metric_cell = {}
    for label, value in metric_rows:
        summary.append((label, value))
        metric_cell[label] = summary.max_row
    summary.append(())
    summary.append(('Quality flags', 'Details'))
    _style_header(summary[summary.max_row])
    if analysis['flags']:
        for index, flag in enumerate(analysis['flags'], start=1):
            summary.append((f'Flag {index}', flag))
            summary.cell(summary.max_row, 2).fill = PatternFill('solid', fgColor=_WARN_FILL)
    else:
        summary.append(('Status', 'All automated BET consistency checks passed.'))
        summary.cell(summary.max_row, 2).fill = PatternFill('solid', fgColor=_GOOD_FILL)
    summary.column_dimensions['A'].width = 39
    summary.column_dimensions['B'].width = 62
    for cell in summary['B']:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if plot_path and os.path.isfile(plot_path):
        image = Image(plot_path)
        image.width, image.height = 850, 370
        summary.add_image(image, 'D2')

    settings_sheet = workbook.create_sheet('Settings')
    settings_sheet.append(('Parameter', 'Value', 'Purpose'))
    _style_header(settings_sheet[1])
    settings_rows = [
        ('sample_id', metadata.get('sample_id'), 'User-entered sample identifier'),
        ('sample_mass_g', metadata.get('sample_mass_g') or parsed.get('sample_mass_g'), 'Parsed unless overridden'),
        ('bet_p_min', analysis['used_p_min'], 'Editable fit-window lower bound'),
        ('bet_p_max', analysis['used_p_max'], 'Editable fit-window upper bound'),
        ('cross_section_nm2', analysis['cross_section_nm2'], 'Adsorbate molecular cross-sectional area'),
        ('molar_volume_cm3_mol', analysis['molar_volume_cm3_mol'], 'STP molar volume used for surface area'),
        ('liquid_molar_volume_cm3_mol', analysis['liquid_molar_volume_cm3_mol'],
         'Condensed adsorbate molar volume used for total pore volume'),
    ]
    settings_rows.extend(
        (f'plot.{key}', value, 'Setting used for the embedded publication plot')
        for key, value in sorted((plot_settings or {}).items())
    )
    for values in settings_rows:
        settings_sheet.append(values)
    for cell in settings_sheet['B'][1:]:
        cell.fill = PatternFill('solid', fgColor=_INPUT_FILL)
    settings_sheet.freeze_panes = 'A2'
    _autosize(settings_sheet)

    isotherm = workbook.create_sheet('Isotherm')
    isotherm.append(('Point', 'Branch', 'Relative pressure (p/p₀)', 'Absolute pressure (mmHg)',
                     'Quantity adsorbed (cm³ STP/g)', 'Elapsed time', 'Source sheet', 'Source row'))
    _style_header(isotherm[1])
    for index, point in enumerate(parsed['points'], start=1):
        isotherm.append((index, point['branch'], point['relative_pressure'],
                         point.get('absolute_pressure_mmhg'), point['quantity_cm3_g_stp'],
                         point.get('elapsed_time'), point.get('source_sheet'), point.get('source_row')))
    isotherm.freeze_panes = 'A2'
    isotherm.auto_filter.ref = isotherm.dimensions
    _autosize(isotherm)

    bet_fit = workbook.create_sheet('BET Fit')
    bet_fit.append(('Relative pressure (p/p₀)', 'Quantity adsorbed (cm³ STP/g)',
                    'BET transform', 'Included', 'Linear fit', 'Residual'))
    _style_header(bet_fit[1])
    min_row, max_row = metric_cell['Fit p/p₀ minimum'], metric_cell['Fit p/p₀ maximum']
    slope_row, intercept_row = metric_cell['Slope'], metric_cell['Y-intercept']
    for index, (pressure, quantity) in enumerate(zip(
            analysis['adsorption_pressure'], analysis['adsorption_quantity']), start=2):
        bet_fit.cell(index, 1, float(pressure))
        bet_fit.cell(index, 2, float(quantity))
        bet_fit.cell(index, 3, f'=IFERROR(A{index}/(B{index}*(1-A{index})),"")')
        bet_fit.cell(index, 4, f'=AND(A{index}>=\'Summary\'!$B${min_row},A{index}<=\'Summary\'!$B${max_row})')
        bet_fit.cell(index, 5, f'=IF(D{index},\'Summary\'!$B${slope_row}*A{index}+\'Summary\'!$B${intercept_row},"")')
        bet_fit.cell(index, 6, f'=IF(D{index},C{index}-E{index},"")')
    bet_fit.freeze_panes = 'A2'
    bet_fit.auto_filter.ref = bet_fit.dimensions
    _autosize(bet_fit)

    source = workbook.create_sheet('Source Metrics')
    source.append(('Metric reported by instrument', 'Reported value'))
    _style_header(source[1])
    for label, value in parsed['source_metrics'].items():
        source.append((label, value))
    source.freeze_panes = 'A2'
    _autosize(source)

    notes = workbook.create_sheet('Analysis Notes')
    notes.append(('Topic', 'Details'))
    _style_header(notes[1])
    notes.append(('BET equation', 'p/[Q(p₀-p)] = 1/(Qm C) + [(C-1)/(Qm C)] (p/p₀)'))
    notes.append(('BET constant', 'C = 1 + slope/intercept. C <= 0 is physically invalid.'))
    notes.append(('Automatic window', 'Contiguous adsorption points are scored using positive C, monolayer-pressure inclusion, Rouquerol monotonicity, Qm agreement, and linearity.'))
    notes.append(('Calculated pore metrics', 'Total pore volume uses the highest-pressure adsorption uptake and the condensed adsorbate molar volume. Average pore diameter is 4V/A.'))
    notes.append(('Reported pore metrics', 'BJH, t-plot, Langmuir, micropore, and related source-report values are preserved on Source Metrics; they are labeled as reported rather than newly refitted.'))
    notes.append(('Raw-data integrity', 'Instrument measurements are preserved on Isotherm; analysis columns are kept separately on BET Fit.'))
    _autosize(notes, max_width=95)
    workbook.save(path)
    return path


def run(filepath, output_dir, metadata, params, plot_context=None):
    parsed = parse_bet_file(filepath)
    sample_id = str(metadata.get('sample_id') or parsed.get('sample_id') or 'BET Sample').strip()
    metadata = dict(metadata)
    metadata['sample_id'] = sample_id
    metadata['sample_mass_g'] = (_optional_number(metadata.get('sample_mass_g'))
                                 or parsed.get('sample_mass_g'))
    analysis = analyze_bet(
        parsed, p_min=_optional_number(params.get('p_min')),
        p_max=_optional_number(params.get('p_max')),
        cross_section_nm2=_optional_number(params.get('cross_section_nm2')) or 0.162,
        molar_volume_cm3_mol=_optional_number(params.get('molar_volume_cm3_mol')) or 22414.0,
        liquid_molar_volume_cm3_mol=(
            _optional_number(params.get('liquid_molar_volume_cm3_mol')) or 34.65))
    plot_path, settings = make_plot(parsed, analysis, output_dir, metadata, params.get('plot_settings'))
    workbook_path = write_workbook(parsed, analysis, plot_path, output_dir, metadata, settings)
    if plot_context is not None:
        plot_context.update({'parsed': parsed, 'analysis': analysis, 'metadata': metadata,
                             'output_dir': output_dir, 'workbook_path': workbook_path,
                             'plot_settings': settings})
    return {
        'sample_id': sample_id, 'sample_mass_g': metadata.get('sample_mass_g'),
        'adsorptive': parsed.get('adsorptive'),
        'surface_area_m2_g': round(analysis['surface_area_m2_g'], 6),
        'source_surface_area_m2_g': _to_float(parsed['source_metrics'].get('BET surface area')),
        'total_pore_volume_cm3_g': round(analysis['total_pore_volume_cm3_g'], 8),
        'average_pore_diameter_nm': round(analysis['average_pore_diameter_nm'], 6),
        'pore_volume_relative_pressure': round(analysis['pore_volume_relative_pressure'], 6),
        'c_constant': round(analysis['c_constant'], 6),
        'monolayer_capacity_cm3_g_stp': round(analysis['monolayer_capacity_cm3_g_stp'], 6),
        'monolayer_pressure': round(analysis['monolayer_pressure'], 6),
        'r_squared': round(analysis['r_squared'], 8),
        'used_p_min': round(analysis['used_p_min'], 6), 'used_p_max': round(analysis['used_p_max'], 6),
        'recommended_p_min': round(analysis['recommended_p_min'], 6),
        'recommended_p_max': round(analysis['recommended_p_max'], 6),
        'window_source': analysis['window_source'], 'n_points': analysis['n_points'],
        'flags': analysis['flags'],
        'source_metrics': {key: str(value) for key, value in parsed['source_metrics'].items()},
        'plot_path': plot_path, 'summary_path': workbook_path,
        'output_dir': output_dir, 'plot_settings': settings,
    }


def regenerate_plot(context, settings):
    plot_path, normalized = make_plot(context['parsed'], context['analysis'],
                                      context['output_dir'], context['metadata'], settings)
    workbook_path = write_workbook(context['parsed'], context['analysis'], plot_path,
                                   context['output_dir'], context['metadata'], normalized)
    context['plot_settings'] = normalized
    context['workbook_path'] = workbook_path
    return {'plot_path': plot_path, 'summary_path': workbook_path,
            'plot_settings': normalized}
