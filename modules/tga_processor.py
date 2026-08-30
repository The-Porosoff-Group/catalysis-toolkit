"""AutoChem TPD/TPO/TPR and pulse-chemisorption analysis."""

from __future__ import annotations

import math
import os
import re
from datetime import datetime

import numpy as np
from scipy.signal import find_peaks, peak_widths


MODULE_INFO = {
    'name': 'TPD / TPO / TPR / Chemisorption',
    'description': 'Temperature-programmed and pulse-chemisorption analysis',
    'status': 'active',
    'icon': '🌡️',
}


METAL_PRESETS = {
    'Ni': {'name': 'Nickel', 'atomic_weight_g_mol': 58.6934, 'density_g_cm3': 8.908,
           'cross_section_nm2': 0.0649, 'stoichiometry_metal_per_gas': 1.0},
    'Co': {'name': 'Cobalt', 'atomic_weight_g_mol': 58.9332, 'density_g_cm3': 8.90,
           'cross_section_nm2': 0.0662, 'stoichiometry_metal_per_gas': 1.0},
    'Pt': {'name': 'Platinum', 'atomic_weight_g_mol': 195.084, 'density_g_cm3': 21.45,
           'cross_section_nm2': 0.0800, 'stoichiometry_metal_per_gas': 1.0},
    'Pd': {'name': 'Palladium', 'atomic_weight_g_mol': 106.42, 'density_g_cm3': 12.023,
           'cross_section_nm2': 0.0787, 'stoichiometry_metal_per_gas': 1.0},
    'W': {'name': 'Tungsten', 'atomic_weight_g_mol': 183.84, 'density_g_cm3': 19.25,
          'cross_section_nm2': 0.0741, 'stoichiometry_metal_per_gas': 1.0},
    'Mo': {'name': 'Molybdenum', 'atomic_weight_g_mol': 95.95, 'density_g_cm3': 10.28,
           'cross_section_nm2': 0.0639, 'stoichiometry_metal_per_gas': 1.0},
    'Fe': {'name': 'Iron', 'atomic_weight_g_mol': 55.845, 'density_g_cm3': 7.874,
           'cross_section_nm2': 0.0636, 'stoichiometry_metal_per_gas': 1.0},
    'Zn': {'name': 'Zinc', 'atomic_weight_g_mol': 65.38, 'density_g_cm3': 7.14,
           'cross_section_nm2': 0.0710, 'stoichiometry_metal_per_gas': 1.0},
    'Mn': {'name': 'Manganese', 'atomic_weight_g_mol': 54.938044, 'density_g_cm3': 7.21,
           'cross_section_nm2': 0.0650, 'stoichiometry_metal_per_gas': 1.0},
    'Cu': {'name': 'Copper', 'atomic_weight_g_mol': 63.546, 'density_g_cm3': 8.96,
           'cross_section_nm2': 0.0680, 'stoichiometry_metal_per_gas': 1.0},
}


_HEADER_FILL = '1F4E78'
_INPUT_FILL = 'FFF2CC'
_WARN_FILL = 'FCE4D6'
_GOOD_FILL = 'E2F0D9'
_R_GAS_ML_ATM = 82.057366080960
_N_A = 6.02214076e23


def _safe_token(value, default='TP'):
    text = re.sub(r'[^A-Za-z0-9_.-]+', '_', str(value or '')).strip('_.-')
    return text or default


def _to_float(value):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = str(value or '').replace(',', '').strip()
    match = re.search(r'[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?', text)
    if not match:
        return None
    try:
        number = float(match.group(0))
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def _read_text(filepath):
    with open(filepath, 'rb') as handle:
        payload = handle.read()
    if payload.startswith((b'\xff\xfe', b'\xfe\xff')):
        return payload.decode('utf-16')
    if payload[:4096].count(b'\x00') > 100:
        return payload.decode('utf-16-le')
    for encoding in ('utf-8-sig', 'cp1252', 'latin-1'):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError('Could not decode the AutoChem text export.')


def _normalize_name(value):
    return re.sub(r'[^a-z0-9]+', '', str(value or '').lower())


def _classify_experiment(name, analysis_type=''):
    text = f'{name} {analysis_type}'.lower()
    if 'chemi' in text or 'pulse' in text:
        return 'Chemisorption'
    if 'desorption' in text or re.search(r'\btpd\b', text):
        return 'TPD'
    if 'oxidation' in text or re.search(r'\btpo\b', text):
        return 'TPO'
    if 'reduction' in text or re.search(r'\btpr\b', text):
        return 'TPR'
    return 'Temperature Programmed'


def _infer_gas(name):
    text = re.sub(r'[_\-/]+', ' ', str(name or '').upper())
    for gas in ('CO2', 'N2O', 'NH3', 'H2', 'O2', 'CO'):
        if re.search(rf'(?<![A-Z0-9]){gas}(?![A-Z0-9])', text):
            return gas
    return ''


def _display_experiment_name(name, experiment_type='', gas=''):
    """Return a clean human-facing selector label for an instrument experiment."""
    text = re.sub(r'[_]+', ' ', str(name or ''))
    text = re.sub(r'\bchemi\s*-?\s*sorption\b', 'Chemisorption', text, flags=re.IGNORECASE)
    text = re.sub(r'[-\s]+', ' ', text).strip()
    if experiment_type == 'Chemisorption':
        return f'{gas} Pulsed Chemisorption'.strip()
    if experiment_type in {'TPD', 'TPO', 'TPR'}:
        return f'{gas} {experiment_type}'.strip()
    return text or experiment_type or 'Experiment'


def _reported_pulse_tables(text):
    """Parse instrument-reported pulse counts and uptake rows when included."""
    tables = {}
    starts = [match.start() for match in re.finditer(r'Pulse\s+Chemisorption\s+Report', text, re.IGNORECASE)]
    for start in starts:
        end_match = re.search(r'(?:\n\s*){2,}.*?Analysis\s+Summary', text[start:], re.IGNORECASE)
        end = start + end_match.end() + 3000 if end_match else min(len(text), start + 15000)
        block = text[start:end]
        experiment_match = re.search(r'Experiment\s+(\d+)\s*:?', block, re.IGNORECASE)
        if not experiment_match or 'Peak Table' not in block:
            continue
        number = int(experiment_match.group(1))
        expected_match = re.search(
            r'Number\s+of\s+(?:injections|pulses|doses)\s*:\s*(\d+)', block, re.IGNORECASE)
        rows = []
        table_text = block.split('Peak Table', 1)[1]
        for line in table_text.splitlines():
            match = re.match(
                r'^\s*(\d+)\s+([-+0-9.,Ee]+)\s+([-+0-9.,Ee]+)\s+([-+0-9.,Ee]+)\s*$', line)
            if not match:
                if rows and not line.strip():
                    break
                continue
            rows.append({
                'number': int(match.group(1)),
                'temperature_c': _to_float(match.group(2)),
                'reported_uptake_umol_g': _to_float(match.group(3)),
                'reported_cumulative_uptake_umol_g': _to_float(match.group(4)),
            })
        if rows:
            tables[number] = {
                'pulses': rows,
                'reported_peak_count': max(row['number'] for row in rows),
                'expected_injection_count': (int(expected_match.group(1))
                                             if expected_match else max(row['number'] for row in rows)),
            }
    return tables


def _numeric_pair(line):
    parts = re.split(r'\s+', line.strip())
    if len(parts) < 2:
        return None
    first, second = _to_float(parts[0]), _to_float(parts[1])
    return (first, second) if first is not None and second is not None else None


def _parse_table(lines, start):
    values = []
    index = start
    while index < len(lines):
        pair = _numeric_pair(lines[index])
        if pair is None:
            break
        values.append(pair)
        index += 1
    return np.asarray(values, dtype=float), index


def parse_autochem_txt(filepath):
    text = _read_text(filepath)
    lines = text.splitlines()
    pulse_tables = _reported_pulse_tables(text)
    metadata = {'source_file': os.path.basename(filepath)}
    sample_match = re.search(r'^\s*Sample:\s*(.*?)\s*$', text, re.MULTILINE)
    if sample_match:
        metadata['sample_id'] = sample_match.group(1).strip()
    mass_match = re.search(r'Sample\s+mass:\s*([0-9.,]+)\s*g', text, re.IGNORECASE)
    if mass_match:
        metadata['sample_mass_g'] = _to_float(mass_match.group(1))
    loop_match = re.search(
        r'Active\s+loop\s+volume\s+at\s+([0-9.,]+)\s*[^:]*:\s*([0-9.,]+)\s*[\u00b5u]mol',
        text, re.IGNORECASE)
    if loop_match:
        metadata['reported_loop_temperature_c'] = _to_float(loop_match.group(1))
        metadata['reported_active_loop_umol'] = _to_float(loop_match.group(2))

    experiments = []
    by_number = {}
    current = None
    raw_start = next((idx for idx, line in enumerate(lines)
                      if line.strip() == 'Signal (a.u.) vs. Time'), len(lines))
    for line in lines[:raw_start]:
        match = re.match(r'^\s*Experiment\s+(\d+)\s*:?\s+(.+?)\s*$', line)
        if match:
            number, name = int(match.group(1)), match.group(2).strip()
            current = by_number.get(number)
            if current is None:
                current = {'number': number, 'name': name, 'analysis_type': '',
                           'calibration': '', 'measured_flow_umol_min': None,
                           'signal_offset': None, 'signal_inverted': None, 'series': {}}
                by_number[number] = current
                experiments.append(current)
            elif len(name) > len(current['name']):
                current['name'] = name
            continue
        if current is None:
            continue
        field_match = re.match(r'^\s*Analysis\s+type:\s*(.+?)\s*$', line, re.IGNORECASE)
        if field_match:
            current['analysis_type'] = field_match.group(1).strip()
            continue
        field_match = re.match(r'^\s*Calibration:\s*(.+?)\s*$', line, re.IGNORECASE)
        if field_match:
            current['calibration'] = field_match.group(1).strip()
            continue
        field_match = re.match(r'^\s*Measured\s+flow\s+rate:\s*(.+?)\s*$', line, re.IGNORECASE)
        if field_match:
            current['measured_flow_umol_min'] = _to_float(field_match.group(1))
            continue
        field_match = re.match(r'^\s*Signal\s+offset:\s*(.+?)\s*$', line, re.IGNORECASE)
        if field_match:
            current['signal_offset'] = _to_float(field_match.group(1))
            continue
        field_match = re.match(r'^\s*Signal\s+inverted:\s*(.+?)\s*$', line, re.IGNORECASE)
        if field_match:
            current['signal_inverted'] = field_match.group(1).strip().lower() == 'yes'

    def experiment_for_label(label):
        normalized = _normalize_name(label)
        best = None
        for experiment in experiments:
            candidate = _normalize_name(experiment['name'])
            if candidate == normalized:
                return experiment
            if candidate and (candidate in normalized or normalized in candidate):
                best = experiment
        if best is not None:
            return best
        experiment = {'number': len(experiments) + 1, 'name': label,
                      'analysis_type': '', 'calibration': '',
                      'measured_flow_umol_min': None, 'signal_offset': None,
                      'signal_inverted': None, 'series': {}}
        experiments.append(experiment)
        return experiment

    index = raw_start
    while index < len(lines) - 1:
        stripped = lines[index].strip()
        signal_match = re.match(r'^Signal \(a\.u\.\)\s*-\s*(.+)$', stripped)
        temperature_match = re.match(r'^Temperature\s*-\s*(.+)$', stripped)
        if not signal_match and not temperature_match:
            index += 1
            continue
        label = (signal_match or temperature_match).group(1).strip()
        header_index = index + 1
        while header_index < len(lines) and not lines[header_index].strip():
            header_index += 1
        header = lines[header_index].strip().lower() if header_index < len(lines) else ''
        values, next_index = _parse_table(lines, header_index + 1)
        if len(values):
            experiment = experiment_for_label(label)
            if signal_match and header.startswith('time'):
                experiment['series']['signal_time'] = values
            elif signal_match and header.startswith('temperature'):
                experiment['series']['signal_temperature'] = values
            elif temperature_match and header.startswith('time'):
                experiment['series']['temperature_time'] = values
        index = max(next_index, index + 1)

    usable = []
    for experiment in experiments:
        experiment['type'] = _classify_experiment(experiment['name'], experiment['analysis_type'])
        experiment['gas'] = _infer_gas(experiment['name'])
        pulse_report = pulse_tables.get(experiment['number'], {})
        experiment['reported_pulses'] = pulse_report.get('pulses', [])
        experiment['reported_peak_count'] = pulse_report.get('reported_peak_count')
        experiment['expected_injection_count'] = pulse_report.get('expected_injection_count')
        experiment['display_name'] = _display_experiment_name(
            experiment['name'], experiment['type'], experiment['gas'])
        experiment['key'] = f"exp{experiment['number']}_{experiment['type'].lower().replace(' ', '_')}"
        if experiment['series'].get('signal_time') is not None:
            usable.append(experiment)
    if not usable:
        raise ValueError('No signal-versus-time traces were found in the AutoChem export.')
    return {'metadata': metadata, 'experiments': usable}


def _odd_window(requested, length):
    if length < 5:
        return 1
    window = int(_to_float(requested) or 101)
    window = max(5, min(window, length - (1 if length % 2 == 0 else 0)))
    return window if window % 2 else window - 1


def _smooth(values, requested_window):
    window = _odd_window(requested_window, len(values))
    if window < 5:
        return values.copy()
    pad = window // 2
    padded = np.pad(values, pad, mode='edge')
    kernel = np.full(window, 1.0 / window)
    return np.convolve(padded, kernel, mode='valid')


def _baseline(x, signal, method='auto'):
    if method == 'none':
        return np.zeros_like(signal)
    edge = max(3, len(signal) // 50)
    if method == 'linear':
        return np.linspace(float(np.median(signal[:edge])), float(np.median(signal[-edge:])), len(signal))
    anchors_x, anchors_y = [], []
    for indices in np.array_split(np.arange(len(signal)), min(50, max(8, len(signal) // 200))):
        if len(indices):
            anchors_x.append(float(np.median(x[indices])))
            anchors_y.append(float(np.percentile(signal[indices], 10)))
    if len(anchors_x) >= 2:
        return np.interp(x, np.asarray(anchors_x), np.asarray(anchors_y))
    return np.linspace(float(np.median(signal[:edge])), float(np.median(signal[-edge:])), len(signal))


def _orientation(corrected, requested='auto'):
    requested = str(requested or 'auto').lower()
    if requested in {'positive', '+', '1'}:
        return 1.0
    if requested in {'negative', '-', '-1'}:
        return -1.0
    positive = float(np.percentile(corrected, 99))
    negative = abs(float(np.percentile(corrected, 1)))
    return 1.0 if positive >= negative else -1.0


def _prepare_trace(experiment, params):
    signal_time = experiment['series']['signal_time']
    time = signal_time[:, 0].astype(float)
    raw = signal_time[:, 1].astype(float)
    temperature = np.full(len(time), np.nan)
    temperature_time = experiment['series'].get('temperature_time')
    if temperature_time is not None and len(temperature_time) >= 2:
        order = np.argsort(temperature_time[:, 0])
        temperature = np.interp(time, temperature_time[order, 0], temperature_time[order, 1])
    elif experiment['type'] != 'Chemisorption':
        ramp_rate = _to_float(params.get('ramp_rate_c_min'))
        if ramp_rate is not None:
            ramp_start = _to_float(params.get('ramp_start_temperature_c'))
            ramp_start = 25.0 if ramp_start is None else ramp_start
            temperature = ramp_start + ramp_rate * (time - time[0])
    valid = np.isfinite(time) & np.isfinite(raw)
    if experiment['type'] != 'Chemisorption' and np.isfinite(temperature).any():
        end = int(np.nanargmax(temperature)) + 1
        if end >= 20:
            valid[end:] = False
        minimum = _to_float(params.get('temperature_min_c'))
        maximum = _to_float(params.get('temperature_max_c'))
        if minimum is not None:
            valid &= temperature >= minimum
        if maximum is not None:
            valid &= temperature <= maximum
    time, raw, temperature = time[valid], raw[valid], temperature[valid]
    if len(time) < 10:
        raise ValueError(f"Experiment {experiment['name']} does not contain enough usable data points.")
    smoothed = _smooth(raw, params.get('smoothing_window'))
    x_for_baseline = time if experiment['type'] == 'Chemisorption' or not np.isfinite(temperature).any() else temperature
    baseline = _baseline(x_for_baseline, smoothed, str(params.get('baseline_method') or 'auto').lower())
    orientation = _orientation(smoothed - baseline, params.get('signal_direction'))
    finite_temperature = np.isfinite(temperature)
    parsed_ramp_rate = None
    ramp_start_temperature = None
    if experiment['type'] != 'Chemisorption' and np.sum(finite_temperature) >= 3:
        finite_time = time[finite_temperature]
        finite_temp = temperature[finite_temperature]
        dt = np.diff(finite_time)
        slopes = np.divide(np.diff(finite_temp), dt, out=np.full_like(dt, np.nan), where=np.abs(dt) > 1e-12)
        ramping = slopes[np.isfinite(slopes) & (np.abs(slopes) > 0.05)]
        if len(ramping):
            parsed_ramp_rate = float(np.median(ramping))
        ramp_start_temperature = float(finite_temp[0])
    return {
        'time_min': time, 'temperature_c': temperature,
        'raw_signal': raw * orientation, 'smoothed_signal': smoothed * orientation,
        'baseline': baseline * orientation, 'corrected_signal': (smoothed - baseline) * orientation,
        'orientation': 'positive' if orientation > 0 else 'negative (flipped)',
        'experiment_type': experiment['type'],
        'x': time if experiment['type'] == 'Chemisorption' or not np.isfinite(temperature).any() else temperature,
        'x_unit': 'min' if experiment['type'] == 'Chemisorption' or not np.isfinite(temperature).any() else '°C',
        'ramp_rate_c_min': parsed_ramp_rate,
        'ramp_start_temperature_c': ramp_start_temperature,
    }


def _detect_peaks(trace, params):
    corrected = trace['corrected_signal']
    x = trace['x']
    amplitude = max(float(np.percentile(corrected, 99) - np.percentile(corrected, 10)), 1e-12)
    prominence = amplitude * max(0.0001, (_to_float(params.get('prominence_percent')) or 3.0) / 100.0)
    x_diffs = np.abs(np.diff(x))
    positive_steps = x_diffs[x_diffs > 1e-9]
    if trace['x_unit'] == '°C' and len(x) > 1:
        step = max(float(np.nanmax(x) - np.nanmin(x)) / (len(x) - 1), 1e-9)
    else:
        step = float(np.median(positive_steps)) if len(positive_steps) else 1.0
    if trace['x_unit'] == 'min':
        distance_units = _to_float(params.get('minimum_pulse_distance_min')) or 0.5
    else:
        distance_units = _to_float(params.get('minimum_peak_distance_c')) or 25.0
    distance_points = max(1, int(distance_units / step))
    peaks, properties = find_peaks(corrected, prominence=prominence, distance=distance_points)
    if not len(peaks):
        return [], prominence
    widths = peak_widths(corrected, peaks, rel_height=0.50 if trace['x_unit'] == 'min' else 0.95)
    results = []
    for peak, left, right, peak_prominence in zip(
            peaks, widths[2], widths[3], properties['prominences']):
        left_index = max(0, int(math.floor(left)))
        right_index = min(len(corrected) - 1, int(math.ceil(right)))
        if right_index <= left_index:
            continue
        if trace['x_unit'] == '°C' and float(x[peak]) <= float(np.nanmin(x)) + 5.0:
            continue
        if (trace.get('experiment_type') == 'TPD'
                and float(x[right_index]) >= float(np.nanmax(x)) - 1.0):
            # A peak whose half-width runs into the end of the temperature ramp
            # is incomplete and is normally an end-of-program artifact.
            continue
        area = float(np.trapezoid(np.clip(corrected[left_index:right_index + 1], 0, None),
                                  trace['time_min'][left_index:right_index + 1]))
        results.append({
            'number': len(results) + 1, 'index': int(peak), 'left_index': left_index,
            'right_index': right_index, 'x_at_max': float(x[peak]),
            'temperature_c': float(trace['temperature_c'][peak]) if np.isfinite(trace['temperature_c'][peak]) else None,
            'time_min': float(trace['time_min'][peak]),
            'height': float(corrected[peak]), 'prominence': float(peak_prominence),
            'area_signal_min': area,
            'left_x': float(x[left_index]), 'right_x': float(x[right_index]),
        })
    return results, prominence


def _limit_tpd_peaks(experiment_type, peaks, params):
    """Retain the most prominent TPD peaks, then restore temperature order."""
    if experiment_type != 'TPD':
        return peaks
    requested = int(_to_float(params.get('maximum_tpd_peaks')) or 6)
    maximum = max(1, min(6, requested))
    retained = sorted(peaks, key=lambda peak: peak.get('prominence', 0.0), reverse=True)[:maximum]
    retained.sort(key=lambda peak: (
        peak.get('temperature_c') is None,
        peak.get('temperature_c') if peak.get('temperature_c') is not None else peak.get('time_min', 0.0)))
    for number, peak in enumerate(retained, start=1):
        peak['number'] = number
    return retained


def _pulse_dose_umol(params, parsed_metadata):
    loop_volume_ml = _to_float(params.get('loop_volume_ml')) or 0.51548
    concentration = _to_float(params.get('active_gas_percent'))
    concentration = 10.0 if concentration is None else concentration
    pressure_atm = _to_float(params.get('loop_pressure_atm')) or 1.0
    loop_temp = _to_float(params.get('loop_temperature_c'))
    if loop_temp is None:
        loop_temp = parsed_metadata.get('reported_loop_temperature_c')
    if loop_temp is None:
        loop_temp = 25.0
    if not (0 < concentration <= 100):
        raise ValueError('Active gas concentration must be greater than 0 and no more than 100%.')
    kelvin = loop_temp + 273.15
    dose = pressure_atm * loop_volume_ml / (_R_GAS_ML_ATM * kelvin) * 1e6 * concentration / 100.0
    return dose, loop_volume_ml, concentration, pressure_atm, loop_temp


def _metal_metrics(uptake_umol_g, params):
    preset_key = str(params.get('metal') or '').strip()
    preset = METAL_PRESETS.get(preset_key, {})
    loading = _to_float(params.get('metal_loading_wt_percent'))
    atomic_weight = _to_float(params.get('atomic_weight_g_mol')) or preset.get('atomic_weight_g_mol')
    density = _to_float(params.get('density_g_cm3')) or preset.get('density_g_cm3')
    cross_section = _to_float(params.get('metal_cross_section_nm2')) or preset.get('cross_section_nm2')
    stoichiometry = _to_float(params.get('stoichiometry_metal_per_gas')) or preset.get('stoichiometry_metal_per_gas', 1.0)
    result = {
        'metal': preset_key, 'metal_loading_wt_percent': loading,
        'atomic_weight_g_mol': atomic_weight, 'density_g_cm3': density,
        'cross_section_nm2': cross_section,
        'stoichiometry_metal_per_gas': stoichiometry,
        'dispersion_percent': None, 'metal_surface_area_m2_g_sample': None,
        'metal_surface_area_m2_g_metal': None, 'particle_diameter_nm': None,
        'cubic_crystallite_size_nm': None,
    }
    if uptake_umol_g is None or cross_section is None or stoichiometry is None:
        return result
    area_sample = uptake_umol_g * stoichiometry * _N_A * 1e-24 * cross_section
    result['metal_surface_area_m2_g_sample'] = area_sample
    if loading and loading > 0:
        loading_fraction = loading / 100.0
        result['metal_surface_area_m2_g_metal'] = area_sample / loading_fraction
        if atomic_weight:
            metal_umol_g = loading_fraction / atomic_weight * 1e6
            result['dispersion_percent'] = uptake_umol_g * stoichiometry / metal_umol_g * 100.0
        if density and result['metal_surface_area_m2_g_metal'] > 0:
            result['particle_diameter_nm'] = 6000.0 / (density * result['metal_surface_area_m2_g_metal'])
            result['cubic_crystallite_size_nm'] = 5000.0 / (density * result['metal_surface_area_m2_g_metal'])
    return result


def _analyze_chemisorption(experiment, trace, peaks, params, parsed_metadata, sample_mass_g):
    dose, loop_volume, concentration, pressure, loop_temp = _pulse_dose_umol(params, parsed_metadata)
    detected_peaks = list(peaks)
    saturated_count = max(1, int(_to_float(params.get('saturated_peak_count')) or 3))
    saturated = detected_peaks[-min(saturated_count, len(detected_peaks)):] if detected_peaks else []
    reference_area = float(np.median([peak['area_signal_min'] for peak in saturated])) if saturated else None
    manual_factor = _to_float(params.get('manual_response_factor_umol_per_area'))
    response_factor = manual_factor or (dose / reference_area if reference_area and reference_area > 0 else None)
    first_area = (float(np.median([peak['area_signal_min']
                                  for peak in detected_peaks[:min(3, len(detected_peaks))]]))
                  if detected_peaks else None)
    requested_direction = str(params.get('chemisorption_direction') or 'auto').strip().lower()
    if requested_direction in {'normal', 'inverse'}:
        direction = requested_direction
    else:
        direction = 'inverse' if first_area is not None and reference_area is not None and first_area > reference_area * 1.03 else 'normal'
    requested_injections = int(_to_float(params.get('expected_injection_count')) or 0)
    reported_injections = int(experiment.get('expected_injection_count') or 0)
    expected_injections = (max(requested_injections, len(detected_peaks)) if requested_injections
                           else max(reported_injections, len(detected_peaks)))
    injection_source = ('User input' if requested_injections
                        else ('Instrument pulse table' if reported_injections else 'Detected peaks'))
    missing_count = max(0, expected_injections - len(detected_peaks))
    synthetic = []
    if missing_count:
        detected_times = np.asarray([peak['time_min'] for peak in detected_peaks], dtype=float)
        interval = (float(np.median(np.diff(detected_times))) if len(detected_times) >= 2
                    else max(float(trace['time_min'][-1] - trace['time_min'][0])
                             / max(expected_injections, 1), 0.5))
        first_time = float(detected_times[0]) if len(detected_times) else float(trace['time_min'][-1])
        for index in range(missing_count):
            pulse_time = max(float(trace['time_min'][0]), first_time - interval * (missing_count - index))
            synthetic.append({
                'number': index + 1, 'index': None, 'left_index': None, 'right_index': None,
                'x_at_max': pulse_time, 'temperature_c': None, 'time_min': pulse_time,
                'height': 0.0, 'prominence': 0.0, 'area_signal_min': 0.0,
                'left_x': None, 'right_x': None, 'fully_adsorbed': True,
                'fully_adsorbed_source': 'Expected injection without a detected response',
            })
    reported_rows = experiment.get('reported_pulses') or []
    for position, peak in enumerate(detected_peaks, start=1):
        peak['number'] = missing_count + position
        reported = reported_rows[position - 1] if position <= len(reported_rows) else {}
        peak['reported_uptake_umol_g'] = reported.get('reported_uptake_umol_g')
        peak['reported_cumulative_uptake_umol_g'] = reported.get('reported_cumulative_uptake_umol_g')
        reported_full = False
        if (sample_mass_g and reported.get('reported_uptake_umol_g') is not None
                and dose > 0):
            reported_full = reported['reported_uptake_umol_g'] >= 0.98 * dose / sample_mass_g
        peak['fully_adsorbed'] = reported_full
        peak['fully_adsorbed_source'] = ('Instrument reported uptake equals the full dose'
                                         if reported_full else '')
    all_pulses = synthetic + detected_peaks
    cumulative = 0.0
    for peak in all_pulses:
        uptake = None
        if peak.get('fully_adsorbed') and peak.get('index') is None:
            uptake = dose
        elif response_factor is not None and reference_area is not None:
            difference = (peak['area_signal_min'] - reference_area
                          if direction == 'inverse' else reference_area - peak['area_signal_min'])
            uptake = max(0.0, difference * response_factor)
        if uptake is not None:
            cumulative += uptake
        peak['pulse_dose_umol'] = dose
        peak['uptake_umol'] = uptake
        peak['cumulative_uptake_umol'] = cumulative if uptake is not None else None
        peak['cumulative_uptake_umol_g'] = cumulative / sample_mass_g if uptake is not None and sample_mass_g else None
        peak['saturated_reference'] = peak in saturated
    peaks[:] = all_pulses
    quantified = response_factor is not None or bool(missing_count)
    uptake_umol_g = cumulative / sample_mass_g if quantified and sample_mass_g else None
    reported_total = next((row.get('reported_cumulative_uptake_umol_g')
                           for row in reversed(reported_rows)
                           if row.get('reported_cumulative_uptake_umol_g') is not None), None)
    return {
        'pulse_dose_umol': dose, 'loop_volume_ml': loop_volume,
        'active_gas_percent': concentration, 'loop_pressure_atm': pressure,
        'loop_temperature_c': loop_temp, 'saturated_peak_count': len(saturated),
        'saturated_reference_area_signal_min': reference_area,
        'response_factor_umol_per_signal_min': response_factor,
        'response_factor_source': 'Manual input' if manual_factor else ('Saturated pulse peaks' if response_factor else 'Unavailable'),
        'chemisorption_direction': direction,
        'expected_injection_count': expected_injections,
        'injection_count_source': injection_source,
        'detected_pulse_count': len(detected_peaks),
        'fully_adsorbed_pulse_count': sum(bool(peak.get('fully_adsorbed')) for peak in all_pulses),
        'reported_total_uptake_umol_g': reported_total,
        'total_uptake_umol': cumulative if quantified else None,
        'total_uptake_umol_g': uptake_umol_g,
        'metal_metrics': _metal_metrics(uptake_umol_g, params),
    }


def analyze_autochem(parsed, params):
    sample_mass_g = _to_float(params.get('sample_mass_g')) or parsed['metadata'].get('sample_mass_g')
    prepared = []
    response_by_gas = {}
    for experiment in parsed['experiments']:
        trace = _prepare_trace(experiment, params)
        peaks, prominence = _detect_peaks(trace, params)
        candidate_count = len(peaks)
        peaks = _limit_tpd_peaks(experiment['type'], peaks, params)
        prepared.append({'experiment': experiment, 'trace': trace, 'peaks': peaks,
                         'peak_candidate_count': candidate_count,
                         'prominence_threshold': prominence})
    for result in prepared:
        experiment = result['experiment']
        if experiment['type'] != 'Chemisorption':
            continue
        chem = _analyze_chemisorption(experiment, result['trace'], result['peaks'],
                                      params, parsed['metadata'], sample_mass_g)
        result.update(chem)
        gas = str(params.get('active_gas') or '').strip()
        if not gas or gas.lower() == 'auto':
            gas = experiment.get('gas') or 'Unknown'
        result['calibration_gas'] = gas
        if chem['response_factor_umol_per_signal_min']:
            response_by_gas[gas] = chem['response_factor_umol_per_signal_min']

    manual_factor = _to_float(params.get('manual_response_factor_umol_per_area'))
    for result in prepared:
        experiment, trace, peaks = result['experiment'], result['trace'], result['peaks']
        positive = np.clip(trace['corrected_signal'], 0, None)
        total_area = float(np.trapezoid(positive, trace['time_min']))
        result['total_area_signal_min'] = total_area
        if experiment['type'] == 'Chemisorption':
            continue
        gas = experiment.get('gas') or str(params.get('active_gas') or '')
        factor = manual_factor or response_by_gas.get(gas)
        result['response_factor_umol_per_signal_min'] = factor
        result['response_factor_source'] = 'Manual input' if manual_factor else ('Matching saturated pulse peaks' if factor else 'Unavailable')
        result['total_amount_umol'] = total_area * factor if factor else None
        result['total_amount_umol_g'] = total_area * factor / sample_mass_g if factor and sample_mass_g else None
        for peak in peaks:
            peak['amount_umol'] = peak['area_signal_min'] * factor if factor else None
            peak['amount_umol_g'] = peak['amount_umol'] / sample_mass_g if factor and sample_mass_g else None
    return prepared, sample_mass_g, response_by_gas


def default_plot_settings(experiment, sample_id='Sample'):
    experiment_type = experiment['type']
    x_axis_basis = 'time' if experiment_type == 'Chemisorption' else 'temperature'
    x_label = 'Time (min)' if x_axis_basis == 'time' else 'Temperature (°C)'
    return {
        'title': f"{sample_id} — {experiment.get('display_name') or experiment['name']}",
        'x_axis_basis': x_axis_basis, 'x_axis_label': x_label,
        'y_axis_label': 'Baseline-corrected detector signal (a.u.)',
        'tick_font_size': 11, 'axis_font_size': 13, 'title_font_size': 15,
        'legend_font_size': 10, 'line_width': 1.8,
        'max_peak_labels': 6 if experiment_type == 'TPD' else 4,
        'png_dpi': 300,
        'figure_width': 8.5, 'figure_height': 4.8,
        'x_axis_min': None, 'x_axis_max': None, 'y_axis_min': None, 'y_axis_max': None,
        'show_baseline': True,
        'show_peak_markers': experiment_type != 'Chemisorption',
        'show_peak_labels': experiment_type != 'Chemisorption',
        'show_y_tick_labels': True,
        'show_integration': True, 'show_grid': False,
        'signal_color': '#3282D2', 'baseline_color': '#7D8590',
        'peak_color': '#D64545', 'integration_color': '#8BC6EC',
    }


def _hex_color(value, default):
    text = str(value or '')
    if re.fullmatch(r'#[0-9A-Fa-f]{6}', text):
        return text.upper()
    if re.fullmatch(r'#[0-9A-Fa-f]{3}', text):
        return '#' + ''.join(char * 2 for char in text[1:]).upper()
    return default


def normalize_plot_settings(settings, experiment, sample_id='Sample'):
    defaults = default_plot_settings(experiment, sample_id)
    settings = settings or {}
    normalized = dict(defaults)
    basis = str(settings.get('x_axis_basis', defaults['x_axis_basis'])).strip().lower()
    normalized['x_axis_basis'] = basis if basis in {'temperature', 'time'} else defaults['x_axis_basis']
    for key in ('title', 'x_axis_label', 'y_axis_label'):
        if key in settings:
            normalized[key] = str(settings[key])
    automatic_labels = {'Time (min)', 'Temperature (°C)', defaults['x_axis_label']}
    if 'x_axis_label' not in settings or str(settings.get('x_axis_label')) in automatic_labels:
        normalized['x_axis_label'] = ('Time (min)' if normalized['x_axis_basis'] == 'time'
                                      else 'Temperature (°C)')
    for key, low, high in (
        ('tick_font_size', 8, 30), ('axis_font_size', 9, 36),
        ('title_font_size', 10, 42), ('legend_font_size', 8, 28),
        ('line_width', 0.5, 6), ('max_peak_labels', 0, 50), ('png_dpi', 72, 600),
        ('figure_width', 5, 18), ('figure_height', 3, 12)):
        value = _to_float(settings.get(key))
        if value is not None:
            normalized[key] = max(low, min(high, value))
    if experiment['type'] == 'TPD':
        normalized['max_peak_labels'] = min(6, normalized['max_peak_labels'])
    for key in ('x_axis_min', 'x_axis_max', 'y_axis_min', 'y_axis_max'):
        if key in settings:
            normalized[key] = _to_float(settings.get(key))
    for key in ('show_baseline', 'show_peak_markers', 'show_peak_labels', 'show_y_tick_labels',
                'show_integration', 'show_grid'):
        value = settings.get(key, defaults[key])
        normalized[key] = value if isinstance(value, bool) else str(value).lower() in {'1', 'true', 'yes', 'on'}
    for key in ('signal_color', 'baseline_color', 'peak_color', 'integration_color'):
        normalized[key] = _hex_color(settings.get(key), defaults[key])
    for label, low_key, high_key in (
            ('X-axis', 'x_axis_min', 'x_axis_max'), ('Y-axis', 'y_axis_min', 'y_axis_max')):
        low, high = normalized[low_key], normalized[high_key]
        if low is not None and high is not None and high <= low:
            raise ValueError(f'{label} maximum must be greater than its minimum.')
    return normalized


def make_experiment_plot(result, output_dir, metadata, settings=None):
    from modules.characterization_plot import render_program_plot

    experiment, trace = result['experiment'], result['trace']
    settings = normalize_plot_settings(settings, experiment, metadata.get('sample_id', 'Sample'))
    if settings['x_axis_basis'] == 'temperature' and np.isfinite(trace['temperature_c']).any():
        x = trace['temperature_c']
    else:
        x = trace['time_min']
        if settings['x_axis_basis'] == 'temperature':
            settings['x_axis_basis'] = 'time'
            settings['x_axis_label'] = 'Time (min)'
    corrected = trace['corrected_signal']
    os.makedirs(output_dir, exist_ok=True)
    prefix = _safe_token(metadata.get('output_prefix') or metadata.get('sample_id'), 'TP')
    path = os.path.join(output_dir, f"{prefix}_{_safe_token(experiment['key'])}_plot.png")
    render_program_plot(path, x, corrected, result['peaks'], settings)
    return path, settings


def _style_header(row):
    from openpyxl.styles import Alignment, Font, PatternFill
    for cell in row:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=_HEADER_FILL)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _autosize(worksheet, max_width=42):
    from openpyxl.utils import get_column_letter
    for column_index, column in enumerate(worksheet.columns, start=1):
        letter = get_column_letter(column_index)
        width = max((len(str(cell.value)) for cell in column[:500] if cell.value is not None), default=8)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def _unique_sheet_name(workbook, requested):
    base = re.sub(r'[\\/*?:\[\]]', '-', requested)[:31] or 'Experiment'
    name, counter = base, 2
    while name in workbook.sheetnames:
        suffix = f' {counter}'
        name = base[:31 - len(suffix)] + suffix
        counter += 1
    return name


def write_workbook(parsed, analyses, plot_paths, output_dir, metadata, params,
                   plot_settings_by_key=None):
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Alignment, Font, PatternFill

    os.makedirs(output_dir, exist_ok=True)
    prefix = _safe_token(metadata.get('output_prefix') or metadata.get('sample_id'), 'TP')
    path = os.path.join(output_dir, f'{prefix}_temperature_programmed_analysis.xlsx')
    workbook = Workbook()
    summary = workbook.active
    summary.title = 'Summary'
    summary.sheet_view.showGridLines = False
    summary['A1'] = 'TPD / TPO / TPR / Chemisorption Analysis'
    summary['A1'].font = Font(size=17, bold=True, color='FFFFFF')
    summary['A1'].fill = PatternFill('solid', fgColor=_HEADER_FILL)
    summary.merge_cells('A1:F1')
    for values in (
        ('Sample ID', metadata.get('sample_id')), ('Source file', parsed['metadata'].get('source_file')),
        ('Sample mass (g)', metadata.get('sample_mass_g')), ('Loop pressure (atm)', params.get('loop_pressure_atm')),
        ('Processed', datetime.now().isoformat(timespec='seconds')), ()):
        summary.append(values)
    summary.append(('Experiment', 'Type', 'Gas', 'Events / peaks', 'Integrated area (signal·min)', 'Total amount (µmol/g)'))
    _style_header(summary[8])
    for result in analyses:
        experiment = result['experiment']
        amount = (result.get('total_uptake_umol_g') if experiment['type'] == 'Chemisorption'
                  else result.get('total_amount_umol_g'))
        summary.append((experiment['name'], experiment['type'], experiment.get('gas'),
                        len(result['peaks']), result.get('total_area_signal_min'), amount))
    summary.append(())
    summary.append(('Calibration / assumption', 'Value', 'Notes'))
    _style_header(summary[summary.max_row])
    assumption_rows = [
        ('Active loop volume (mL)', params.get('loop_volume_ml'), 'Default 0.51548 mL; user-editable in the toolkit'),
        ('Active gas concentration (%)', params.get('active_gas_percent'), 'Volume or mole percent'),
        ('Loop pressure (atm)', params.get('loop_pressure_atm'), '1 atm fallback as requested'),
        ('Loop temperature (°C)', params.get('loop_temperature_c'), 'Parsed from the report when available'),
        ('Reported active-loop dose (µmol)', parsed['metadata'].get('reported_active_loop_umol'), 'Instrument-reported comparison value'),
        ('Baseline method', params.get('baseline_method'), 'Analysis setting'),
        ('Smoothing window (points)', params.get('smoothing_window'), 'Centered moving-average window'),
        ('Peak prominence (%)', params.get('prominence_percent'), 'Relative to corrected signal range'),
        ('Maximum TPD peaks', params.get('maximum_tpd_peaks'),
         'At most the most prominent six candidates are retained'),
        ('Expected chemisorption injections', params.get('expected_injection_count'),
         'Parsed from an instrument pulse table when available; user value overrides upward'),
        ('Ramp rate (°C/min)', params.get('ramp_rate_c_min'),
         'Used only when temperature cannot be parsed from the report'),
        ('Ramp start temperature (°C)', params.get('ramp_start_temperature_c'),
         'Fallback for deriving temperature from time and ramp rate'),
    ]
    for values in assumption_rows:
        summary.append(values)
    summary.append(())
    summary.append(('Chemisorption metric', 'Value', 'Units / status'))
    _style_header(summary[summary.max_row])
    chem = next((item for item in analyses if item['experiment']['type'] == 'Chemisorption'), None)
    if chem:
        metal = chem.get('metal_metrics', {})
        rows = [
            ('Pulse dose', chem.get('pulse_dose_umol'), 'µmol/pulse'),
            ('Response factor', chem.get('response_factor_umol_per_signal_min'), 'µmol/(signal·min)'),
            ('Response-factor source', chem.get('response_factor_source'), ''),
            ('Chemisorption direction', chem.get('chemisorption_direction'), ''),
            ('Expected injections', chem.get('expected_injection_count'), chem.get('injection_count_source')),
            ('Detected pulse responses', chem.get('detected_pulse_count'), ''),
            ('Fully adsorbed pulses', chem.get('fully_adsorbed_pulse_count'), ''),
            ('Total uptake', chem.get('total_uptake_umol_g'), 'µmol/g sample'),
            ('Total uptake (reported)', chem.get('reported_total_uptake_umol_g'), 'µmol/g sample'),
            ('Metal dispersion', metal.get('dispersion_percent'), '%'),
            ('Metal surface area', metal.get('metal_surface_area_m2_g_sample'), 'm²/g sample'),
            ('Metal surface area', metal.get('metal_surface_area_m2_g_metal'), 'm²/g metal'),
            ('Particle diameter', metal.get('particle_diameter_nm'), 'nm'),
            ('Cubic crystallite size', metal.get('cubic_crystallite_size_nm'), 'nm'),
        ]
        for values in rows:
            summary.append(values)
    else:
        summary.append(('Status', None, 'No pulse-chemisorption experiment was found.'))
    _autosize(summary, 60)

    settings_sheet = workbook.create_sheet('Settings')
    settings_sheet.append(('Scope', 'Parameter', 'Value', 'Purpose'))
    _style_header(settings_sheet[1])
    analysis_purposes = {
        'loop_volume_ml': 'Active sample-loop volume',
        'active_gas_percent': 'Active gas concentration in the loop',
        'loop_pressure_atm': 'Loop pressure; defaults to 1 atm',
        'loop_temperature_c': 'Parsed loop temperature or user fallback',
        'saturated_peak_count': 'Final saturated pulses used for response factor',
        'expected_injection_count': 'Expected pulse injections; parsed from the sample report when available',
        'manual_response_factor': 'Optional user response-factor override',
        'chemisorption_direction': 'Normal or inverse pulse uptake interpretation',
        'baseline_method': 'Baseline correction method',
        'smoothing_window': 'Centered moving-average window in points',
        'prominence_percent': 'Automatic peak prominence threshold',
        'minimum_peak_distance_c': 'Minimum programmed-peak spacing',
        'maximum_tpd_peaks': 'Maximum prominent TPD peaks retained; capped at six',
        'minimum_pulse_distance_min': 'Minimum pulse-peak spacing',
        'ramp_rate_c_min': 'Fallback heating or cooling rate when temperature is unavailable',
        'ramp_start_temperature_c': 'Fallback temperature at the start of the trace',
        'metal': 'Transition-metal preset',
        'metal_loading_wt_percent': 'Metal loading used for dispersion estimates',
        'stoichiometry_metal_per_gas': 'Metal atoms represented by one adsorbed gas molecule',
    }
    for key, purpose in analysis_purposes.items():
        settings_sheet.append(('Analysis', key, params.get(key), purpose))
    for experiment_key, plot_settings in sorted((plot_settings_by_key or {}).items()):
        for key, value in sorted(plot_settings.items()):
            settings_sheet.append((experiment_key, f'plot.{key}', value,
                                   'Setting used for the embedded experiment plot'))
    for cell in settings_sheet['C'][1:]:
        cell.fill = PatternFill('solid', fgColor=_INPUT_FILL)
    settings_sheet.freeze_panes = 'A2'
    _autosize(settings_sheet, 62)

    presets = workbook.create_sheet('Metal Presets')
    presets.append(('Metal', 'Name', 'Atomic weight (g/mol)', 'Density (g/cm³)',
                    'Cross-section (nm²/atom)', 'Default metal atoms/gas molecule'))
    _style_header(presets[1])
    for symbol, values in METAL_PRESETS.items():
        presets.append((symbol, values['name'], values['atomic_weight_g_mol'],
                        values['density_g_cm3'], values['cross_section_nm2'],
                        values['stoichiometry_metal_per_gas']))
    presets.append(())
    presets.append(('Note', 'Preset values are editable estimates; verify gas-specific adsorption stoichiometry for the system under study.'))
    _autosize(presets, 72)

    for result in analyses:
        experiment, trace = result['experiment'], result['trace']
        sheet = workbook.create_sheet(_unique_sheet_name(workbook, f"{experiment['type']} - {experiment['name']}"))
        sheet.sheet_view.showGridLines = False
        sheet.append((f"{experiment['type']} Analysis", 'Value', 'Units / provenance'))
        _style_header(sheet[1])
        amount_label = 'Total uptake' if experiment['type'] == 'Chemisorption' else 'Total amount'
        amount_value = result.get('total_uptake_umol_g') if experiment['type'] == 'Chemisorption' else result.get('total_amount_umol_g')
        overview = [
            ('Experiment name', experiment['name'], ''), ('Analysis type', experiment['analysis_type'], ''),
            ('Gas', experiment.get('gas'), ''), ('Calibration', experiment.get('calibration'), ''),
            ('Signal orientation', trace['orientation'], ''),
            ('Peak count', len(result['peaks']), ''),
            ('Integrated area', result.get('total_area_signal_min'), 'signal·min'),
            (amount_label, amount_value, 'µmol/g sample'),
            ('Response factor', result.get('response_factor_umol_per_signal_min'), 'µmol/(signal·min)'),
            ('Response-factor source', result.get('response_factor_source'), ''),
            ('Ramp rate', trace.get('ramp_rate_c_min'), '°C/min; parsed or derived'),
        ]
        if experiment['type'] == 'TPD':
            overview.extend([
                ('Peak candidates before cap', result.get('peak_candidate_count'), ''),
                ('Maximum TPD peaks', params.get('maximum_tpd_peaks'),
                 'Most prominent candidates retained'),
            ])
        if experiment['type'] == 'Chemisorption':
            overview.extend([
                ('Expected injections', result.get('expected_injection_count'), result.get('injection_count_source')),
                ('Detected pulse responses', result.get('detected_pulse_count'), ''),
                ('Fully adsorbed pulses', result.get('fully_adsorbed_pulse_count'), ''),
                ('Total uptake (reported)', result.get('reported_total_uptake_umol_g'), 'µmol/g sample'),
            ])
        for values in overview:
            sheet.append(values)
        sheet.append(())
        peak_header_row = sheet.max_row + 1
        if experiment['type'] == 'Chemisorption':
            sheet.append(('Injection', 'Time (min)', 'Peak height', 'Area (signal·min)',
                          'Pulse dose (µmol)', 'Calculated uptake (µmol)',
                          'Calculated cumulative uptake (µmol/g)', 'Reported uptake (µmol/g)',
                          'Fully adsorbed', 'Status / reference'))
        else:
            sheet.append(('Peak', 'Temperature (°C)', 'Time (min)', 'Peak height', 'Area (signal·min)',
                          'Amount (µmol)', 'Amount (µmol/g)', 'Left bound', 'Right bound'))
        _style_header(sheet[peak_header_row])
        for peak in result['peaks']:
            if experiment['type'] == 'Chemisorption':
                status = (peak.get('fully_adsorbed_source') or
                          ('Saturated response reference' if peak.get('saturated_reference') else ''))
                sheet.append((peak['number'], peak['time_min'], peak['height'], peak['area_signal_min'],
                              peak.get('pulse_dose_umol'), peak.get('uptake_umol'),
                              peak.get('cumulative_uptake_umol_g'), peak.get('reported_uptake_umol_g'),
                              peak.get('fully_adsorbed'), status))
            else:
                sheet.append((peak['number'], peak['temperature_c'], peak['time_min'], peak['height'],
                              peak['area_signal_min'], peak.get('amount_umol'), peak.get('amount_umol_g'),
                              peak['left_x'], peak['right_x']))
        raw_header_row = sheet.max_row + 3
        sheet.cell(raw_header_row, 1, 'Point')
        sheet.cell(raw_header_row, 2, 'Time (min)')
        sheet.cell(raw_header_row, 3, 'Temperature (°C)')
        sheet.cell(raw_header_row, 4, 'Raw oriented signal (a.u.)')
        sheet.cell(raw_header_row, 5, 'Smoothed signal (a.u.)')
        sheet.cell(raw_header_row, 6, 'Baseline (a.u.)')
        sheet.cell(raw_header_row, 7, 'Corrected signal (a.u.)')
        _style_header(sheet[raw_header_row])
        for index in range(len(trace['time_min'])):
            sheet.append((index + 1, float(trace['time_min'][index]),
                          float(trace['temperature_c'][index]) if np.isfinite(trace['temperature_c'][index]) else None,
                          float(trace['raw_signal'][index]), float(trace['smoothed_signal'][index]),
                          float(trace['baseline'][index]), float(trace['corrected_signal'][index])))
        sheet.freeze_panes = f'A{raw_header_row + 1}'
        sheet.auto_filter.ref = f'A{raw_header_row}:G{sheet.max_row}'
        for header_row in (1, peak_header_row, raw_header_row):
            sheet.row_dimensions[header_row].height = 31
        for column, width in {
                'A': 25, 'B': 34, 'C': 26, 'D': 24, 'E': 23,
                'F': 20, 'G': 23, 'H': 22, 'I': 22, 'J': 34}.items():
            sheet.column_dimensions[column].width = width
        plot_path = plot_paths.get(experiment['key'])
        if plot_path and os.path.isfile(plot_path):
            image = Image(plot_path)
            image.width, image.height = 760, 420
            sheet.add_image(image, 'J2')
    workbook.save(path)
    return path


def _ui_experiment(result, plot_path, settings):
    experiment = result['experiment']
    peaks = []
    for peak in result['peaks']:
        peaks.append({
            'number': peak['number'], 'temperature_c': peak.get('temperature_c'),
            'time_min': peak.get('time_min'), 'area_signal_min': peak.get('area_signal_min'),
            'amount_umol_g': peak.get('amount_umol_g'), 'uptake_umol': peak.get('uptake_umol'),
            'cumulative_uptake_umol_g': peak.get('cumulative_uptake_umol_g'),
            'saturated_reference': peak.get('saturated_reference', False),
            'fully_adsorbed': peak.get('fully_adsorbed', False),
            'fully_adsorbed_source': peak.get('fully_adsorbed_source', ''),
            'reported_uptake_umol_g': peak.get('reported_uptake_umol_g'),
        })
    metrics = {
        'peak_count': len(peaks), 'total_area_signal_min': result.get('total_area_signal_min'),
        'peak_candidate_count': result.get('peak_candidate_count'),
        'response_factor_umol_per_signal_min': result.get('response_factor_umol_per_signal_min'),
        'response_factor_source': result.get('response_factor_source'),
    }
    if experiment['type'] == 'Chemisorption':
        metrics.update({
            'pulse_dose_umol': result.get('pulse_dose_umol'),
            'total_uptake_umol_g': result.get('total_uptake_umol_g'),
            'reported_total_uptake_umol_g': result.get('reported_total_uptake_umol_g'),
            'chemisorption_direction': result.get('chemisorption_direction'),
            'expected_injection_count': result.get('expected_injection_count'),
            'injection_count_source': result.get('injection_count_source'),
            'detected_pulse_count': result.get('detected_pulse_count'),
            'fully_adsorbed_pulse_count': result.get('fully_adsorbed_pulse_count'),
            **result.get('metal_metrics', {}),
        })
    else:
        metrics['total_amount_umol_g'] = result.get('total_amount_umol_g')
    return {
        'key': experiment['key'], 'number': experiment['number'], 'name': experiment['name'],
        'display_name': experiment.get('display_name') or experiment['name'],
        'type': experiment['type'], 'gas': experiment.get('gas'), 'calibration': experiment.get('calibration'),
        'metrics': metrics, 'peaks': peaks, 'plot_path': plot_path, 'plot_settings': settings,
        'ramp_rate_c_min': result['trace'].get('ramp_rate_c_min'),
    }


def run(filepath, output_dir, metadata, params, plot_context=None):
    parsed = parse_autochem_txt(filepath)
    metadata = dict(metadata)
    metadata['sample_id'] = str(metadata.get('sample_id') or parsed['metadata'].get('sample_id') or 'Sample').strip()
    params = dict(params)
    if not params.get('sample_mass_g'):
        params['sample_mass_g'] = parsed['metadata'].get('sample_mass_g')
    params['loop_volume_ml'] = _to_float(params.get('loop_volume_ml')) or 0.51548
    params['active_gas_percent'] = _to_float(params.get('active_gas_percent')) or 10.0
    params['loop_pressure_atm'] = _to_float(params.get('loop_pressure_atm')) or 1.0
    params['loop_temperature_c'] = (_to_float(params.get('loop_temperature_c'))
                                    or parsed['metadata'].get('reported_loop_temperature_c') or 25.0)
    params['baseline_method'] = str(params.get('baseline_method') or 'auto')
    params['smoothing_window'] = int(_to_float(params.get('smoothing_window')) or 301)
    params['prominence_percent'] = _to_float(params.get('prominence_percent')) or 8.0
    params['maximum_tpd_peaks'] = max(
        1, min(6, int(_to_float(params.get('maximum_tpd_peaks')) or 6)))
    analyses, sample_mass_g, response_by_gas = analyze_autochem(parsed, params)
    metadata['sample_mass_g'] = sample_mass_g
    plot_paths, settings_by_key = {}, {}
    for result in analyses:
        path, settings = make_experiment_plot(result, output_dir, metadata)
        key = result['experiment']['key']
        plot_paths[key], settings_by_key[key] = path, settings
    workbook_path = write_workbook(parsed, analyses, plot_paths, output_dir, metadata,
                                   params, settings_by_key)
    if plot_context is not None:
        plot_context.update({'parsed': parsed, 'analyses': analyses, 'metadata': metadata,
                             'params': params, 'output_dir': output_dir,
                             'plot_paths': plot_paths, 'settings_by_key': settings_by_key,
                             'workbook_path': workbook_path})
    return {
        'sample_id': metadata['sample_id'], 'sample_mass_g': sample_mass_g,
        'reported_sample_mass_g': parsed['metadata'].get('sample_mass_g'),
        'reported_active_loop_umol': parsed['metadata'].get('reported_active_loop_umol'),
        'response_factors': response_by_gas,
        'experiments': [_ui_experiment(result, plot_paths[result['experiment']['key']],
                                       settings_by_key[result['experiment']['key']]) for result in analyses],
        'summary_path': workbook_path, 'output_dir': output_dir,
    }


def regenerate_plot(context, experiment_key, settings):
    result = next((item for item in context['analyses']
                   if item['experiment']['key'] == experiment_key), None)
    if result is None:
        raise ValueError('The selected experiment is not available in this plotting session.')
    plot_path, normalized = make_experiment_plot(result, context['output_dir'],
                                                 context['metadata'], settings)
    context['plot_paths'][experiment_key] = plot_path
    context['settings_by_key'][experiment_key] = normalized
    workbook_path = write_workbook(context['parsed'], context['analyses'],
                                   context['plot_paths'], context['output_dir'],
                                   context['metadata'], context['params'],
                                   context['settings_by_key'])
    context['workbook_path'] = workbook_path
    return {'plot_path': plot_path, 'summary_path': workbook_path,
            'plot_settings': normalized}
