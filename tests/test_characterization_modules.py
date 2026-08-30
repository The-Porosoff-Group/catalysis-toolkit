import math
import os
import tempfile
import unittest

import numpy as np
from openpyxl import load_workbook

from modules import bet_processor
from modules import tga_processor


class BetProcessorTests(unittest.TestCase):
    def test_recommended_window_recovers_physical_bet_parameters(self):
        pressure = np.linspace(0.01, 0.32, 28)
        qm = 2.35
        c_value = 55.0
        quantity = qm * c_value * pressure / (
            (1.0 - pressure) * (1.0 + (c_value - 1.0) * pressure))
        parsed = {
            'points': [
                {'branch': 'Adsorption', 'relative_pressure': float(x),
                 'quantity_cm3_g_stp': float(q), 'source_sheet': 'Synthetic',
                 'source_row': index + 2, 'absolute_pressure_mmhg': None,
                 'elapsed_time': ''}
                for index, (x, q) in enumerate(zip(pressure, quantity))
            ],
            'source_metrics': {},
        }

        result = bet_processor.analyze_bet(parsed)

        self.assertAlmostEqual(result['c_constant'], c_value, places=6)
        self.assertAlmostEqual(result['monolayer_capacity_cm3_g_stp'], qm, places=6)
        self.assertGreater(result['r_squared'], 0.999999)
        self.assertFalse(result['flags'])
        expected_volume = quantity[-1] * 34.65 / 22414.0
        self.assertAlmostEqual(result['total_pore_volume_cm3_g'], expected_volume, places=10)
        self.assertAlmostEqual(
            result['average_pore_diameter_nm'],
            4000.0 * expected_volume / result['surface_area_m2_g'], places=8)

    def test_manual_unphysical_window_is_flagged(self):
        pressure = np.linspace(0.01, 0.35, 20)
        quantity = np.linspace(1.0, 3.0, 20)
        parsed = {
            'points': [
                {'branch': 'Adsorption', 'relative_pressure': float(x),
                 'quantity_cm3_g_stp': float(q)}
                for x, q in zip(pressure, quantity)
            ],
            'source_metrics': {},
        }
        result = bet_processor.analyze_bet(parsed, p_min=0.01, p_max=0.35)
        self.assertTrue(result['flags'])


class TemperatureProgrammedProcessorTests(unittest.TestCase):
    def test_clean_labels_and_plot_defaults_follow_experiment_type(self):
        chem = {'type': 'Chemisorption', 'name': 'CO2 _pulsed Chemi-sorption'}
        tpd = {'type': 'TPD', 'name': 'CO2_TPD'}
        self.assertEqual(
            tga_processor._display_experiment_name(chem['name'], chem['type'], 'CO2'),
            'CO2 Pulsed Chemisorption')
        self.assertFalse(tga_processor.default_plot_settings(chem)['show_peak_labels'])
        settings = tga_processor.normalize_plot_settings(
            {'x_axis_basis': 'time', 'x_axis_label': 'Temperature (°C)',
             'show_y_tick_labels': False, 'max_peak_labels': 20}, tpd)
        self.assertEqual(settings['x_axis_label'], 'Time (min)')
        self.assertFalse(settings['show_y_tick_labels'])
        self.assertEqual(settings['max_peak_labels'], 6)

    def test_tpd_peak_limit_retains_most_prominent_in_temperature_order(self):
        peaks = [
            {'number': index + 1, 'temperature_c': temperature, 'time_min': temperature / 10,
             'prominence': prominence}
            for index, (temperature, prominence) in enumerate(
                [(100, 1), (200, 8), (300, 3), (400, 7),
                 (500, 2), (600, 6), (700, 4), (800, 5)])
        ]

        retained = tga_processor._limit_tpd_peaks('TPD', peaks, {'maximum_tpd_peaks': 4})

        self.assertEqual([peak['temperature_c'] for peak in retained], [200, 400, 600, 800])
        self.assertEqual([peak['number'] for peak in retained], [1, 2, 3, 4])
        self.assertEqual(len(tga_processor._limit_tpd_peaks('TPD', peaks, {'maximum_tpd_peaks': 20})), 6)

    @staticmethod
    def _synthetic_report():
        pulse_time = np.linspace(0, 6, 601)
        pulse_signal = np.full_like(pulse_time, 0.02)
        for center, amplitude in zip((1, 2, 3, 4, 5), (0.75, 0.88, 1.0, 1.0, 1.0)):
            pulse_signal += amplitude * np.exp(-0.5 * ((pulse_time - center) / 0.055) ** 2)
        pulse_temperature = np.full_like(pulse_time, 40.0)

        program_time = np.linspace(0, 10, 1001)
        program_temperature = 50.0 + 50.0 * program_time
        program_signal = 0.01 + 0.8 * np.exp(-0.5 * ((program_temperature - 220) / 22) ** 2)
        program_signal += 0.5 * np.exp(-0.5 * ((program_temperature - 410) / 28) ** 2)

        def table(label, x_label, x, y):
            rows = [label, f'{x_label}\tSignal (a.u.)']
            rows.extend(f'{a:.8g}\t{b:.8g}' for a, b in zip(x, y))
            return '\n'.join(rows)

        def temperature_table(label, time, temperature):
            rows = [label, 'Time (minutes)\tTemperature (°C)']
            rows.extend(f'{a:.8g}\t{b:.8g}' for a, b in zip(time, temperature))
            return '\n'.join(rows)

        return f'''MicroActive for AutoChem II 2920
Sample: Synthetic catalyst
Started: 1/1/2026 1:00 PM     Sample mass: 0.1000 g

Experiment 1: CO pulsed Chemi-sorption
Analysis type: Pulse Chemisorption
Calibration: None

Experiment 2: CO_TPD
Analysis type: Temperature Programmed Desorption
Calibration: None

Pulse Chemisorption Report
Experiment 1: CO pulsed Chemi-sorption
Number of Injections: 6
Peak Table
Peak Temperature Quantity Cumulative
----
1 40.0 0.10 0.10
2 40.0 0.08 0.18
3 40.0 0.03 0.21
4 40.0 0.00 0.21
5 40.0 0.00 0.21

Signal (a.u.) vs. Time

{table('Signal (a.u.) - CO pulsed Chemi-sorption', 'Time (minutes)', pulse_time, pulse_signal)}

{table('Signal (a.u.) - CO_TPD', 'Time (minutes)', program_time, program_signal)}

Temperature vs. Time

{temperature_table('Temperature - CO pulsed Chemi-sorption', pulse_time, pulse_temperature)}

{temperature_table('Temperature - CO_TPD', program_time, program_temperature)}
'''

    def _write_report(self, directory):
        path = os.path.join(directory, 'synthetic_autochem.txt')
        with open(path, 'wb') as handle:
            handle.write(self._synthetic_report().encode('utf-16'))
        return path

    def test_saturated_pulses_calibrate_matching_tpd(self):
        with tempfile.TemporaryDirectory() as directory:
            path = self._write_report(directory)
            parsed = tga_processor.parse_autochem_txt(path)
            analyses, mass, factors = tga_processor.analyze_autochem(parsed, {
                'sample_mass_g': 0.1,
                'loop_volume_ml': 0.51548,
                'active_gas_percent': 10,
                'loop_pressure_atm': 1,
                'loop_temperature_c': 25,
                'smoothing_window': 11,
                'prominence_percent': 5,
                'minimum_peak_distance_c': 40,
                'minimum_pulse_distance_min': 0.5,
                'baseline_method': 'linear',
            })

        chem = next(item for item in analyses if item['experiment']['type'] == 'Chemisorption')
        tpd = next(item for item in analyses if item['experiment']['type'] == 'TPD')
        self.assertEqual(mass, 0.1)
        self.assertIn('CO', factors)
        self.assertEqual(len(chem['peaks']), 6)
        self.assertEqual(chem['detected_pulse_count'], 5)
        self.assertEqual(chem['fully_adsorbed_pulse_count'], 1)
        self.assertTrue(chem['peaks'][0]['fully_adsorbed'])
        self.assertIsNone(chem['peaks'][0]['index'])
        self.assertIsNotNone(chem['total_uptake_umol_g'])
        self.assertGreater(chem['total_uptake_umol_g'], 0)
        self.assertEqual(len(tpd['peaks']), 2)
        self.assertIsNotNone(tpd['total_amount_umol_g'])

    def test_transition_metal_metrics_match_report_equations(self):
        metrics = tga_processor._metal_metrics(0.61131, {
            'metal': 'W', 'metal_loading_wt_percent': 15,
        })
        self.assertAlmostEqual(metrics['dispersion_percent'], 0.0749, delta=0.001)
        self.assertAlmostEqual(metrics['metal_surface_area_m2_g_sample'], 0.0273, delta=0.001)
        self.assertAlmostEqual(metrics['particle_diameter_nm'], 1710, delta=30)
        self.assertAlmostEqual(metrics['cubic_crystallite_size_nm'], 1425, delta=30)

    def test_standardized_workbook_contains_summary_and_experiment_tabs(self):
        with tempfile.TemporaryDirectory() as directory:
            source = self._write_report(directory)
            result = tga_processor.run(
                source, directory,
                {'sample_id': 'Synthetic', 'output_prefix': 'synthetic'},
                {
                    'sample_mass_g': 0.1, 'loop_volume_ml': 0.51548,
                    'active_gas_percent': 10, 'loop_pressure_atm': 1,
                    'loop_temperature_c': 25, 'saturated_peak_count': 3,
                    'smoothing_window': 11, 'prominence_percent': 5,
                    'minimum_peak_distance_c': 40,
                    'minimum_pulse_distance_min': 0.5,
                    'baseline_method': 'linear', 'metal': 'W',
                    'metal_loading_wt_percent': 15,
                })
            workbook = load_workbook(result['summary_path'], read_only=False, data_only=False)
            try:
                self.assertIn('Summary', workbook.sheetnames)
                self.assertIn('Metal Presets', workbook.sheetnames)
                self.assertTrue(any(name.startswith('Chemisorption') for name in workbook.sheetnames))
                self.assertTrue(any(name.startswith('TPD') for name in workbook.sheetnames))
                experiment_sheets = [sheet for sheet in workbook.worksheets
                                     if sheet.title.startswith(('Chemisorption', 'TPD'))]
                self.assertTrue(all(sheet._images for sheet in experiment_sheets))
            finally:
                workbook.close()


if __name__ == '__main__':
    unittest.main()
