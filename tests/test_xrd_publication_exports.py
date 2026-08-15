import copy
import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook
from PIL import Image

from modules.xrd import _write_summary_xlsx
from modules.xrd.presentation import (
    export_file_prefix,
    format_chemical_formula,
    format_space_group,
    phase_legend_label,
    phase_tick_label,
)
from modules.xrd.xrd_plots import make_xrd_plot


def _gaussian(x, center, width, amplitude):
    return amplitude * np.exp(-0.5 * ((x - center) / width) ** 2)


def publication_result():
    two_theta = np.linspace(20, 60, 801)
    background = 65 + 0.7 * (two_theta - 20)
    phase_one = (
        _gaussian(two_theta, 34.4, 0.24, 620)
        + _gaussian(two_theta, 38.1, 0.28, 310)
        + _gaussian(two_theta, 52.3, 0.31, 220)
    )
    phase_two = (
        _gaussian(two_theta, 31.6, 0.23, 370)
        + _gaussian(two_theta, 48.2, 0.30, 260)
        + _gaussian(two_theta, 57.4, 0.32, 180)
    )
    calculated = background + phase_one + phase_two
    observed = calculated + 16 * np.sin(two_theta * 1.7)
    return {
        'tt': two_theta.tolist(),
        'y_obs': observed.tolist(),
        'y_calc': calculated.tolist(),
        'y_background': background.tolist(),
        'phase_patterns': [phase_one.tolist(), phase_two.tolist()],
        'residuals': (observed - calculated).tolist(),
        'statistics': {'Rwp': 5.93, 'Rp': 4.81, 'chi2': 1.37, 'GoF': 1.17},
        'phase_results': [
            {
                'name': 'beta_Mo2C-Pbcn', 'formula': 'Mo2C',
                'spacegroup': 'Pbcn', 'spacegroup_number': 60,
                'system': 'orthorhombic', 'a': 4.73, 'b': 6.05, 'c': 5.21,
                'alpha': 90, 'beta': 90, 'gamma': 90,
                'weight_fraction_%': 62.4, 'weight_fraction_err_%': 1.2,
                'tick_positions': [34.4, 38.1, 52.3],
                'tick_reflections': [
                    {'two_theta': 34.4, 'hkl': [1, 1, 1]},
                    {'two_theta': 38.1, 'hkl': [2, 0, 0]},
                    {'two_theta': 52.3, 'hkl': [2, 2, 1]},
                ],
            },
            {
                'name': 'WC1-x_P-6m2', 'formula': 'WC1-x',
                'spacegroup': 'P-6m2', 'spacegroup_number': 187,
                'system': 'hexagonal', 'a': 2.91, 'b': 2.91, 'c': 2.84,
                'alpha': 90, 'beta': 90, 'gamma': 120,
                'weight_fraction_%': 37.6, 'weight_fraction_err_%': 1.2,
                'tick_positions': [31.6, 48.2, 57.4],
                'tick_reflections': [
                    {'two_theta': 31.6, 'hkl': [1, 0, 0]},
                    {'two_theta': 48.2, 'hkl': [1, 0, 1]},
                    {'two_theta': 57.4, 'hkl': [1, 1, 0]},
                ],
            },
        ],
        'zero_shift': 0.013,
        'wavelength': 1.54056,
    }


class XrdPublicationExportTests(unittest.TestCase):
    def test_scientific_labels_use_subscripts_and_overbars(self):
        self.assertEqual(format_chemical_formula('Mo2C'), 'Mo₂C')
        self.assertEqual(format_chemical_formula('WC1-x'), 'WC₁₋ₓ')
        self.assertNotIn('-', format_space_group('P-6m2'))
        label = phase_legend_label(publication_result()['phase_results'][1])
        self.assertIn('WC₁₋ₓ', label)
        self.assertIn('space group P6̅m2', label)
        self.assertIn('weight %', label)
        self.assertNotIn('_', label)
        self.assertEqual(
            phase_tick_label(publication_result()['phase_results'][1]),
            'WC₁₋ₓ (P6̅m2)')

    def test_export_prefix_contains_date_and_sample(self):
        self.assertEqual(
            export_file_prefix({
                'analysis_date': '2026-08-15T09:30:00',
                'sample_id': 'β-Mo2C_350C.xlsx',
            }),
            '2026-08-15_β-Mo2C_350C',
        )

    def test_light_and_dark_figures_are_compact_300_dpi_landscape(self):
        metadata = {
            'sample_id': 'β-Mo2C_350C',
            'wavelength_label': 'Cu Kα2 (1.54439 Å)',
            'method': 'GSAS-II',
        }
        with tempfile.TemporaryDirectory() as directory:
            for theme in ('light', 'dark'):
                path = os.path.join(directory, f'figure_{theme}.png')
                make_xrd_plot(
                    copy.deepcopy(publication_result()), metadata, path,
                    theme=theme)
                with Image.open(path) as image:
                    self.assertEqual(image.size, (1950, 1275))
                    self.assertGreater(image.width / image.height, 1.5)
                    self.assertAlmostEqual(image.info['dpi'][0], 300, delta=1)
                    corner = image.convert('RGB').getpixel((2, 2))
                    if theme == 'light':
                        self.assertGreater(min(corner), 240)
                    else:
                        self.assertLess(max(corner), 40)

    def test_light_is_the_default_export_theme(self):
        metadata = {'sample_id': 'Default_theme_sample'}
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, 'figure_default.png')
            make_xrd_plot(
                copy.deepcopy(publication_result()), metadata, path)
            with Image.open(path) as image:
                corner = image.convert('RGB').getpixel((2, 2))
                self.assertGreater(min(corner), 240)

    def test_web_preview_is_fixed_dark_and_export_choice_defaults_light(self):
        root = Path(__file__).resolve().parents[1]
        for relative_path in ('templates/index.html',
                              'templates/xrd_toolkit/index.html'):
            html = (root / relative_path).read_text(encoding='utf-8')
            self.assertIn('id="xrd-export-theme"', html)
            self.assertIn(
                '<option value="light" selected>Light background', html)
            self.assertNotIn('id="xrd-plot-theme"', html)
            self.assertIn("paper:'#0d1117'", html)
            self.assertIn('data.plot_paths?.dark', html)
            self.assertIn("modeBarButtonsToRemove:['toImage']", html)
            self.assertIn('id="xrd-figure-title"', html)
            self.assertIn("fd.append('figure_title'", html)
            self.assertIn('escHtml(figureTitle)', html)

    def test_custom_figure_title_is_used_without_changing_sample_identity(self):
        metadata = {
            'sample_id': 'traceable_sample_01',
            'figure_title': 'Tungsten carbide catalyst after reduction',
        }
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, 'custom_title.png')
            with patch('matplotlib.axes.Axes.set_title') as set_title:
                make_xrd_plot(
                    copy.deepcopy(publication_result()), metadata, path,
                    theme='light')
            self.assertEqual(
                set_title.call_args.args[0],
                'Tungsten carbide catalyst after reduction')

    def test_workbook_filename_and_content_include_sample_date_and_hkl(self):
        metadata = {
            'sample_id': 'β-Mo2C_350C',
            'figure_title': 'Carbide catalyst after reduction',
            'analysis_date': '2026-08-15',
            'source_file': 'M1.xlsx',
        }
        with tempfile.TemporaryDirectory() as directory:
            path = _write_summary_xlsx(
                copy.deepcopy(publication_result()), metadata,
                'GSAS-II', directory)
            self.assertEqual(
                os.path.basename(path),
                '2026-08-15_β-Mo2C_350C_xrd_refinement_results.xlsx')

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                summary_rows = list(
                    workbook['Summary'].iter_rows(values_only=True))
                summary_by_parameter = {
                    row[0]: row[1:] for row in summary_rows[1:]}
                self.assertIn(
                    '2026-08-15', summary_by_parameter['Analysis date'])
                self.assertIn(
                    'Carbide catalyst after reduction',
                    summary_by_parameter['Figure title'])
                self.assertIn(
                    'M1.xlsx', summary_by_parameter['Source data file'])

                plot_rows = list(
                    workbook['Plot Data'].iter_rows(values_only=True))
                headers = list(plot_rows[0])
                hkl_header = next(
                    header for header in headers
                    if header and 'Mo₂C (space group Pbcn) [hkl]' in header)
                hkl_index = headers.index(hkl_header)
                self.assertEqual(plot_rows[1][hkl_index], '(1 1 1)')
            finally:
                workbook.close()


if __name__ == '__main__':
    unittest.main()
